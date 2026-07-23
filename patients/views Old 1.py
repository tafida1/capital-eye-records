from django.contrib import messages
from django.contrib.auth.decorators import login_required 

from django.shortcuts import render, redirect, get_object_or_404

from accounts.decorators import records_staff_required, clinical_staff_required, finance_staff_required

from .forms import (
    PatientForm,
    FamilyGroupForm,
    PatientVisitForm,
    PatientVisitQuickForm,
    ConsultationForm,
    EyeExaminationForm,
    DiagnosisTreatmentForm,
    PrescriptionForm,
    SurgeryProcedureForm,
    SurgeryProcedureQuickForm,
    AppointmentForm,
    AppointmentQuickForm,
    BillForm,
    BillQuickForm,
    PaymentForm,
)

from .models import (
    Patient,
    FamilyGroup,
    PatientVisit,
    Consultation,
    EyeExamination,
    DiagnosisTreatment,
    Prescription,
    SurgeryProcedure,
    Appointment,
    Bill,
    Payment,
)

from django.template.loader import get_template
from audit_logs.utils import log_activity
from audit_logs.models import AuditLog
from django.http import HttpResponse
from xhtml2pdf import pisa
from io import BytesIO
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
from django.utils import timezone
from django.db.models import Q, Sum, Count




def render_to_pdf(template_src, context_dict, filename):
    template = get_template(template_src)
    html = template.render(context_dict)

    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)

    if pdf.err:
        return HttpResponse("PDF generation error", status=500)

    response = HttpResponse(result.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="{filename}"'
    return response


def autosize_excel_columns(sheet):
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            try:
                value_length = len(str(cell.value)) if cell.value else 0
                if value_length > max_length:
                    max_length = value_length
            except Exception:
                pass

        sheet.column_dimensions[column_letter].width = min(max_length + 3, 40)



@login_required
def patient_list(request):
    query = request.GET.get("q", "").strip()

    patients = Patient.objects.select_related("family_group", "registered_by").all()

    if query:
        patients = patients.filter(
            Q(file_number__icontains=query)
            | Q(full_name__icontains=query)
            | Q(phone_number__icontains=query)
            | Q(family_group__family_code__icontains=query)
            | Q(family_group__family_name__icontains=query)
            | Q(diagnosis__icontains=query)
            | Q(eye_complaint__icontains=query)
            | Q(surgery_procedure_details__icontains=query)
        )

    return render(request, "patients/patient_list.html", {
        "patients": patients,
        "query": query,
    })


@login_required
@records_staff_required
def patient_create(request):
    form = PatientForm(request.POST or None)

    if request.method == "POST":
        if form.is_valid():
            patient = form.save(commit=False)
            patient.registered_by = request.user

            if patient.family_group and not patient.family_group_name:
                patient.family_group_name = patient.family_group.family_name

            patient.save()

            log_activity(
                request,
                AuditLog.ActionType.CREATE,
                "Patients",
                f"Registered patient {patient.full_name} with file number {patient.file_number}.",
                object_id=patient.pk,
                object_repr=patient.file_number,
            )

            messages.success(
                request,
                f"Patient registered successfully. File Number: {patient.file_number}"
            )
            return redirect("patient_detail", pk=patient.pk)

        messages.error(request, "Please correct the errors below.")

    return render(request, "patients/patient_form.html", {
        "form": form,
        "title": "Register New Patient",
        "button_text": "Register Patient",
    })


@login_required
def patient_detail(request, pk):
    patient = get_object_or_404(
        Patient.objects.select_related("family_group", "registered_by"),
        pk=pk,
    )

    visits = patient.visits.select_related("created_by")[:8]
    appointments = patient.appointments.select_related("assigned_to", "created_by")[:5]
    surgeries = patient.surgeries.select_related("surgeon", "assistant", "created_by")[:5]
    bills = patient.bills.select_related("created_by")[:5]

    unpaid_bills = patient.bills.exclude(
        status=Bill.BillStatus.PAID
    ).exclude(
        status=Bill.BillStatus.CANCELLED
    )

    total_balance = sum(bill.balance for bill in unpaid_bills)

    context = {
        "patient": patient,
        "visits": visits,
        "appointments": appointments,
        "surgeries": surgeries,
        "bills": bills,
        "total_balance": total_balance,
        "unpaid_bills_count": unpaid_bills.count(),
    }

    return render(request, "patients/patient_detail.html", context)


@login_required
@records_staff_required
def patient_update(request, pk):
    patient = get_object_or_404(Patient, pk=pk)
    form = PatientForm(request.POST or None, instance=patient)

    if request.method == "POST":
        if form.is_valid():
            patient = form.save(commit=False)

            if patient.family_group and not patient.family_group_name:
                patient.family_group_name = patient.family_group.family_name

            patient.save()

            log_activity(
                request,
                AuditLog.ActionType.UPDATE,
                "Patients",
                f"Updated patient record {patient.full_name}.",
                object_id=patient.pk,
                object_repr=patient.file_number,
            )

            messages.success(request, "Patient record updated successfully.")
            return redirect("patient_detail", pk=patient.pk)

        messages.error(request, "Please correct the errors below.")

    return render(request, "patients/patient_form.html", {
        "form": form,
        "patient": patient,
        "title": "Update Patient Record",
        "button_text": "Save Changes",
    })


@login_required
def family_group_list(request):
    query = request.GET.get("q", "").strip()

    family_groups = FamilyGroup.objects.select_related("created_by").all()

    if query:
        family_groups = family_groups.filter(
            Q(family_code__icontains=query)
            | Q(family_name__icontains=query)
            | Q(head_of_family__icontains=query)
            | Q(primary_phone__icontains=query)
        )

    return render(request, "patients/family_group_list.html", {
        "family_groups": family_groups,
        "query": query,
    })


@login_required
@records_staff_required
def family_group_create(request):
    form = FamilyGroupForm(request.POST or None)

    if request.method == "POST":
        if form.is_valid():
            family_group = form.save(commit=False)
            family_group.created_by = request.user
            family_group.save()

            messages.success(
                request,
                f"Family group created successfully. Family Code: {family_group.family_code}"
            )
            return redirect("family_group_detail", pk=family_group.pk)

        messages.error(request, "Please correct the errors below.")

    return render(request, "patients/family_group_form.html", {
        "form": form,
        "title": "Create Family / Group Record",
        "button_text": "Create Family Group",
    })


@login_required
def family_group_detail(request, pk):
    family_group = get_object_or_404(FamilyGroup, pk=pk)
    members = family_group.members.all()

    return render(request, "patients/family_group_detail.html", {
        "family_group": family_group,
        "members": members,
    })


@login_required
@records_staff_required
def family_group_update(request, pk):
    family_group = get_object_or_404(FamilyGroup, pk=pk)
    form = FamilyGroupForm(request.POST or None, instance=family_group)

    if request.method == "POST":
        if form.is_valid():
            form.save()
            messages.success(request, "Family group updated successfully.")
            return redirect("family_group_detail", pk=family_group.pk)

        messages.error(request, "Please correct the errors below.")

    return render(request, "patients/family_group_form.html", {
        "form": form,
        "family_group": family_group,
        "title": "Update Family / Group Record",
        "button_text": "Save Changes",
    })



@login_required
def visit_list(request):
    query = request.GET.get("q", "").strip()
    status = request.GET.get("status", "").strip()

    visits = PatientVisit.objects.select_related("patient", "created_by").all()

    if query:
        visits = visits.filter(
            Q(visit_number__icontains=query)
            | Q(patient__file_number__icontains=query)
            | Q(patient__full_name__icontains=query)
            | Q(patient__phone_number__icontains=query)
            | Q(chief_complaint__icontains=query)
            | Q(brief_history__icontains=query)
        )

    if status:
        visits = visits.filter(status=status)

    return render(request, "patients/visit_list.html", {
        "visits": visits,
        "query": query,
        "status": status,
        "status_choices": PatientVisit.VisitStatus.choices,
    })


@login_required
@records_staff_required
def visit_create(request):
    form = PatientVisitForm(request.POST or None)

    if request.method == "POST":
        if form.is_valid():
            visit = form.save(commit=False)
            visit.created_by = request.user
            visit.save()

            messages.success(
                request,
                f"Patient visit created successfully. Visit Number: {visit.visit_number}"
            )
            return redirect("visit_detail", pk=visit.pk)

        messages.error(request, "Please correct the errors below.")

    return render(request, "patients/visit_form.html", {
        "form": form,
        "title": "Create Patient Visit / Encounter",
        "button_text": "Create Visit",
    })


@login_required
@records_staff_required
def patient_visit_create(request, patient_pk):
    patient = get_object_or_404(Patient, pk=patient_pk)
    form = PatientVisitQuickForm(request.POST or None)

    if request.method == "POST":
        if form.is_valid():
            visit = form.save(commit=False)
            visit.patient = patient
            visit.created_by = request.user
            visit.save()

            messages.success(
                request,
                f"Visit created successfully. Visit Number: {visit.visit_number}"
            )
            return redirect("visit_detail", pk=visit.pk)

        messages.error(request, "Please correct the errors below.")

    return render(request, "patients/patient_visit_form.html", {
        "form": form,
        "patient": patient,
        "title": "Create Visit for Patient",
        "button_text": "Create Visit",
    })


@login_required
def visit_detail(request, pk):
    visit = get_object_or_404(
        PatientVisit.objects.select_related(
            "patient",
            "created_by",
        ),
        pk=pk,
    )

    prescriptions = visit.prescriptions.select_related("prescribed_by").all()
    surgeries = visit.surgeries.select_related("surgeon", "assistant").all()
    bills = visit.bills.select_related("created_by").all()

    context = {
        "visit": visit,
        "prescriptions": prescriptions,
        "surgeries": surgeries,
        "bills": bills,
    }

    return render(request, "patients/visit_detail.html", context)




@login_required
@records_staff_required
def visit_update(request, pk):
    visit = get_object_or_404(PatientVisit, pk=pk)
    form = PatientVisitForm(request.POST or None, instance=visit)

    if request.method == "POST":
        if form.is_valid():
            form.save()
            messages.success(request, "Visit record updated successfully.")
            return redirect("visit_detail", pk=visit.pk)

        messages.error(request, "Please correct the errors below.")

    return render(request, "patients/visit_form.html", {
        "form": form,
        "visit": visit,
        "title": "Update Patient Visit / Encounter",
        "button_text": "Save Changes",
    })


@login_required
def patient_visit_history(request, patient_pk):
    patient = get_object_or_404(Patient, pk=patient_pk)
    visits = patient.visits.select_related("created_by").all()

    return render(request, "patients/patient_visit_history.html", {
        "patient": patient,
        "visits": visits,
    })


@login_required
@clinical_staff_required
def consultation_create(request, visit_pk):
    visit = get_object_or_404(PatientVisit.objects.select_related("patient"), pk=visit_pk)

    if hasattr(visit, "consultation"):
        messages.info(request, "This visit already has a consultation record. You can edit it instead.")
        return redirect("consultation_update", pk=visit.consultation.pk)

    form = ConsultationForm(request.POST or None)

    if request.method == "POST":
        if form.is_valid():
            consultation = form.save(commit=False)
            consultation.visit = visit
            consultation.doctor = request.user
            consultation.save()

            visit.status = PatientVisit.VisitStatus.WITH_DOCTOR
            visit.save(update_fields=["status"])

            messages.success(request, "Consultation record created successfully.")
            return redirect("visit_detail", pk=visit.pk)

        messages.error(request, "Please correct the errors below.")

    return render(request, "patients/consultation_form.html", {
        "form": form,
        "visit": visit,
        "title": "Create Consultation Record",
        "button_text": "Save Consultation",
    })


@login_required
@clinical_staff_required
def consultation_update(request, pk):
    consultation = get_object_or_404(
        Consultation.objects.select_related("visit", "visit__patient"),
        pk=pk,
    )

    form = ConsultationForm(request.POST or None, instance=consultation)

    if request.method == "POST":
        if form.is_valid():
            form.save()
            messages.success(request, "Consultation record updated successfully.")
            return redirect("visit_detail", pk=consultation.visit.pk)

        messages.error(request, "Please correct the errors below.")

    return render(request, "patients/consultation_form.html", {
        "form": form,
        "visit": consultation.visit,
        "consultation": consultation,
        "title": "Update Consultation Record",
        "button_text": "Save Changes",
    })


@login_required
@clinical_staff_required
def eye_examination_create(request, visit_pk):
    visit = get_object_or_404(PatientVisit.objects.select_related("patient"), pk=visit_pk)

    if hasattr(visit, "eye_examination"):
        messages.info(request, "This visit already has an eye examination record. You can edit it instead.")
        return redirect("eye_examination_update", pk=visit.eye_examination.pk)

    form = EyeExaminationForm(request.POST or None)

    if request.method == "POST":
        if form.is_valid():
            exam = form.save(commit=False)
            exam.visit = visit
            exam.examined_by = request.user
            exam.save()

            visit.status = PatientVisit.VisitStatus.WITH_DOCTOR
            visit.save(update_fields=["status"])

            messages.success(request, "Eye examination record created successfully.")
            return redirect("visit_detail", pk=visit.pk)

        messages.error(request, "Please correct the errors below.")

    return render(request, "patients/eye_examination_form.html", {
        "form": form,
        "visit": visit,
        "title": "Create Eye Examination Record",
        "button_text": "Save Eye Examination",
    })


@login_required
@clinical_staff_required
def eye_examination_update(request, pk):
    exam = get_object_or_404(
        EyeExamination.objects.select_related("visit", "visit__patient"),
        pk=pk,
    )

    form = EyeExaminationForm(request.POST or None, instance=exam)

    if request.method == "POST":
        if form.is_valid():
            form.save()
            messages.success(request, "Eye examination record updated successfully.")
            return redirect("visit_detail", pk=exam.visit.pk)

        messages.error(request, "Please correct the errors below.")

    return render(request, "patients/eye_examination_form.html", {
        "form": form,
        "visit": exam.visit,
        "exam": exam,
        "title": "Update Eye Examination Record",
        "button_text": "Save Changes",
    })



@login_required
@clinical_staff_required
def diagnosis_treatment_create(request, visit_pk):
    visit = get_object_or_404(PatientVisit.objects.select_related("patient"), pk=visit_pk)

    if hasattr(visit, "diagnosis_treatment"):
        messages.info(request, "This visit already has a diagnosis/treatment record. You can edit it instead.")
        return redirect("diagnosis_treatment_update", pk=visit.diagnosis_treatment.pk)

    form = DiagnosisTreatmentForm(request.POST or None)

    if request.method == "POST":
        if form.is_valid():
            record = form.save(commit=False)
            record.visit = visit
            record.created_by = request.user
            record.save()

            visit.status = PatientVisit.VisitStatus.WITH_DOCTOR
            visit.save(update_fields=["status"])

            messages.success(request, "Diagnosis and treatment record saved successfully.")
            return redirect("visit_detail", pk=visit.pk)

        messages.error(request, "Please correct the errors below.")

    return render(request, "patients/diagnosis_treatment_form.html", {
        "form": form,
        "visit": visit,
        "title": "Create Diagnosis and Treatment Record",
        "button_text": "Save Diagnosis/Treatment",
    })


@login_required
@clinical_staff_required
def diagnosis_treatment_update(request, pk):
    record = get_object_or_404(
        DiagnosisTreatment.objects.select_related("visit", "visit__patient"),
        pk=pk,
    )

    form = DiagnosisTreatmentForm(request.POST or None, instance=record)

    if request.method == "POST":
        if form.is_valid():
            form.save()
            messages.success(request, "Diagnosis and treatment record updated successfully.")
            return redirect("visit_detail", pk=record.visit.pk)

        messages.error(request, "Please correct the errors below.")

    return render(request, "patients/diagnosis_treatment_form.html", {
        "form": form,
        "visit": record.visit,
        "record": record,
        "title": "Update Diagnosis and Treatment Record",
        "button_text": "Save Changes",
    })


@login_required
@clinical_staff_required
def prescription_create(request, visit_pk):
    visit = get_object_or_404(PatientVisit.objects.select_related("patient"), pk=visit_pk)
    form = PrescriptionForm(request.POST or None)

    if request.method == "POST":
        if form.is_valid():
            prescription = form.save(commit=False)
            prescription.visit = visit
            prescription.prescribed_by = request.user
            prescription.save()

            messages.success(request, "Prescription added successfully.")
            return redirect("visit_detail", pk=visit.pk)

        messages.error(request, "Please correct the errors below.")

    return render(request, "patients/prescription_form.html", {
        "form": form,
        "visit": visit,
        "title": "Add Prescription",
        "button_text": "Save Prescription",
    })


@login_required
@clinical_staff_required
def prescription_update(request, pk):
    prescription = get_object_or_404(
        Prescription.objects.select_related("visit", "visit__patient"),
        pk=pk,
    )

    form = PrescriptionForm(request.POST or None, instance=prescription)

    if request.method == "POST":
        if form.is_valid():
            form.save()
            messages.success(request, "Prescription updated successfully.")
            return redirect("visit_detail", pk=prescription.visit.pk)

        messages.error(request, "Please correct the errors below.")

    return render(request, "patients/prescription_form.html", {
        "form": form,
        "visit": prescription.visit,
        "prescription": prescription,
        "title": "Update Prescription",
        "button_text": "Save Changes",
    })



@login_required
def surgery_list(request):
    query = request.GET.get("q", "").strip()
    status = request.GET.get("status", "").strip()

    surgeries = SurgeryProcedure.objects.select_related(
        "patient",
        "visit",
        "surgeon",
        "assistant",
        "created_by",
    ).all()

    if query:
        surgeries = surgeries.filter(
            Q(procedure_number__icontains=query)
            | Q(patient__file_number__icontains=query)
            | Q(patient__full_name__icontains=query)
            | Q(patient__phone_number__icontains=query)
            | Q(procedure_name__icontains=query)
            | Q(procedure_type__icontains=query)
            | Q(pre_op_diagnosis__icontains=query)
            | Q(post_op_diagnosis__icontains=query)
        )

    if status:
        surgeries = surgeries.filter(status=status)

    return render(request, "patients/surgery_list.html", {
        "surgeries": surgeries,
        "query": query,
        "status": status,
        "status_choices": SurgeryProcedure.ProcedureStatus.choices,
    })


@login_required
@clinical_staff_required
def surgery_create(request):
    form = SurgeryProcedureForm(request.POST or None)

    if request.method == "POST":
        if form.is_valid():
            surgery = form.save(commit=False)
            surgery.created_by = request.user
            surgery.save()

            messages.success(
                request,
                f"Surgery/procedure record created successfully. Procedure Number: {surgery.procedure_number}"
            )
            return redirect("surgery_detail", pk=surgery.pk)

        messages.error(request, "Please correct the errors below.")

    return render(request, "patients/surgery_form.html", {
        "form": form,
        "title": "Create Surgery / Procedure Record",
        "button_text": "Save Surgery / Procedure",
    })


@login_required
@clinical_staff_required
def patient_surgery_create(request, patient_pk):
    patient = get_object_or_404(Patient, pk=patient_pk)
    form = SurgeryProcedureQuickForm(request.POST or None)

    if request.method == "POST":
        if form.is_valid():
            surgery = form.save(commit=False)
            surgery.patient = patient
            surgery.created_by = request.user
            surgery.save()

            messages.success(
                request,
                f"Surgery/procedure record created successfully. Procedure Number: {surgery.procedure_number}"
            )
            return redirect("surgery_detail", pk=surgery.pk)

        messages.error(request, "Please correct the errors below.")

    return render(request, "patients/patient_surgery_form.html", {
        "form": form,
        "patient": patient,
        "title": "Create Surgery / Procedure for Patient",
        "button_text": "Save Surgery / Procedure",
    })


@login_required
@clinical_staff_required
def visit_surgery_create(request, visit_pk):
    visit = get_object_or_404(PatientVisit.objects.select_related("patient"), pk=visit_pk)
    form = SurgeryProcedureQuickForm(request.POST or None)

    if request.method == "POST":
        if form.is_valid():
            surgery = form.save(commit=False)
            surgery.patient = visit.patient
            surgery.visit = visit
            surgery.created_by = request.user
            surgery.save()

            messages.success(
                request,
                f"Surgery/procedure record created successfully. Procedure Number: {surgery.procedure_number}"
            )
            return redirect("surgery_detail", pk=surgery.pk)

        messages.error(request, "Please correct the errors below.")

    return render(request, "patients/visit_surgery_form.html", {
        "form": form,
        "visit": visit,
        "patient": visit.patient,
        "title": "Create Surgery / Procedure for Visit",
        "button_text": "Save Surgery / Procedure",
    })


@login_required
def surgery_detail(request, pk):
    surgery = get_object_or_404(
        SurgeryProcedure.objects.select_related(
            "patient",
            "visit",
            "surgeon",
            "assistant",
            "created_by",
        ),
        pk=pk,
    )

    return render(request, "patients/surgery_detail.html", {
        "surgery": surgery,
    })


@login_required
@clinical_staff_required
def surgery_update(request, pk):
    surgery = get_object_or_404(SurgeryProcedure, pk=pk)
    form = SurgeryProcedureForm(request.POST or None, instance=surgery)

    if request.method == "POST":
        if form.is_valid():
            form.save()
            messages.success(request, "Surgery/procedure record updated successfully.")
            return redirect("surgery_detail", pk=surgery.pk)

        messages.error(request, "Please correct the errors below.")

    return render(request, "patients/surgery_form.html", {
        "form": form,
        "surgery": surgery,
        "title": "Update Surgery / Procedure Record",
        "button_text": "Save Changes",
    })


@login_required
def patient_surgery_history(request, patient_pk):
    patient = get_object_or_404(Patient, pk=patient_pk)
    surgeries = patient.surgeries.select_related("visit", "surgeon", "assistant", "created_by").all()

    return render(request, "patients/patient_surgery_history.html", {
        "patient": patient,
        "surgeries": surgeries,
    })



@login_required
def appointment_list(request):
    query = request.GET.get("q", "").strip()
    status = request.GET.get("status", "").strip()
    date = request.GET.get("date", "").strip()

    appointments = Appointment.objects.select_related(
        "patient",
        "assigned_to",
        "created_by",
    ).all()

    if query:
        appointments = appointments.filter(
            Q(appointment_number__icontains=query)
            | Q(patient__file_number__icontains=query)
            | Q(patient__full_name__icontains=query)
            | Q(patient__phone_number__icontains=query)
            | Q(reason__icontains=query)
        )

    if status:
        appointments = appointments.filter(status=status)

    if date:
        appointments = appointments.filter(appointment_date=date)

    return render(request, "patients/appointment_list.html", {
        "appointments": appointments,
        "query": query,
        "status": status,
        "date": date,
        "status_choices": Appointment.AppointmentStatus.choices,
    })


@login_required
@records_staff_required
def appointment_create(request):
    form = AppointmentForm(request.POST or None)

    if request.method == "POST":
        if form.is_valid():
            appointment = form.save(commit=False)
            appointment.created_by = request.user
            appointment.save()

            messages.success(
                request,
                f"Appointment created successfully. Appointment Number: {appointment.appointment_number}"
            )
            return redirect("appointment_detail", pk=appointment.pk)

        messages.error(request, "Please correct the errors below.")

    return render(request, "patients/appointment_form.html", {
        "form": form,
        "title": "Create Appointment",
        "button_text": "Save Appointment",
    })


@login_required
@records_staff_required
def patient_appointment_create(request, patient_pk):
    patient = get_object_or_404(Patient, pk=patient_pk)
    form = AppointmentQuickForm(request.POST or None)

    if request.method == "POST":
        if form.is_valid():
            appointment = form.save(commit=False)
            appointment.patient = patient
            appointment.created_by = request.user
            appointment.save()

            messages.success(
                request,
                f"Appointment created successfully. Appointment Number: {appointment.appointment_number}"
            )
            return redirect("appointment_detail", pk=appointment.pk)

        messages.error(request, "Please correct the errors below.")

    return render(request, "patients/patient_appointment_form.html", {
        "form": form,
        "patient": patient,
        "title": "Create Appointment for Patient",
        "button_text": "Save Appointment",
    })


@login_required
def appointment_detail(request, pk):
    appointment = get_object_or_404(
        Appointment.objects.select_related("patient", "assigned_to", "created_by"),
        pk=pk,
    )

    return render(request, "patients/appointment_detail.html", {
        "appointment": appointment,
    })


@login_required
@records_staff_required
def appointment_update(request, pk):
    appointment = get_object_or_404(Appointment, pk=pk)
    form = AppointmentForm(request.POST or None, instance=appointment)

    if request.method == "POST":
        if form.is_valid():
            form.save()
            messages.success(request, "Appointment updated successfully.")
            return redirect("appointment_detail", pk=appointment.pk)

        messages.error(request, "Please correct the errors below.")

    return render(request, "patients/appointment_form.html", {
        "form": form,
        "appointment": appointment,
        "title": "Update Appointment",
        "button_text": "Save Changes",
    })


@login_required
def patient_appointment_history(request, patient_pk):
    patient = get_object_or_404(Patient, pk=patient_pk)
    appointments = patient.appointments.select_related("assigned_to", "created_by").all()

    return render(request, "patients/patient_appointment_history.html", {
        "patient": patient,
        "appointments": appointments,
    })



@login_required
def appointment_calendar(request):
    today = timezone.localdate()
    selected_date = request.GET.get("date", "")

    if selected_date:
        calendar_date = selected_date
    else:
        calendar_date = today

    appointments = Appointment.objects.select_related(
        "patient",
        "assigned_to",
        "created_by",
    ).filter(
        appointment_date=calendar_date
    )

    return render(request, "patients/appointment_calendar.html", {
        "appointments": appointments,
        "calendar_date": calendar_date,
        "today": today,
        "status_choices": Appointment.AppointmentStatus.choices,
    })


@login_required
def clinic_queue(request):
    today = timezone.localdate()

    visits = PatientVisit.objects.select_related(
        "patient",
        "created_by",
    ).filter(
        visit_date__date=today
    )

    appointments = Appointment.objects.select_related(
        "patient",
        "assigned_to",
    ).filter(
        appointment_date=today
    )

    return render(request, "patients/clinic_queue.html", {
        "visits": visits,
        "appointments": appointments,
        "today": today,
    })


@login_required
@clinical_staff_required
def doctor_worklist(request):
    today = timezone.localdate()

    visits = PatientVisit.objects.select_related(
        "patient",
        "created_by",
    ).filter(
        visit_date__date=today
    ).exclude(
        status=PatientVisit.VisitStatus.COMPLETED
    )

    appointments = Appointment.objects.select_related(
        "patient",
        "assigned_to",
    ).filter(
        appointment_date=today
    )

    if not request.user.is_superuser:
        appointments = appointments.filter(
            Q(assigned_to=request.user) | Q(assigned_to__isnull=True)
        )

    return render(request, "patients/doctor_worklist.html", {
        "visits": visits,
        "appointments": appointments,
        "today": today,
    })


@login_required
@records_staff_required
def appointment_mark_arrived(request, pk):
    appointment = get_object_or_404(Appointment, pk=pk)
    appointment.status = Appointment.AppointmentStatus.ARRIVED
    appointment.save(update_fields=["status", "updated_at"])

    messages.success(request, "Appointment marked as arrived.")
    return redirect("clinic_queue")


@login_required
@records_staff_required
def appointment_mark_completed(request, pk):
    appointment = get_object_or_404(Appointment, pk=pk)
    appointment.status = Appointment.AppointmentStatus.COMPLETED
    appointment.save(update_fields=["status", "updated_at"])

    messages.success(request, "Appointment marked as completed.")
    return redirect("clinic_queue")


@login_required
@clinical_staff_required
def visit_mark_completed(request, pk):
    visit = get_object_or_404(PatientVisit, pk=pk)
    visit.status = PatientVisit.VisitStatus.COMPLETED
    visit.save(update_fields=["status", "updated_at"])

    messages.success(request, "Visit marked as completed.")
    return redirect("doctor_worklist")




@login_required
@finance_staff_required
def bill_list(request):
    query = request.GET.get("q", "").strip()
    status = request.GET.get("status", "").strip()

    bills = Bill.objects.select_related(
        "patient",
        "visit",
        "surgery",
        "appointment",
        "created_by",
    ).all()

    filtered_bills = bills

    if query:
        filtered_bills = filtered_bills.filter(
            Q(bill_number__icontains=query)
            | Q(patient__file_number__icontains=query)
            | Q(patient__full_name__icontains=query)
            | Q(patient__phone_number__icontains=query)
            | Q(bill_title__icontains=query)
        )

    if status:
        filtered_bills = filtered_bills.filter(status=status)

    total_bills = bills.count()
    unpaid_bills = bills.exclude(status=Bill.BillStatus.PAID).exclude(status=Bill.BillStatus.CANCELLED).count()

    return render(request, "patients/bill_list.html", {
        "bills": filtered_bills,
        "query": query,
        "status": status,
        "status_choices": Bill.BillStatus.choices,
        "total_bills": total_bills,
        "unpaid_bills": unpaid_bills,
    })


@login_required
@finance_staff_required
def bill_create(request):
    form = BillForm(request.POST or None)

    if request.method == "POST":
        if form.is_valid():
            bill = form.save(commit=False)
            bill.created_by = request.user
            bill.save()

            messages.success(request, f"Bill created successfully. Bill Number: {bill.bill_number}")
            return redirect("bill_detail", pk=bill.pk)

        messages.error(request, "Please correct the errors below.")

    return render(request, "patients/bill_form.html", {
        "form": form,
        "title": "Create Bill",
        "button_text": "Save Bill",
    })


@login_required
@finance_staff_required
def patient_bill_create(request, patient_pk):
    patient = get_object_or_404(Patient, pk=patient_pk)
    form = BillQuickForm(request.POST or None)

    if request.method == "POST":
        if form.is_valid():
            bill = form.save(commit=False)
            bill.patient = patient
            bill.created_by = request.user
            bill.save()

            messages.success(request, f"Bill created successfully. Bill Number: {bill.bill_number}")
            return redirect("bill_detail", pk=bill.pk)

        messages.error(request, "Please correct the errors below.")

    return render(request, "patients/patient_bill_form.html", {
        "form": form,
        "patient": patient,
        "title": "Create Bill for Patient",
        "button_text": "Save Bill",
    })


@login_required
@finance_staff_required
def bill_detail(request, pk):
    bill = get_object_or_404(
        Bill.objects.select_related(
            "patient",
            "visit",
            "surgery",
            "appointment",
            "created_by",
        ),
        pk=pk,
    )

    payments = bill.payments.select_related("received_by").all()

    context = {
        "bill": bill,
        "payments": payments,
        "can_pay": bill.balance > 0 and bill.status != Bill.BillStatus.CANCELLED,
    }

    return render(request, "patients/bill_detail.html", context)


@login_required
@finance_staff_required
def bill_update(request, pk):
    bill = get_object_or_404(Bill, pk=pk)
    form = BillForm(request.POST or None, instance=bill)

    if request.method == "POST":
        if form.is_valid():
            form.save()
            messages.success(request, "Bill updated successfully.")
            return redirect("bill_detail", pk=bill.pk)

        messages.error(request, "Please correct the errors below.")

    return render(request, "patients/bill_form.html", {
        "form": form,
        "bill": bill,
        "title": "Update Bill",
        "button_text": "Save Changes",
    })


@login_required
@finance_staff_required
def payment_create(request, bill_pk):
    bill = get_object_or_404(Bill.objects.select_related("patient"), pk=bill_pk)
    form = PaymentForm(request.POST or None, bill=bill)

    if request.method == "POST":
        if form.is_valid():
            payment = form.save(commit=False)
            payment.bill = bill
            payment.received_by = request.user
            payment.save()

            log_activity(
                request,
                AuditLog.ActionType.CREATE,
                "Payments",
                f"Recorded payment of ₦{payment.amount} for bill {bill.bill_number}.",
                object_id=payment.pk,
                object_repr=payment.receipt_number,
            )

            messages.success(request, f"Payment recorded successfully. Receipt Number: {payment.receipt_number}")
            return redirect("bill_detail", pk=bill.pk)

        messages.error(request, "Please correct the errors below.")

    return render(request, "patients/payment_form.html", {
        "form": form,
        "bill": bill,
        "title": "Record Payment",
        "button_text": "Save Payment",
    })

    

@login_required
@finance_staff_required
def patient_bill_history(request, patient_pk):
    patient = get_object_or_404(Patient, pk=patient_pk)
    bills = patient.bills.select_related("created_by").all()

    return render(request, "patients/patient_bill_history.html", {
        "patient": patient,
        "bills": bills,
    })



@login_required
def global_search(request):
    query = request.GET.get("q", "").strip()

    patients = Patient.objects.none()
    visits = PatientVisit.objects.none()
    appointments = Appointment.objects.none()
    surgeries = SurgeryProcedure.objects.none()
    bills = Bill.objects.none()
    prescriptions = Prescription.objects.none()
    diagnoses = DiagnosisTreatment.objects.none()
    consultations = Consultation.objects.none()

    if query:
        patients = Patient.objects.filter(
            Q(file_number__icontains=query)
            | Q(full_name__icontains=query)
            | Q(phone_number__icontains=query)
            | Q(address__icontains=query)
            | Q(diagnosis__icontains=query)
            | Q(treatment__icontains=query)
            | Q(eye_complaint__icontains=query)
        )[:20]

        visits = PatientVisit.objects.select_related("patient").filter(
            Q(visit_number__icontains=query)
            | Q(patient__file_number__icontains=query)
            | Q(patient__full_name__icontains=query)
            | Q(chief_complaint__icontains=query)
            | Q(brief_history__icontains=query)
        )[:20]

        appointments = Appointment.objects.select_related("patient").filter(
            Q(appointment_number__icontains=query)
            | Q(patient__file_number__icontains=query)
            | Q(patient__full_name__icontains=query)
            | Q(patient__phone_number__icontains=query)
            | Q(reason__icontains=query)
        )[:20]

        surgeries = SurgeryProcedure.objects.select_related("patient").filter(
            Q(procedure_number__icontains=query)
            | Q(patient__file_number__icontains=query)
            | Q(patient__full_name__icontains=query)
            | Q(procedure_name__icontains=query)
            | Q(procedure_type__icontains=query)
            | Q(pre_op_diagnosis__icontains=query)
            | Q(post_op_diagnosis__icontains=query)
        )[:20]

        bills = Bill.objects.select_related("patient").filter(
            Q(bill_number__icontains=query)
            | Q(patient__file_number__icontains=query)
            | Q(patient__full_name__icontains=query)
            | Q(patient__phone_number__icontains=query)
            | Q(bill_title__icontains=query)
        )[:20]

        prescriptions = Prescription.objects.select_related("visit", "visit__patient").filter(
            Q(drug_name__icontains=query)
            | Q(dosage__icontains=query)
            | Q(frequency__icontains=query)
            | Q(visit__visit_number__icontains=query)
            | Q(visit__patient__file_number__icontains=query)
            | Q(visit__patient__full_name__icontains=query)
        )[:20]

        diagnoses = DiagnosisTreatment.objects.select_related("visit", "visit__patient").filter(
            Q(primary_diagnosis__icontains=query)
            | Q(secondary_diagnosis__icontains=query)
            | Q(differential_diagnosis__icontains=query)
            | Q(treatment_plan__icontains=query)
            | Q(visit__visit_number__icontains=query)
            | Q(visit__patient__file_number__icontains=query)
            | Q(visit__patient__full_name__icontains=query)
        )[:20]

        consultations = Consultation.objects.select_related("visit", "visit__patient").filter(
            Q(presenting_complaint__icontains=query)
            | Q(provisional_diagnosis__icontains=query)
            | Q(final_diagnosis__icontains=query)
            | Q(treatment_plan__icontains=query)
            | Q(visit__visit_number__icontains=query)
            | Q(visit__patient__file_number__icontains=query)
            | Q(visit__patient__full_name__icontains=query)
        )[:20]

    return render(request, "patients/global_search.html", {
        "query": query,
        "patients": patients,
        "visits": visits,
        "appointments": appointments,
        "surgeries": surgeries,
        "bills": bills,
        "prescriptions": prescriptions,
        "diagnoses": diagnoses,
        "consultations": consultations,
    })



@login_required
@records_staff_required
def patient_excel_import(request):
    if request.method == "POST":
        excel_file = request.FILES.get("excel_file")

        if not excel_file:
            messages.error(request, "Please select an Excel file.")
            return redirect("patient_excel_import")

        if not excel_file.name.endswith(".xlsx"):
            messages.error(request, "Only .xlsx files are allowed.")
            return redirect("patient_excel_import")

        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active

        imported_count = 0
        skipped_count = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):
            full_name = row[0]
            gender = row[1]
            phone_number = row[2]
            address = row[3]
            occupation = row[4]
            next_of_kin_name = row[5]
            next_of_kin_phone = row[6]
            medical_history = row[7]
            allergy_history = row[8]
            eye_complaint = row[9]
            diagnosis = row[10]
            treatment = row[11]
            payment_status = row[12]
            notes = row[13]

            if not full_name:
                skipped_count += 1
                continue

            duplicate = Patient.objects.filter(
                full_name__iexact=str(full_name).strip(),
                phone_number__iexact=str(phone_number).strip() if phone_number else "",
            ).exists()

            if duplicate:
                skipped_count += 1
                continue

            Patient.objects.create(
                full_name=str(full_name).strip(),
                gender=str(gender).upper() if gender else Patient.Gender.MALE,
                phone_number=str(phone_number).strip() if phone_number else "",
                address=address or "",
                occupation=occupation or "",
                next_of_kin_name=next_of_kin_name or "",
                next_of_kin_phone=str(next_of_kin_phone).strip() if next_of_kin_phone else "",
                medical_history=medical_history or "",
                allergy_history=allergy_history or "",
                eye_complaint=eye_complaint or "",
                diagnosis=diagnosis or "",
                treatment=treatment or "",
                payment_status=payment_status if payment_status else Patient.PaymentStatus.NOT_APPLICABLE,
                notes=notes or "",
                registered_by=request.user,
            )

            imported_count += 1

        log_activity(
            request,
            AuditLog.ActionType.IMPORT,
            "Patients",
            f"Imported Excel patient records. Imported: {imported_count}, Skipped: {skipped_count}.",
        )

        messages.success(
            request,
            f"Excel import completed. Imported: {imported_count}, Skipped: {skipped_count}"
        )
        return redirect("patient_list")

    return render(request, "patients/patient_excel_import.html")



@login_required
def patient_excel_export(request):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Patients"

    headers = [
        "File Number",
        "Full Name",
        "Gender",
        "Age",
        "Date of Birth",
        "Phone Number",
        "Address",
        "Occupation",
        "Next of Kin",
        "Next of Kin Phone",
        "Medical History",
        "Allergy History",
        "Eye Complaint",
        "Diagnosis",
        "Treatment",
        "Payment Status",
        "Registration Date",
        "Registered By",
        "Notes",
    ]

    sheet.append(headers)

    patients = Patient.objects.select_related("registered_by").all()

    for patient in patients:
        sheet.append([
            patient.file_number,
            patient.full_name,
            patient.get_gender_display(),
            patient.display_age or "",
            patient.date_of_birth.strftime("%Y-%m-%d") if patient.date_of_birth else "",
            patient.phone_number,
            patient.address,
            patient.occupation,
            patient.next_of_kin_name,
            patient.next_of_kin_phone,
            patient.medical_history,
            patient.allergy_history,
            patient.eye_complaint,
            patient.diagnosis,
            patient.treatment,
            patient.get_payment_status_display(),
            patient.registration_date.strftime("%Y-%m-%d %H:%M"),
            str(patient.registered_by) if patient.registered_by else "",
            patient.notes,
        ])

    for col_num, column_title in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        sheet.column_dimensions[column_letter].width = 22

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    filename = f"capital_eye_patients_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    workbook.save(response)

    log_activity(
        request,
        AuditLog.ActionType.EXPORT,
        "Patients",
        "Exported patient records to Excel.",
    )

    return response



@login_required
@records_staff_required
def patient_excel_template(request):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Patient Import Template"

    headers = [
        "Full Name",
        "Gender",
        "Phone Number",
        "Address",
        "Occupation",
        "Next of Kin Name",
        "Next of Kin Phone",
        "Medical History",
        "Allergy History",
        "Eye Complaint",
        "Diagnosis",
        "Treatment",
        "Payment Status",
        "Notes",
    ]

    sheet.append(headers)

    sample = [
        "John Musa",
        "MALE",
        "08000000000",
        "Kano, Nigeria",
        "Trader",
        "Aisha Musa",
        "08111111111",
        "No known chronic illness",
        "No known allergy",
        "Blurred vision",
        "Cataract",
        "Eye drops and follow-up",
        "NOT_APPLICABLE",
        "Imported sample row. Delete before real import.",
    ]

    sheet.append(sample)

    for col_num, column_title in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        sheet.column_dimensions[column_letter].width = 25

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = 'attachment; filename="patient_import_template.xlsx"'

    workbook.save(response)
    return response



@login_required
def patient_card_pdf(request, pk):
    patient = get_object_or_404(Patient, pk=pk)

    log_activity(
        request,
        AuditLog.ActionType.PRINT,
        "PDF Reports",
        f"Generated PDF for patient {patient.file_number}.",
        object_id=patient.pk,
        object_repr=patient.file_number,
    )

    return render_to_pdf(
        "patients/pdf/patient_card_pdf.html",
        {"patient": patient},
        f"{patient.file_number}_patient_card.pdf",
    )


@login_required
def patient_report_pdf(request, pk):
    patient = get_object_or_404(Patient, pk=pk)

    visits = patient.visits.all()
    appointments = patient.appointments.all()
    surgeries = patient.surgeries.all()
    bills = patient.bills.all()

    log_activity(
        request,
        AuditLog.ActionType.PRINT,
        "PDF Reports",
        f"Generated PDF for patient {patient.file_number}.",
        object_id=patient.pk,
        object_repr=patient.file_number,
    )

    return render_to_pdf(
        "patients/pdf/patient_report_pdf.html",
        {
            "patient": patient,
            "visits": visits,
            "appointments": appointments,
            "surgeries": surgeries,
            "bills": bills,
        },
        f"{patient.file_number}_full_report.pdf",
    )


@login_required
def visit_report_pdf(request, pk):
    visit = get_object_or_404(
        PatientVisit.objects.select_related("patient"),
        pk=pk,
    )

    log_activity(
        request,
        AuditLog.ActionType.PRINT,
        "PDF Reports",
        f"Generated PDF for patient {visit.patient.file_number}.",
        object_id=visit.patient.pk,
        object_repr=visit.patient.file_number,
    )

    return render_to_pdf(
        "patients/pdf/visit_report_pdf.html",
        {"visit": visit},
        f"{visit.visit_number}_visit_report.pdf",
    )


@login_required
@finance_staff_required
def bill_receipt_pdf(request, pk):
    bill = get_object_or_404(
        Bill.objects.select_related("patient", "created_by"),
        pk=pk,
    )

    payments = bill.payments.select_related("received_by").all()

    log_activity(
        request,
        AuditLog.ActionType.PRINT,
        "PDF Reports",
        f"Generated PDF for patient {bill.patient.file_number}.",
        object_id=bill.patient.pk,
        object_repr=bill.patient.file_number,
    )

    return render_to_pdf(
        "patients/pdf/bill_receipt_pdf.html",
        {
            "bill": bill,
            "payments": payments,
        },
        f"{bill.bill_number}_receipt.pdf",
    )



@login_required
def surgery_theatre_dashboard(request):
    today = timezone.localdate()
    selected_date = request.GET.get("date", "")

    if selected_date:
        theatre_date = selected_date
    else:
        theatre_date = today

    surgeries = SurgeryProcedure.objects.select_related(
        "patient",
        "visit",
        "surgeon",
        "assistant",
        "created_by",
    ).filter(
        scheduled_date__date=theatre_date
    )

    planned_count = surgeries.filter(status=SurgeryProcedure.ProcedureStatus.PLANNED).count()
    in_progress_count = surgeries.filter(status=SurgeryProcedure.ProcedureStatus.IN_PROGRESS).count()
    completed_count = surgeries.filter(status=SurgeryProcedure.ProcedureStatus.COMPLETED).count()
    postponed_count = surgeries.filter(status=SurgeryProcedure.ProcedureStatus.POSTPONED).count()

    return render(request, "patients/surgery_theatre_dashboard.html", {
        "surgeries": surgeries,
        "theatre_date": theatre_date,
        "today": today,
        "planned_count": planned_count,
        "in_progress_count": in_progress_count,
        "completed_count": completed_count,
        "postponed_count": postponed_count,
    })


@login_required
@clinical_staff_required
def surgery_mark_in_progress(request, pk):
    surgery = get_object_or_404(SurgeryProcedure, pk=pk)
    surgery.status = SurgeryProcedure.ProcedureStatus.IN_PROGRESS
    surgery.save(update_fields=["status", "updated_at"])

    messages.success(request, "Surgery/procedure marked as in progress.")
    return redirect("surgery_theatre_dashboard")


@login_required
@clinical_staff_required
def surgery_mark_completed(request, pk):
    surgery = get_object_or_404(SurgeryProcedure, pk=pk)
    surgery.status = SurgeryProcedure.ProcedureStatus.COMPLETED

    if not surgery.procedure_date:
        surgery.procedure_date = timezone.now()

    surgery.save(update_fields=["status", "procedure_date", "updated_at"])

    messages.success(request, "Surgery/procedure marked as completed.")
    return redirect("surgery_theatre_dashboard")


@login_required
@clinical_staff_required
def surgery_mark_postponed(request, pk):
    surgery = get_object_or_404(SurgeryProcedure, pk=pk)
    surgery.status = SurgeryProcedure.ProcedureStatus.POSTPONED
    surgery.save(update_fields=["status", "updated_at"])

    messages.success(request, "Surgery/procedure marked as postponed.")
    return redirect("surgery_theatre_dashboard")


@login_required
@clinical_staff_required
def surgery_mark_cancelled(request, pk):
    surgery = get_object_or_404(SurgeryProcedure, pk=pk)
    surgery.status = SurgeryProcedure.ProcedureStatus.CANCELLED
    surgery.save(update_fields=["status", "updated_at"])

    messages.success(request, "Surgery/procedure marked as cancelled.")
    return redirect("surgery_theatre_dashboard")



@login_required
def reports_dashboard(request):
    today = timezone.localdate()

    total_patients = Patient.objects.count()
    total_visits = PatientVisit.objects.count()
    total_appointments = Appointment.objects.count()
    total_surgeries = SurgeryProcedure.objects.count()
    total_bills = Bill.objects.count()

    total_revenue = Payment.objects.aggregate(
        total=Sum("amount")
    )["total"] or 0

    today_revenue = Payment.objects.filter(
        payment_date__date=today
    ).aggregate(
        total=Sum("amount")
    )["total"] or 0

    outstanding_balance = sum(
        bill.balance
        for bill in Bill.objects.exclude(status=Bill.BillStatus.PAID).exclude(status=Bill.BillStatus.CANCELLED)
    )

    return render(request, "patients/reports_dashboard.html", {
        "today": today,
        "total_patients": total_patients,
        "total_visits": total_visits,
        "total_appointments": total_appointments,
        "total_surgeries": total_surgeries,
        "total_bills": total_bills,
        "total_revenue": total_revenue,
        "today_revenue": today_revenue,
        "outstanding_balance": outstanding_balance,
    })


@login_required
@finance_staff_required
def financial_report(request):
    today = timezone.localdate()
    start_date = request.GET.get("start_date") or today
    end_date = request.GET.get("end_date") or today

    payments = Payment.objects.select_related(
        "bill",
        "bill__patient",
        "received_by",
    ).filter(
        payment_date__date__gte=start_date,
        payment_date__date__lte=end_date,
    )

    bills = Bill.objects.select_related(
        "patient",
        "created_by",
    ).filter(
        created_at__date__gte=start_date,
        created_at__date__lte=end_date,
    )

    total_payments = payments.aggregate(total=Sum("amount"))["total"] or 0
    total_billed = bills.aggregate(total=Sum("total_amount"))["total"] or 0
    total_discount = bills.aggregate(total=Sum("discount"))["total"] or 0
    total_paid_on_bills = bills.aggregate(total=Sum("amount_paid"))["total"] or 0

    outstanding = sum(bill.balance for bill in bills)

    payment_method_summary = payments.values("payment_method").annotate(
        total=Sum("amount"),
        count=Count("id"),
    ).order_by("payment_method")

    return render(request, "patients/reports/financial_report.html", {
        "start_date": start_date,
        "end_date": end_date,
        "payments": payments,
        "bills": bills,
        "total_payments": total_payments,
        "total_billed": total_billed,
        "total_discount": total_discount,
        "total_paid_on_bills": total_paid_on_bills,
        "outstanding": outstanding,
        "payment_method_summary": payment_method_summary,
    })


@login_required
def clinical_report(request):
    today = timezone.localdate()
    start_date = request.GET.get("start_date") or today
    end_date = request.GET.get("end_date") or today

    visits = PatientVisit.objects.select_related(
        "patient",
        "created_by",
    ).filter(
        visit_date__date__gte=start_date,
        visit_date__date__lte=end_date,
    )

    consultations = Consultation.objects.select_related(
        "visit",
        "visit__patient",
        "doctor",
    ).filter(
        consultation_date__date__gte=start_date,
        consultation_date__date__lte=end_date,
    )

    prescriptions = Prescription.objects.select_related(
        "visit",
        "visit__patient",
        "prescribed_by",
    ).filter(
        prescribed_at__date__gte=start_date,
        prescribed_at__date__lte=end_date,
    )

    diagnosis_records = DiagnosisTreatment.objects.select_related(
        "visit",
        "visit__patient",
    ).filter(
        created_at__date__gte=start_date,
        created_at__date__lte=end_date,
    )

    visit_type_summary = visits.values("visit_type").annotate(
        count=Count("id")
    ).order_by("visit_type")

    return render(request, "patients/reports/clinical_report.html", {
        "start_date": start_date,
        "end_date": end_date,
        "visits": visits,
        "consultations": consultations,
        "prescriptions": prescriptions,
        "diagnosis_records": diagnosis_records,
        "visit_type_summary": visit_type_summary,
    })


@login_required
def appointment_report(request):
    today = timezone.localdate()
    start_date = request.GET.get("start_date") or today
    end_date = request.GET.get("end_date") or today
    status = request.GET.get("status", "").strip()

    appointments = Appointment.objects.select_related(
        "patient",
        "assigned_to",
        "created_by",
    ).filter(
        appointment_date__gte=start_date,
        appointment_date__lte=end_date,
    )

    if status:
        appointments = appointments.filter(status=status)

    status_summary = appointments.values("status").annotate(
        count=Count("id")
    ).order_by("status")

    return render(request, "patients/reports/appointment_report.html", {
        "start_date": start_date,
        "end_date": end_date,
        "status": status,
        "appointments": appointments,
        "status_summary": status_summary,
        "status_choices": Appointment.AppointmentStatus.choices,
    })


@login_required
def surgery_report(request):
    today = timezone.localdate()
    start_date = request.GET.get("start_date") or today
    end_date = request.GET.get("end_date") or today
    status = request.GET.get("status", "").strip()

    surgeries = SurgeryProcedure.objects.select_related(
        "patient",
        "visit",
        "surgeon",
        "assistant",
    ).filter(
        scheduled_date__date__gte=start_date,
        scheduled_date__date__lte=end_date,
    )

    if status:
        surgeries = surgeries.filter(status=status)

    status_summary = surgeries.values("status").annotate(
        count=Count("id")
    ).order_by("status")

    return render(request, "patients/reports/surgery_report.html", {
        "start_date": start_date,
        "end_date": end_date,
        "status": status,
        "surgeries": surgeries,
        "status_summary": status_summary,
        "status_choices": SurgeryProcedure.ProcedureStatus.choices,
    })


    