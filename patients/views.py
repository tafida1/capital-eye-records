import mimetypes
import os
import json
import math

from urllib.parse import quote

from django.contrib import messages
from django.contrib.auth.decorators import login_required 

from django.shortcuts import render, redirect, get_object_or_404

from django.core.exceptions import PermissionDenied
from django.db import transaction
from django.views.decorators.http import require_POST

from accounts.decorators import records_staff_required, clinical_staff_required, finance_staff_required

from .forms import (
    PatientForm,
    FamilyGroupForm,
    PatientVisitForm,
    PatientVisitQuickForm,
    ConsultationForm,
    EyeExaminationForm,
    ClinicalAttachmentForm,
    ClinicalAttachmentReviewForm,
    DiagnosisTreatmentForm,
    PrescriptionForm,
    SurgeryProcedureForm,
    SurgeryProcedureQuickForm,
    AppointmentForm,
    AppointmentQuickForm,
    BillForm,
    BillQuickForm,
    PaymentForm,

    ContactLensAssessmentForm,
    ContactLensTrialForm,
    ContactLensPrescriptionForm,
    ContactLensPrescriptionApprovalForm,
    ContactLensPrescriptionDispensingForm,
    ContactLensFollowUpForm,
)

from .models import (
    Patient,
    FamilyGroup,
    PatientVisit,
    Consultation,
    EyeExamination,
    ClinicalAttachment,
    ClinicalImageAnnotation,
    DiagnosisTreatment,
    Prescription,
    SurgeryProcedure,
    Appointment,
    Bill,
    Payment,

    ContactLensAssessment,
    ContactLensTrial,
    ContactLensPrescription,
    ContactLensFollowUp,
)

from django.template.loader import get_template, render_to_string
from audit_logs.utils import log_activity
from audit_logs.models import AuditLog
from notifications.utils import create_notification
from notifications.models import Notification
from django.http import (
    FileResponse,
    Http404,
    HttpResponse,
    JsonResponse,
)
from xhtml2pdf import pisa
from io import BytesIO
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime, time, timedelta
from django.utils import timezone
from django.db.models import Max, Q, Sum, Count
from django.db.models.functions import TruncMonth
from django.core.paginator import Paginator




def render_to_pdf(template_src, context_dict, filename):
    template = get_template(template_src)
    html = template.render(context_dict)

    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)

    if pdf.err:
        return HttpResponse("PDF generation error", status=500)

    response = HttpResponse(result.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="{filename}"'
    return response


def autosize_excel_columns(sheet):
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            try:
                value_length = len(str(cell.value)) if cell.value else 0
                if value_length > max_length:
                    max_length = value_length
            except Exception:
                pass

        sheet.column_dimensions[column_letter].width = min(max_length + 3, 40)


CLINICAL_ATTACHMENT_MANAGEMENT_ROLES = {
    "SUPER_ADMIN",
    "HOSPITAL_ADMIN",
    "DOCTOR",
    "OPHTHALMOLOGIST",
    "OPTOMETRIST",
    "LAB_STAFF",
    "PROCEDURE_STAFF",
}


CLINICAL_ATTACHMENT_REVIEW_ROLES = {
    "SUPER_ADMIN",
    "HOSPITAL_ADMIN",
    "DOCTOR",
    "OPHTHALMOLOGIST",
    "OPTOMETRIST",
}


def user_role_value(user):
    """
    Return the authenticated user's role as a normalized string.
    """

    return str(
        getattr(user, "role", "") or ""
    ).strip().upper()


def user_can_manage_clinical_attachments(user):
    """
    Determine whether the user may upload, update, deactivate,
    or restore clinical attachments.
    """

    if user is None or not user.is_authenticated:
        return False

    if user.is_superuser:
        return True

    if getattr(user, "is_staff", False):
        return True

    return (
        user_role_value(user)
        in CLINICAL_ATTACHMENT_MANAGEMENT_ROLES
    )


def user_can_review_clinical_attachments(user):
    """
    Determine whether the user may clinically review an attachment.
    """

    if user is None or not user.is_authenticated:
        return False

    if user.is_superuser:
        return True

    role = user_role_value(user)

    return role in CLINICAL_ATTACHMENT_REVIEW_ROLES


def ensure_attachment_management_permission(user):
    if not user_can_manage_clinical_attachments(user):
        raise PermissionDenied(
            "You do not have permission to manage clinical attachments."
        )


def ensure_attachment_review_permission(user):
    if not user_can_review_clinical_attachments(user):
        raise PermissionDenied(
            "You do not have permission to review clinical attachments."
        )


def user_can_annotate_clinical_attachments(user):
    """
    Only authorized clinical reviewers or administrators may create,
    edit, finalize or remove clinical image annotations.
    """

    return user_can_review_clinical_attachments(user)


def ensure_annotation_permission(user):
    if not user_can_annotate_clinical_attachments(user):
        raise PermissionDenied(
            "You do not have permission to annotate clinical images."
        )


# ============================================================
# CONTACT LENS PERMISSION HELPERS
# ============================================================


def user_can_manage_contact_lenses(user):
    """
    Users already authorized to manage clinical investigation
    records may create and update Contact Lens clinical records.

    This reuses the current clinical permission architecture rather
    than introducing a duplicate role system.
    """

    if not user or not user.is_authenticated:
        return False

    if user.is_superuser:
        return True

    return user_can_manage_clinical_attachments(user)


def user_can_approve_contact_lens_prescriptions(user):
    """
    Approval is restricted to users already authorized to perform
    clinical investigation reviews, such as doctors, optometrists
    and appropriate administrators.
    """

    if not user or not user.is_authenticated:
        return False

    if user.is_superuser:
        return True

    return user_can_review_clinical_attachments(user)


def user_can_dispense_contact_lens_prescriptions(user):
    """
    The first implementation uses the existing clinical-management
    permission. This may later be narrowed to an Optical/Dispensing
    role when the clinic defines that role.
    """

    if not user or not user.is_authenticated:
        return False

    if user.is_superuser:
        return True

    return user_can_manage_clinical_attachments(user)


def user_can_print_contact_lens_prescriptions(user):
    """
    Allow authenticated clinical users with Contact Lens viewing
    privileges to print prescriptions.

    Approval and dispensing permissions remain separately controlled.
    """

    if not user or not user.is_authenticated:
        return False

    if user.is_superuser:
        return True

    return (
        user_can_manage_contact_lenses(user)
        or user_can_approve_contact_lens_prescriptions(user)
        or user_can_dispense_contact_lens_prescriptions(user)
    )


def ensure_contact_lens_print_permission(user):
    if not user_can_print_contact_lens_prescriptions(user):
        raise PermissionDenied(
            "You do not have permission to print Contact Lens prescriptions."
        )


def ensure_contact_lens_management_permission(user):
    if not user_can_manage_contact_lenses(user):
        raise PermissionDenied(
            "You do not have permission to manage Contact Lens records."
        )


def ensure_contact_lens_approval_permission(user):
    if not user_can_approve_contact_lens_prescriptions(user):
        raise PermissionDenied(
            "You do not have permission to approve Contact Lens prescriptions."
        )


def ensure_contact_lens_dispensing_permission(user):
    if not user_can_dispense_contact_lens_prescriptions(user):
        raise PermissionDenied(
            "You do not have permission to dispense Contact Lens prescriptions."
        )


def get_contact_lens_assessment_queryset():
    return (
        ContactLensAssessment.objects
        .select_related(
            "patient",
            "visit",
            "eye_examination",
            "assessed_by",
        )
        .prefetch_related(
            "trial_lenses",
            "prescriptions",
        )
    )


def get_contact_lens_trial_queryset():
    return (
        ContactLensTrial.objects
        .select_related(
            "assessment",
            "assessment__patient",
            "assessment__visit",
            "fitted_by",
        )
    )


def get_contact_lens_prescription_queryset(
    include_inactive=False,
):
    queryset = (
        ContactLensPrescription.objects
        .select_related(
            "assessment",
            "patient",
            "visit",
            "prescribed_by",
            "approved_by",
            "dispensed_by",
        )
        .prefetch_related(
            "follow_ups",
        )
    )

    if not include_inactive:
        queryset = queryset.filter(
            is_active=True
        )

    return queryset


def contact_lens_prescription_print_context(
    prescription,
    *,
    request=None,
):
    """
    Build a shared context for browser printing and PDF generation.
    """

    today = timezone.localdate()

    is_expired = bool(
        prescription.valid_until
        and prescription.valid_until < today
    )

    right_has_prescription = any(
        value not in {None, ""}
        for value in [
            prescription.right_lens_design,
            prescription.right_brand_name,
            prescription.right_base_curve,
            prescription.right_diameter,
            prescription.right_sphere,
            prescription.right_cylinder,
            prescription.right_add_power,
        ]
    )

    left_has_prescription = any(
        value not in {None, ""}
        for value in [
            prescription.left_lens_design,
            prescription.left_brand_name,
            prescription.left_base_curve,
            prescription.left_diameter,
            prescription.left_sphere,
            prescription.left_cylinder,
            prescription.left_add_power,
        ]
    )

    approved_statuses = {
        ContactLensPrescription
        .PrescriptionStatus
        .APPROVED,

        ContactLensPrescription
        .PrescriptionStatus
        .DISPENSED,
    }

    return {
        "prescription": prescription,
        "assessment": prescription.assessment,
        "patient": prescription.patient,
        "visit": prescription.visit,

        "right_has_prescription": right_has_prescription,
        "left_has_prescription": left_has_prescription,

        "is_expired": is_expired,

        "is_final_prescription": (
            prescription.status in approved_statuses
        ),

        "generated_at": timezone.localtime(),
        "generated_by": (
            request.user
            if request is not None
            else None
        ),
    }


def get_contact_lens_follow_up_queryset():
    return (
        ContactLensFollowUp.objects
        .select_related(
            "prescription",
            "prescription__assessment",
            "patient",
            "reviewed_by",
        )
    )


# ============================================================
# CONTACT LENS FOLLOW-UP MONITORING HELPERS
# ============================================================


def contact_lens_follow_up_due_category(
    follow_up,
    *,
    today=None,
):
    """
    Classify a Contact Lens follow-up by its due status.

    Completed, cancelled and missed records are treated separately
    from active scheduled reviews.
    """

    if today is None:
        today = timezone.localdate()

    if (
        follow_up.status
        == ContactLensFollowUp
        .FollowUpStatus
        .COMPLETED
    ):
        return "COMPLETED"

    if (
        follow_up.status
        == ContactLensFollowUp
        .FollowUpStatus
        .CANCELLED
    ):
        return "CANCELLED"

    if (
        follow_up.status
        == ContactLensFollowUp
        .FollowUpStatus
        .MISSED
    ):
        return "MISSED"

    follow_up_date = timezone.localtime(
        follow_up.follow_up_date
    ).date()

    if follow_up_date < today:
        return "OVERDUE"

    if follow_up_date == today:
        return "DUE_TODAY"

    if follow_up_date <= today + timedelta(days=7):
        return "DUE_SOON"

    return "UPCOMING"


def contact_lens_follow_up_safety_flags(
    follow_up,
):
    """
    Return a list of clinical safety flags detected from an existing
    Contact Lens follow-up record.

    This does not diagnose the patient. It highlights information
    already entered by clinical staff.
    """

    flags = []

    complications = (
        follow_up.complications
        or ""
    ).strip()

    if complications:
        flags.append(
            {
                "code": "COMPLICATION",
                "label": "Complication recorded",
                "severity": "danger",
            }
        )

    if (
        follow_up.comfort_score_right
        is not None
        and follow_up.comfort_score_right <= 4
    ):
        flags.append(
            {
                "code": "LOW_RIGHT_COMFORT",
                "label": (
                    "Low right-eye comfort score"
                ),
                "severity": "warning",
            }
        )

    if (
        follow_up.comfort_score_left
        is not None
        and follow_up.comfort_score_left <= 4
    ):
        flags.append(
            {
                "code": "LOW_LEFT_COMFORT",
                "label": (
                    "Low left-eye comfort score"
                ),
                "severity": "warning",
            }
        )

    high_risk_lens_conditions = {
        ContactLensFollowUp
        .LensCondition
        .DAMAGED,

        ContactLensFollowUp
        .LensCondition
        .POOR_HYGIENE,

        ContactLensFollowUp
        .LensCondition
        .REPLACEMENT_REQUIRED,
    }

    if (
        follow_up.lens_condition
        in high_risk_lens_conditions
    ):
        flags.append(
            {
                "code": "LENS_CONDITION",
                "label": (
                    follow_up
                    .get_lens_condition_display()
                ),
                "severity": "danger",
            }
        )

    if follow_up.lens_parameters_changed:
        flags.append(
            {
                "code": "PARAMETERS_CHANGED",
                "label": (
                    "Lens parameters changed"
                ),
                "severity": "info",
            }
        )

    if (
        follow_up.status
        == ContactLensFollowUp
        .FollowUpStatus
        .MISSED
    ):
        flags.append(
            {
                "code": "MISSED_REVIEW",
                "label": "Follow-up missed",
                "severity": "warning",
            }
        )

    return flags


def contact_lens_prescription_expiry_category(
    prescription,
    *,
    today=None,
):
    """
    Classify the expiry status of an active Contact Lens prescription.
    """

    if today is None:
        today = timezone.localdate()

    if not prescription.valid_until:
        return ""

    if prescription.valid_until < today:
        return "EXPIRED"

    if prescription.valid_until == today:
        return "EXPIRES_TODAY"

    if (
        prescription.valid_until
        <= today + timedelta(days=30)
    ):
        return "EXPIRING_SOON"

    return ""


CLINICAL_ANNOTATION_MAX_OBJECTS = 200
CLINICAL_ANNOTATION_MAX_JSON_SIZE = 500 * 1024
CLINICAL_ANNOTATION_MAX_POINTS_PER_OBJECT = 5000

CLINICAL_ANNOTATION_ALLOWED_TOOLS = {
    "freehand",
    "arrow",
    "rectangle",
    "ellipse",
    "text",
}


def annotation_number_is_valid(value):
    """
    Accept finite numeric values only.
    """

    return (
        isinstance(value, (int, float))
        and not isinstance(value, bool)
        and math.isfinite(value)
    )


def validate_normalized_coordinate(value):
    """
    Annotation coordinates are normalized from 0.0 to 1.0.
    A small tolerance is allowed for strokes touching the boundary.
    """

    return (
        annotation_number_is_valid(value)
        and -0.05 <= value <= 1.05
    )


def validate_annotation_object(annotation_object):
    if not isinstance(annotation_object, dict):
        raise ValueError(
            "Each annotation object must be a JSON object."
        )

    tool_type = annotation_object.get("type")

    if tool_type not in CLINICAL_ANNOTATION_ALLOWED_TOOLS:
        raise ValueError(
            "The annotation contains an unsupported drawing tool."
        )

    stroke_width = annotation_object.get(
        "strokeWidth",
        2,
    )

    if (
        not annotation_number_is_valid(stroke_width)
        or not 1 <= stroke_width <= 20
    ):
        raise ValueError(
            "Annotation stroke width must be between 1 and 20."
        )

    colour = str(
        annotation_object.get(
            "color",
            "#ff0000",
        )
    )

    if (
        len(colour) > 20
        or not colour.startswith("#")
    ):
        raise ValueError(
            "The annotation contains an invalid colour value."
        )

    if tool_type == "freehand":
        points = annotation_object.get("points")

        if not isinstance(points, list):
            raise ValueError(
                "A freehand annotation must contain drawing points."
            )

        if len(points) > CLINICAL_ANNOTATION_MAX_POINTS_PER_OBJECT:
            raise ValueError(
                "A freehand annotation contains too many points."
            )

        for point in points:
            if (
                not isinstance(point, dict)
                or not validate_normalized_coordinate(
                    point.get("x")
                )
                or not validate_normalized_coordinate(
                    point.get("y")
                )
            ):
                raise ValueError(
                    "A freehand annotation contains invalid coordinates."
                )

        return

    if tool_type in {
        "arrow",
        "rectangle",
        "ellipse",
    }:
        required_coordinates = (
            "startX",
            "startY",
            "endX",
            "endY",
        )

        for coordinate_name in required_coordinates:
            if not validate_normalized_coordinate(
                annotation_object.get(
                    coordinate_name
                )
            ):
                raise ValueError(
                    "The annotation contains invalid coordinates."
                )

        return

    if tool_type == "text":
        if not validate_normalized_coordinate(
            annotation_object.get("x")
        ):
            raise ValueError(
                "The text annotation has an invalid X coordinate."
            )

        if not validate_normalized_coordinate(
            annotation_object.get("y")
        ):
            raise ValueError(
                "The text annotation has an invalid Y coordinate."
            )

        text_value = str(
            annotation_object.get(
                "text",
                "",
            )
        ).strip()

        if not text_value:
            raise ValueError(
                "A text annotation cannot be empty."
            )

        if len(text_value) > 500:
            raise ValueError(
                "A text annotation must not exceed 500 characters."
            )


def validate_annotation_payload(annotation_data):
    """
    Validate and return a safe annotation JSON structure.
    """

    if not isinstance(annotation_data, dict):
        raise ValueError(
            "The annotation payload must be a JSON object."
        )

    objects = annotation_data.get(
        "objects",
        [],
    )

    if not isinstance(objects, list):
        raise ValueError(
            "The annotation objects must be supplied as a list."
        )

    if len(objects) > CLINICAL_ANNOTATION_MAX_OBJECTS:
        raise ValueError(
            (
                "This annotation contains too many drawing objects. "
                f"The maximum is {CLINICAL_ANNOTATION_MAX_OBJECTS}."
            )
        )

    for annotation_object in objects:
        validate_annotation_object(
            annotation_object
        )

    normalized_payload = {
        "schemaVersion": 1,
        "objects": objects,
    }

    encoded_payload = json.dumps(
        normalized_payload,
        separators=(",", ":"),
    ).encode("utf-8")

    if len(encoded_payload) > CLINICAL_ANNOTATION_MAX_JSON_SIZE:
        raise ValueError(
            "The annotation data is too large to save."
        )

    return normalized_payload


def get_clinical_attachment_queryset(include_inactive=False):
    """
    Return the optimized attachment queryset used by all attachment views.
    """

    queryset = ClinicalAttachment.objects.select_related(
        "patient",
        "visit",
        "uploaded_by",
        "reviewed_by",
    )

    if not include_inactive:
        queryset = queryset.filter(is_active=True)

    return queryset


def safe_attachment_response_filename(attachment):
    """
    Return a safe user-facing response filename.
    """

    filename = (
        attachment.original_filename
        or os.path.basename(
            attachment.attachment_file.name
        )
        or "clinical_attachment"
    )

    filename = filename.replace(
        "\r",
        "",
    ).replace(
        "\n",
        "",
    )

    return filename[:255]


def attachment_content_type(attachment):
    content_type, _encoding = mimetypes.guess_type(
        safe_attachment_response_filename(attachment)
    )

    if content_type:
        return content_type

    return "application/octet-stream"


def build_attachment_file_response(
    attachment,
    *,
    as_attachment,
):
    """
    Open the protected file through Django storage and return it
    through an authenticated response.
    """

    if (
        not attachment.attachment_file
        or not attachment.attachment_file.name
    ):
        raise Http404(
            "The clinical file is not available."
        )

    storage = attachment.attachment_file.storage
    stored_name = attachment.attachment_file.name

    if not storage.exists(stored_name):
        raise Http404(
            "The clinical file could not be found on storage."
        )

    try:
        file_handle = storage.open(
            stored_name,
            "rb",
        )
    except (OSError, ValueError):
        raise Http404(
            "The clinical file could not be opened."
        )

    filename = safe_attachment_response_filename(
        attachment
    )

    response = FileResponse(
        file_handle,
        as_attachment=as_attachment,
        filename=filename,
        content_type=attachment_content_type(
            attachment
        ),
    )

    response["X-Content-Type-Options"] = "nosniff"
    response["Cache-Control"] = (
        "private, no-store, no-cache, must-revalidate"
    )
    response["Pragma"] = "no-cache"
    response["Expires"] = "0"

    encoded_filename = quote(filename)

    disposition = (
        "attachment"
        if as_attachment
        else "inline"
    )

    response["Content-Disposition"] = (
        f"{disposition}; "
        f"filename*=UTF-8''{encoded_filename}"
    )

    return response


def normalize_timeline_datetime(value):
    """
    Convert a date or datetime into an aware datetime suitable
    for chronological timeline sorting.
    """

    if value is None:
        return None

    if isinstance(value, datetime):
        result = value
    else:
        result = datetime.combine(
            value,
            time.min,
        )

    if timezone.is_naive(result):
        result = timezone.make_aware(
            result,
            timezone.get_current_timezone(),
        )

    return result


def timeline_user_name(user):
    """
    Return a safe display name for a clinical staff user.
    """

    if user is None:
        return "Not recorded"

    full_name = str(
        user.get_full_name() or ""
    ).strip()

    return full_name or user.username


# ============================================================
# CONTACT LENS TIMELINE HELPERS
# ============================================================


def contact_lens_timeline_status_class(status):
    """
    Return a safe CSS class for Contact Lens timeline status badges.
    """

    status_value = str(status or "").strip().upper()

    status_map = {
        "PENDING": "pending",
        "SUITABLE": "completed",
        "SUITABLE_WITH_CAUTION": "warning",
        "TEMPORARILY_UNSUITABLE": "warning",
        "UNSUITABLE": "danger",

        "DRAFT": "draft",
        "PENDING_APPROVAL": "pending",
        "APPROVED": "completed",
        "DISPENSED": "completed",
        "EXPIRED": "warning",
        "CANCELLED": "danger",

        "SCHEDULED": "pending",
        "COMPLETED": "completed",
        "MISSED": "warning",
    }

    return status_map.get(
        status_value,
        "default",
    )


def contact_lens_eye_parameter_summary(
    prescription,
    *,
    prefix,
    eye_label,
):
    """
    Produce a concise timeline summary of one eye's final Contact
    Lens prescription parameters.
    """

    design_value = getattr(
        prescription,
        f"{prefix}_lens_design",
        "",
    )

    if design_value == "OTHER":
        design_label = getattr(
            prescription,
            f"{prefix}_lens_design_other",
            "",
        )
    else:
        display_method = getattr(
            prescription,
            f"get_{prefix}_lens_design_display",
            None,
        )

        design_label = (
            display_method()
            if callable(display_method)
            else design_value
        )

    values = []

    if design_label:
        values.append(str(design_label))

    parameter_labels = [
        (f"{prefix}_sphere", "SPH"),
        (f"{prefix}_cylinder", "CYL"),
        (f"{prefix}_axis", "Axis"),
        (f"{prefix}_base_curve", "BC"),
        (f"{prefix}_diameter", "DIA"),
        (f"{prefix}_add_power", "Add"),
    ]

    for field_name, label in parameter_labels:
        value = getattr(
            prescription,
            field_name,
            None,
        )

        if value not in {None, ""}:
            values.append(f"{label} {value}")

    if not values:
        return ""

    return f"{eye_label}: " + ", ".join(values)


def ophthalmology_dashboard_date_range(request):
    """
    Resolve the selected dashboard date range.

    Supported periods:
    today, 7days, 30days, 90days, year and all.
    """

    selected_period = (
        request.GET.get("period", "30days")
        .strip()
        .lower()
    )

    allowed_periods = {
        "today",
        "7days",
        "30days",
        "90days",
        "year",
        "all",
    }

    if selected_period not in allowed_periods:
        selected_period = "30days"

    today = timezone.localdate()
    end_date = today

    if selected_period == "today":
        start_date = today

    elif selected_period == "7days":
        start_date = today - timedelta(days=6)

    elif selected_period == "30days":
        start_date = today - timedelta(days=29)

    elif selected_period == "90days":
        start_date = today - timedelta(days=89)

    elif selected_period == "year":
        start_date = today.replace(
            month=1,
            day=1,
        )

    else:
        start_date = None
        end_date = None

    return {
        "selected_period": selected_period,
        "start_date": start_date,
        "end_date": end_date,
    }


def model_has_field(model_class, field_name):
    """
    Return True when a Django model contains the named field.
    """

    return any(
        field.name == field_name
        for field in model_class._meta.get_fields()
    )


def first_existing_model_field(model_class, candidates):
    """
    Return the first field name that exists on a model.

    This keeps dashboard summaries compatible with the current
    project model names without inventing unavailable fields.
    """

    for candidate in candidates:
        if model_has_field(model_class, candidate):
            return candidate

    return None


def record_display_value(record, field_names, default=""):
    """
    Return the first non-empty value from the supplied field names.
    """

    for field_name in field_names:
        value = getattr(
            record,
            field_name,
            None,
        )

        if value not in {
            None,
            "",
        }:
            return value

    return default


@login_required
def patient_list(request):
    query = request.GET.get("q", "").strip()

    patients = Patient.objects.select_related("family_group", "registered_by").all()

    if query:
        patients = patients.filter(
            Q(file_number__icontains=query)
            | Q(full_name__icontains=query)
            | Q(phone_number__icontains=query)
            | Q(family_group__family_code__icontains=query)
            | Q(family_group__family_name__icontains=query)
            | Q(diagnosis__icontains=query)
            | Q(eye_complaint__icontains=query)
            | Q(surgery_procedure_details__icontains=query)
        )

    paginator = Paginator(patients, 20)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(request, "patients/patient_list.html", {
        "patients": page_obj,
        "page_obj": page_obj,
        "query": query,
    })



@login_required
@records_staff_required
def patient_create(request):
    form = PatientForm(request.POST or None)

    if request.method == "POST":
        if form.is_valid():
            patient = form.save(commit=False)
            patient.registered_by = request.user

            if patient.family_group and not patient.family_group_name:
                patient.family_group_name = patient.family_group.family_name

            patient.save()

            log_activity(
                request,
                AuditLog.ActionType.CREATE,
                "Patients",
                f"Registered patient {patient.full_name} with file number {patient.file_number}.",
                object_id=patient.pk,
                object_repr=patient.file_number,
            )

            messages.success(
                request,
                f"Patient registered successfully. File Number: {patient.file_number}"
            )
            return redirect("patient_detail", pk=patient.pk)

        messages.error(request, "Please correct the errors below.")

    return render(request, "patients/patient_form.html", {
        "form": form,
        "title": "Register New Patient",
        "button_text": "Register Patient",
    })


@login_required
def patient_detail(request, pk):
    patient = get_object_or_404(
        Patient.objects.select_related(
            "family_group",
            "registered_by",
        ),
        pk=pk,
    )

    visits = patient.visits.select_related(
        "created_by"
    )[:8]

    appointments = patient.appointments.select_related(
        "assigned_to",
        "created_by",
    )[:5]

    surgeries = patient.surgeries.select_related(
        "surgeon",
        "assistant",
        "created_by",
    )[:5]

    bills = patient.bills.select_related(
        "created_by"
    )[:5]

    unpaid_bills = patient.bills.exclude(
        status=Bill.BillStatus.PAID
    ).exclude(
        status=Bill.BillStatus.CANCELLED
    )

    total_balance = sum(
        bill.balance
        for bill in unpaid_bills
    )

    # =====================================================
    # CLINICAL ATTACHMENTS / INVESTIGATION RESULTS
    # =====================================================

    active_attachments = (
        patient.clinical_attachments
        .filter(is_active=True)
        .select_related(
            "visit",
            "uploaded_by",
            "reviewed_by",
        )
    )

    attachment_total_count = (
        active_attachments.count()
    )

    attachment_pending_count = (
        active_attachments.filter(
            review_status=(
                ClinicalAttachment
                .ReviewStatus
                .PENDING
            )
        ).count()
    )

    attachment_reviewed_count = (
        active_attachments.filter(
            review_status=(
                ClinicalAttachment
                .ReviewStatus
                .REVIEWED
            )
        ).count()
    )

    attachment_attention_count = (
        active_attachments.filter(
            review_status=(
                ClinicalAttachment
                .ReviewStatus
                .NEEDS_ATTENTION
            )
        ).count()
    )

    latest_clinical_attachments = (
        active_attachments.order_by(
            "-investigation_date",
            "-uploaded_at",
        )[:6]
    )

    context = {
        "patient": patient,
        "visits": visits,
        "appointments": appointments,
        "surgeries": surgeries,
        "bills": bills,
        "total_balance": total_balance,
        "unpaid_bills_count": (
            unpaid_bills.count()
        ),

        # Clinical Attachments
        "latest_clinical_attachments": (
            latest_clinical_attachments
        ),
        "attachment_total_count": (
            attachment_total_count
        ),
        "attachment_pending_count": (
            attachment_pending_count
        ),
        "attachment_reviewed_count": (
            attachment_reviewed_count
        ),
        "attachment_attention_count": (
            attachment_attention_count
        ),
        "can_manage_clinical_attachments": (
            user_can_manage_clinical_attachments(
                request.user
            )
        ),
        "can_review_clinical_attachments": (
            user_can_review_clinical_attachments(
                request.user
            )
        ),
    }

    return render(
        request,
        "patients/patient_detail.html",
        context,
    )



@login_required
@records_staff_required
def patient_update(request, pk):
    patient = get_object_or_404(Patient, pk=pk)
    form = PatientForm(request.POST or None, instance=patient)

    if request.method == "POST":
        if form.is_valid():
            patient = form.save(commit=False)

            if patient.family_group and not patient.family_group_name:
                patient.family_group_name = patient.family_group.family_name

            patient.save()

            log_activity(
                request,
                AuditLog.ActionType.UPDATE,
                "Patients",
                f"Updated patient record {patient.full_name}.",
                object_id=patient.pk,
                object_repr=patient.file_number,
            )

            messages.success(request, "Patient record updated successfully.")
            return redirect("patient_detail", pk=patient.pk)

        messages.error(request, "Please correct the errors below.")

    return render(request, "patients/patient_form.html", {
        "form": form,
        "patient": patient,
        "title": "Update Patient Record",
        "button_text": "Save Changes",
    })


@login_required
def family_group_list(request):
    query = request.GET.get("q", "").strip()

    family_groups = FamilyGroup.objects.select_related("created_by").all()

    if query:
        family_groups = family_groups.filter(
            Q(family_code__icontains=query)
            | Q(family_name__icontains=query)
            | Q(head_of_family__icontains=query)
            | Q(primary_phone__icontains=query)
        )

    return render(request, "patients/family_group_list.html", {
        "family_groups": family_groups,
        "query": query,
    })


@login_required
@records_staff_required
def family_group_create(request):
    form = FamilyGroupForm(request.POST or None)

    if request.method == "POST":
        if form.is_valid():
            family_group = form.save(commit=False)
            family_group.created_by = request.user
            family_group.save()

            messages.success(
                request,
                f"Family group created successfully. Family Code: {family_group.family_code}"
            )
            return redirect("family_group_detail", pk=family_group.pk)

        messages.error(request, "Please correct the errors below.")

    return render(request, "patients/family_group_form.html", {
        "form": form,
        "title": "Create Family / Group Record",
        "button_text": "Create Family Group",
    })


@login_required
def family_group_detail(request, pk):
    family_group = get_object_or_404(FamilyGroup, pk=pk)
    members = family_group.members.all()

    return render(request, "patients/family_group_detail.html", {
        "family_group": family_group,
        "members": members,
    })


@login_required
@records_staff_required
def family_group_update(request, pk):
    family_group = get_object_or_404(FamilyGroup, pk=pk)
    form = FamilyGroupForm(request.POST or None, instance=family_group)

    if request.method == "POST":
        if form.is_valid():
            form.save()
            messages.success(request, "Family group updated successfully.")
            return redirect("family_group_detail", pk=family_group.pk)

        messages.error(request, "Please correct the errors below.")

    return render(request, "patients/family_group_form.html", {
        "form": form,
        "family_group": family_group,
        "title": "Update Family / Group Record",
        "button_text": "Save Changes",
    })



@login_required
def visit_list(request):
    query = request.GET.get("q", "").strip()
    status = request.GET.get("status", "").strip()

    visits = PatientVisit.objects.select_related("patient", "created_by").all()

    if query:
        visits = visits.filter(
            Q(visit_number__icontains=query)
            | Q(patient__file_number__icontains=query)
            | Q(patient__full_name__icontains=query)
            | Q(patient__phone_number__icontains=query)
            | Q(chief_complaint__icontains=query)
            | Q(brief_history__icontains=query)
        )

    if status:
        visits = visits.filter(status=status)

    return render(request, "patients/visit_list.html", {
        "visits": visits,
        "query": query,
        "status": status,
        "status_choices": PatientVisit.VisitStatus.choices,
    })


@login_required
@records_staff_required
def visit_create(request):
    form = PatientVisitForm(request.POST or None)

    if request.method == "POST":
        if form.is_valid():
            visit = form.save(commit=False)
            visit.created_by = request.user
            visit.save()

            messages.success(
                request,
                f"Patient visit created successfully. Visit Number: {visit.visit_number}"
            )
            return redirect("visit_detail", pk=visit.pk)

        messages.error(request, "Please correct the errors below.")

    return render(request, "patients/visit_form.html", {
        "form": form,
        "title": "Create Patient Visit / Encounter",
        "button_text": "Create Visit",
    })


@login_required
@records_staff_required
def patient_visit_create(request, patient_pk):
    patient = get_object_or_404(Patient, pk=patient_pk)
    form = PatientVisitQuickForm(request.POST or None)

    if request.method == "POST":
        if form.is_valid():
            visit = form.save(commit=False)
            visit.patient = patient
            visit.created_by = request.user
            visit.save()

            messages.success(
                request,
                f"Visit created successfully. Visit Number: {visit.visit_number}"
            )
            return redirect("visit_detail", pk=visit.pk)

        messages.error(request, "Please correct the errors below.")

    return render(request, "patients/patient_visit_form.html", {
        "form": form,
        "patient": patient,
        "title": "Create Visit for Patient",
        "button_text": "Create Visit",
    })



@login_required
def visit_detail(request, pk):
    visit = get_object_or_404(
        PatientVisit.objects.select_related(
            "patient",
            "created_by",
        ),
        pk=pk,
    )

    prescriptions = visit.prescriptions.select_related("prescribed_by").all()
    surgeries = visit.surgeries.select_related("surgeon", "assistant").all()
    bills = visit.bills.select_related("created_by").all()

    clinical_attachments = (
        visit.clinical_attachments
        .filter(is_active=True)
        .select_related(
            "uploaded_by",
            "reviewed_by",
        )
        .order_by(
            "-investigation_date",
            "-uploaded_at",
        )
    )

    contact_lens_assessment = (
        ContactLensAssessment.objects
        .filter(visit=visit)
        .first()
    )

    context = {
        "visit": visit,
        "prescriptions": prescriptions,
        "surgeries": surgeries,
        "bills": bills,
        "clinical_attachments": clinical_attachments,
        "contact_lens_assessment": contact_lens_assessment,
        "can_manage_clinical_attachments": (
            user_can_manage_clinical_attachments(
                request.user
            )
        ),
        "can_review_clinical_attachments": (
            user_can_review_clinical_attachments(
                request.user
            )
        ),
    }

    return render(request, "patients/visit_detail.html", context)



@login_required
@records_staff_required
def visit_update(request, pk):
    visit = get_object_or_404(PatientVisit, pk=pk)
    form = PatientVisitForm(request.POST or None, instance=visit)

    if request.method == "POST":
        if form.is_valid():
            form.save()
            messages.success(request, "Visit record updated successfully.")
            return redirect("visit_detail", pk=visit.pk)

        messages.error(request, "Please correct the errors below.")

    return render(request, "patients/visit_form.html", {
        "form": form,
        "visit": visit,
        "title": "Update Patient Visit / Encounter",
        "button_text": "Save Changes",
    })


@login_required
def patient_visit_history(request, patient_pk):
    patient = get_object_or_404(Patient, pk=patient_pk)
    visits = patient.visits.select_related("created_by").all()

    return render(request, "patients/patient_visit_history.html", {
        "patient": patient,
        "visits": visits,
    })


@login_required
@clinical_staff_required
def consultation_create(request, visit_pk):
    visit = get_object_or_404(PatientVisit.objects.select_related("patient"), pk=visit_pk)

    if hasattr(visit, "consultation"):
        messages.info(request, "This visit already has a consultation record. You can edit it instead.")
        return redirect("consultation_update", pk=visit.consultation.pk)

    form = ConsultationForm(request.POST or None)

    if request.method == "POST":
        if form.is_valid():
            consultation = form.save(commit=False)
            consultation.visit = visit
            consultation.doctor = request.user
            consultation.save()

            visit.status = PatientVisit.VisitStatus.WITH_DOCTOR
            visit.save(update_fields=["status"])

            messages.success(request, "Consultation record created successfully.")
            return redirect("visit_detail", pk=visit.pk)

        messages.error(request, "Please correct the errors below.")

    return render(request, "patients/consultation_form.html", {
        "form": form,
        "visit": visit,
        "title": "Create Consultation Record",
        "button_text": "Save Consultation",
    })


@login_required
@clinical_staff_required
def consultation_update(request, pk):
    consultation = get_object_or_404(
        Consultation.objects.select_related("visit", "visit__patient"),
        pk=pk,
    )

    form = ConsultationForm(request.POST or None, instance=consultation)

    if request.method == "POST":
        if form.is_valid():
            form.save()
            messages.success(request, "Consultation record updated successfully.")
            return redirect("visit_detail", pk=consultation.visit.pk)

        messages.error(request, "Please correct the errors below.")

    return render(request, "patients/consultation_form.html", {
        "form": form,
        "visit": consultation.visit,
        "consultation": consultation,
        "title": "Update Consultation Record",
        "button_text": "Save Changes",
    })


@login_required
@clinical_staff_required
def eye_examination_create(request, visit_pk):
    visit = get_object_or_404(PatientVisit.objects.select_related("patient"), pk=visit_pk)

    if hasattr(visit, "eye_examination"):
        messages.info(request, "This visit already has an eye examination record. You can edit it instead.")
        return redirect("eye_examination_update", pk=visit.eye_examination.pk)

    form = EyeExaminationForm(
        request.POST or None,
        request_user=request.user,
    )

    if request.method == "POST":
        if form.is_valid():
            exam = form.save(commit=False)
            exam.visit = visit
            exam.examined_by = request.user
            exam.save()

            visit.status = PatientVisit.VisitStatus.WITH_DOCTOR
            visit.save(update_fields=["status"])

            messages.success(request, "Eye examination record created successfully.")
            return redirect("visit_detail", pk=visit.pk)

        messages.error(request, "Please correct the errors below.")

    return render(request, "patients/eye_examination_form.html", {
        "form": form,
        "visit": visit,
        "title": "Create Eye Examination Record",
        "button_text": "Save Eye Examination",
    })


@login_required
@clinical_staff_required
def eye_examination_update(request, pk):
    exam = get_object_or_404(
        EyeExamination.objects.select_related("visit", "visit__patient"),
        pk=pk,
    )

    form = EyeExaminationForm(
        request.POST or None,
        instance=exam,
        request_user=request.user,
    )

    if request.method == "POST":
        if form.is_valid():
            form.save()
            messages.success(request, "Eye examination record updated successfully.")
            return redirect("visit_detail", pk=exam.visit.pk)

        messages.error(request, "Please correct the errors below.")

    return render(request, "patients/eye_examination_form.html", {
        "form": form,
        "visit": exam.visit,
        "exam": exam,
        "title": "Update Eye Examination Record",
        "button_text": "Save Changes",
    })



@login_required
@clinical_staff_required
def clinical_attachment_create(request, visit_pk):
    """
    Upload one clinical attachment for a patient visit.

    Patient, visit and uploader are assigned by trusted server-side logic.
    """

    visit = get_object_or_404(
        PatientVisit.objects.select_related("patient"),
        pk=visit_pk,
    )

    ensure_attachment_management_permission(request.user)

    form = ClinicalAttachmentForm(
        request.POST or None,
        request.FILES or None,
        visit=visit,
        request_user=request.user,
    )

    if request.method == "POST":
        if form.is_valid():
            with transaction.atomic():
                attachment = form.save()

                log_activity(
                    request,
                    AuditLog.ActionType.CREATE,
                    "Clinical Attachments",
                    (
                        f"Uploaded clinical attachment "
                        f"'{attachment.title}' for visit "
                        f"{visit.visit_number} and patient "
                        f"{visit.patient.file_number}."
                    ),
                    object_id=attachment.pk,
                    object_repr=attachment.title,
                )

            messages.success(
                request,
                "Clinical investigation file uploaded successfully.",
            )

            return redirect(
                "clinical_attachment_detail",
                pk=attachment.pk,
            )

        messages.error(
            request,
            "The clinical file could not be uploaded. "
            "Please correct the errors below.",
        )

    return render(
        request,
        "patients/clinical_attachments/attachment_form.html",
        {
            "form": form,
            "visit": visit,
            "patient": visit.patient,
            "attachment": None,
            "title": "Upload Clinical Investigation",
            "button_text": "Upload Clinical File",
        },
    )


@login_required
@clinical_staff_required
def visit_clinical_attachment_list(request, visit_pk):
    """
    Display active clinical attachments belonging to one visit.
    """

    visit = get_object_or_404(
        PatientVisit.objects.select_related("patient"),
        pk=visit_pk,
    )

    attachments = (
        get_clinical_attachment_queryset()
        .filter(visit=visit)
    )

    query = request.GET.get("q", "").strip()
    category = request.GET.get("category", "").strip()
    eye_side = request.GET.get("eye_side", "").strip()
    review_status = request.GET.get(
        "review_status",
        "",
    ).strip()

    if query:
        attachments = attachments.filter(
            Q(title__icontains=query)
            | Q(description__icontains=query)
            | Q(original_filename__icontains=query)
            | Q(category_other__icontains=query)
            | Q(review_notes__icontains=query)
        )

    valid_categories = {
        value
        for value, _label
        in ClinicalAttachment.Category.choices
    }

    if category in valid_categories:
        attachments = attachments.filter(category=category)

    valid_eye_sides = {
        value
        for value, _label
        in ClinicalAttachment.EyeSide.choices
    }

    if eye_side in valid_eye_sides:
        attachments = attachments.filter(
            eye_side=eye_side,
        )

    valid_review_statuses = {
        value
        for value, _label
        in ClinicalAttachment.ReviewStatus.choices
    }

    if review_status in valid_review_statuses:
        attachments = attachments.filter(
            review_status=review_status,
        )

    paginator = Paginator(attachments, 20)

    page_obj = paginator.get_page(
        request.GET.get("page")
    )

    context = {
        "visit": visit,
        "patient": visit.patient,
        "attachments": page_obj,
        "page_obj": page_obj,
        "query": query,
        "selected_category": category,
        "selected_eye_side": eye_side,
        "selected_review_status": review_status,
        "category_choices": ClinicalAttachment.Category.choices,
        "eye_side_choices": ClinicalAttachment.EyeSide.choices,
        "review_status_choices": (
            ClinicalAttachment.ReviewStatus.choices
        ),
        "can_manage_attachments": (
            user_can_manage_clinical_attachments(request.user)
        ),
        "can_review_attachments": (
            user_can_review_clinical_attachments(request.user)
        ),
    }

    return render(
        request,
        "patients/clinical_attachments/attachment_list.html",
        context,
    )


@login_required
@clinical_staff_required
def patient_clinical_attachment_list(request, patient_pk):
    """
    Display the complete active attachment history for one patient.
    """

    patient = get_object_or_404(
        Patient,
        pk=patient_pk,
    )

    attachments = (
        get_clinical_attachment_queryset()
        .filter(patient=patient)
    )

    query = request.GET.get("q", "").strip()
    category = request.GET.get("category", "").strip()
    review_status = request.GET.get(
        "review_status",
        "",
    ).strip()

    if query:
        attachments = attachments.filter(
            Q(title__icontains=query)
            | Q(description__icontains=query)
            | Q(original_filename__icontains=query)
            | Q(visit__visit_number__icontains=query)
            | Q(category_other__icontains=query)
        )

    if category:
        attachments = attachments.filter(
            category=category,
        )

    if review_status:
        attachments = attachments.filter(
            review_status=review_status,
        )

    paginator = Paginator(attachments, 20)

    page_obj = paginator.get_page(
        request.GET.get("page")
    )

    return render(
        request,
        "patients/clinical_attachments/patient_attachment_history.html",
        {
            "patient": patient,
            "attachments": page_obj,
            "page_obj": page_obj,
            "query": query,
            "selected_category": category,
            "selected_review_status": review_status,
            "category_choices": (
                ClinicalAttachment.Category.choices
            ),
            "review_status_choices": (
                ClinicalAttachment.ReviewStatus.choices
            ),
            "can_manage_attachments": (
                user_can_manage_clinical_attachments(
                    request.user
                )
            ),
            "can_review_attachments": (
                user_can_review_clinical_attachments(
                    request.user
                )
            ),
        },
    )


@login_required
@clinical_staff_required
def clinical_attachment_detail(request, pk):
    attachment = get_object_or_404(
        get_clinical_attachment_queryset(
            include_inactive=True,
        ),
        pk=pk,
    )

    return render(
        request,
        "patients/clinical_attachments/attachment_detail.html",
        {
            "attachment": attachment,
            "visit": attachment.visit,
            "patient": attachment.patient,
            "can_manage_attachment": (
                user_can_manage_clinical_attachments(
                    request.user
                )
            ),
            "can_review_attachment": (
                user_can_review_clinical_attachments(
                    request.user
                )
            ),
            "can_permanently_delete": (
                request.user.is_superuser
            ),
        },
    )


@login_required
@clinical_staff_required
def clinical_attachment_viewer(request, pk):
    """
    Display an active image or PDF attachment inside the protected
    professional clinical investigation viewer.

    The original file remains protected behind the authenticated
    clinical_attachment_preview route.
    """

    attachment = get_object_or_404(
        get_clinical_attachment_queryset(),
        pk=pk,
    )

    if not (
        attachment.is_image
        or attachment.is_pdf
    ):
        messages.info(
            request,
            (
                "This document type cannot be displayed in the "
                "investigation viewer. Download it instead."
            ),
        )

        return redirect(
            "clinical_attachment_download",
            pk=attachment.pk,
        )

    # ---------------------------------------------------------
    # NAVIGATION SCOPE
    #
    # Default: other active viewable files belonging to this visit.
    # Optional: ?scope=patient shows files across all patient visits.
    # ---------------------------------------------------------

    requested_scope = (
        request.GET.get("scope", "visit")
        .strip()
        .lower()
    )

    if requested_scope not in {
        "visit",
        "patient",
    }:
        requested_scope = "visit"

    viewable_attachments = (
        get_clinical_attachment_queryset()
        .filter(
            Q(file_extension__iexact=".jpg")
            | Q(file_extension__iexact=".jpeg")
            | Q(file_extension__iexact=".png")
            | Q(file_extension__iexact=".pdf")
        )
    )

    if requested_scope == "patient":
        viewable_attachments = viewable_attachments.filter(
            patient=attachment.patient,
        )
    else:
        viewable_attachments = viewable_attachments.filter(
            visit=attachment.visit,
        )

    viewable_attachments = list(
        viewable_attachments.order_by(
            "investigation_date",
            "uploaded_at",
            "pk",
        )
    )

    previous_attachment = None
    next_attachment = None
    current_position = 1
    total_viewable = len(viewable_attachments)

    for index, item in enumerate(viewable_attachments):
        if item.pk != attachment.pk:
            continue

        current_position = index + 1

        if index > 0:
            previous_attachment = viewable_attachments[
                index - 1
            ]

        if index < total_viewable - 1:
            next_attachment = viewable_attachments[
                index + 1
            ]

        break

    log_activity(
        request,
        AuditLog.ActionType.PRINT,
        "Clinical Attachments",
        (
            f"Opened professional investigation viewer for "
            f"'{attachment.title}' belonging to patient "
            f"{attachment.patient.file_number}."
        ),
        object_id=attachment.pk,
        object_repr=attachment.title,
    )

    return render(
        request,
        (
            "patients/clinical_attachments/"
            "attachment_viewer.html"
        ),
        {
            "attachment": attachment,
            "patient": attachment.patient,
            "visit": attachment.visit,
            "previous_attachment": previous_attachment,
            "next_attachment": next_attachment,
            "current_position": current_position,
            "total_viewable": total_viewable,
            "viewer_scope": requested_scope,
            "can_review_attachment": (
                user_can_review_clinical_attachments(
                    request.user
                )
            ),
            "can_manage_attachment": (
                user_can_manage_clinical_attachments(
                    request.user
                )
            ),
        },
    )


@login_required
@clinical_staff_required
def clinical_attachment_update(request, pk):
    attachment = get_object_or_404(
        get_clinical_attachment_queryset(
            include_inactive=True,
        ),
        pk=pk,
    )

    ensure_attachment_management_permission(request.user)

    old_file_name = (
        attachment.attachment_file.name
        if attachment.attachment_file
        else ""
    )

    form = ClinicalAttachmentForm(
        request.POST or None,
        request.FILES or None,
        instance=attachment,
        visit=attachment.visit,
        request_user=request.user,
    )

    if request.method == "POST":
        if form.is_valid():
            with transaction.atomic():
                updated_attachment = form.save()

                new_file_name = (
                    updated_attachment.attachment_file.name
                    if updated_attachment.attachment_file
                    else ""
                )

                # Delete the previous stored file only after a successful
                # replacement has been saved.
                if (
                    old_file_name
                    and new_file_name
                    and old_file_name != new_file_name
                ):
                    storage = (
                        updated_attachment
                        .attachment_file
                        .storage
                    )

                    if storage.exists(old_file_name):
                        storage.delete(old_file_name)

                log_activity(
                    request,
                    AuditLog.ActionType.UPDATE,
                    "Clinical Attachments",
                    (
                        f"Updated clinical attachment "
                        f"'{updated_attachment.title}' for visit "
                        f"{updated_attachment.visit.visit_number}."
                    ),
                    object_id=updated_attachment.pk,
                    object_repr=updated_attachment.title,
                )

            messages.success(
                request,
                "Clinical attachment updated successfully.",
            )

            return redirect(
                "clinical_attachment_detail",
                pk=updated_attachment.pk,
            )

        messages.error(
            request,
            "Please correct the attachment errors below.",
        )

    return render(
        request,
        "patients/clinical_attachments/attachment_form.html",
        {
            "form": form,
            "visit": attachment.visit,
            "patient": attachment.patient,
            "attachment": attachment,
            "title": "Update Clinical Attachment",
            "button_text": "Save Attachment Changes",
        },
    )


@login_required
@clinical_staff_required
def clinical_attachment_preview(request, pk):
    attachment = get_object_or_404(
        get_clinical_attachment_queryset(),
        pk=pk,
    )

    if not (
        attachment.is_image
        or attachment.is_pdf
    ):
        messages.info(
            request,
            "This file type cannot be displayed inside the browser. "
            "It will be downloaded instead.",
        )

        return redirect(
            "clinical_attachment_download",
            pk=attachment.pk,
        )

    log_activity(
        request,
        AuditLog.ActionType.PRINT,
        "Clinical Attachments",
        (
            f"Viewed clinical attachment "
            f"'{attachment.title}' for patient "
            f"{attachment.patient.file_number}."
        ),
        object_id=attachment.pk,
        object_repr=attachment.title,
    )

    return build_attachment_file_response(
        attachment,
        as_attachment=False,
    )


@login_required
@clinical_staff_required
def clinical_attachment_download(request, pk):
    attachment = get_object_or_404(
        get_clinical_attachment_queryset(),
        pk=pk,
    )

    log_activity(
        request,
        AuditLog.ActionType.PRINT,
        "Clinical Attachments",
        (
            f"Downloaded clinical attachment "
            f"'{attachment.title}' for patient "
            f"{attachment.patient.file_number}."
        ),
        object_id=attachment.pk,
        object_repr=attachment.title,
    )

    return build_attachment_file_response(
        attachment,
        as_attachment=True,
    )


@login_required
@clinical_staff_required
def clinical_attachment_review(request, pk):
    attachment = get_object_or_404(
        get_clinical_attachment_queryset(),
        pk=pk,
    )

    ensure_attachment_review_permission(request.user)

    form = ClinicalAttachmentReviewForm(
        request.POST or None,
        instance=attachment,
        request_user=request.user,
    )

    if request.method == "POST":
        if form.is_valid():
            with transaction.atomic():
                reviewed_attachment = form.save()

                log_activity(
                    request,
                    AuditLog.ActionType.UPDATE,
                    "Clinical Attachments",
                    (
                        f"Updated review status for attachment "
                        f"'{reviewed_attachment.title}' to "
                        f"{reviewed_attachment.get_review_status_display()}."
                    ),
                    object_id=reviewed_attachment.pk,
                    object_repr=reviewed_attachment.title,
                )

            messages.success(
                request,
                "Clinical attachment review saved successfully.",
            )

            return redirect(
                "clinical_attachment_detail",
                pk=reviewed_attachment.pk,
            )

        messages.error(
            request,
            "Please correct the clinical review errors below.",
        )

    return render(
        request,
        "patients/clinical_attachments/attachment_review_form.html",
        {
            "form": form,
            "attachment": attachment,
            "visit": attachment.visit,
            "patient": attachment.patient,
            "title": "Review Clinical Attachment",
            "button_text": "Save Clinical Review",
        },
    )


@login_required
@clinical_staff_required
@require_POST
def clinical_attachment_deactivate(request, pk):
    attachment = get_object_or_404(
        get_clinical_attachment_queryset(),
        pk=pk,
    )

    ensure_attachment_management_permission(request.user)

    attachment.is_active = False

    attachment.save(
        update_fields=[
            "is_active",
            "updated_at",
        ]
    )

    log_activity(
        request,
        AuditLog.ActionType.UPDATE,
        "Clinical Attachments",
        (
            f"Deactivated clinical attachment "
            f"'{attachment.title}' for visit "
            f"{attachment.visit.visit_number}."
        ),
        object_id=attachment.pk,
        object_repr=attachment.title,
    )

    messages.success(
        request,
        "Clinical attachment removed from active clinical records. "
        "The audit record and stored file were retained.",
    )

    return redirect(
        "clinical_attachment_detail",
        pk=attachment.pk,
    )


@login_required
@clinical_staff_required
@require_POST
def clinical_attachment_restore(request, pk):
    attachment = get_object_or_404(
        get_clinical_attachment_queryset(
            include_inactive=True,
        ),
        pk=pk,
        is_active=False,
    )

    ensure_attachment_management_permission(request.user)

    attachment.is_active = True

    attachment.save(
        update_fields=[
            "is_active",
            "updated_at",
        ]
    )

    log_activity(
        request,
        AuditLog.ActionType.UPDATE,
        "Clinical Attachments",
        (
            f"Restored clinical attachment "
            f"'{attachment.title}' for visit "
            f"{attachment.visit.visit_number}."
        ),
        object_id=attachment.pk,
        object_repr=attachment.title,
    )

    messages.success(
        request,
        "Clinical attachment restored successfully.",
    )

    return redirect(
        "clinical_attachment_detail",
        pk=attachment.pk,
    )


@login_required
@require_POST
def clinical_attachment_permanent_delete(request, pk):
    if not request.user.is_superuser:
        raise PermissionDenied(
            "Only a system superuser may permanently delete "
            "a clinical attachment."
        )

    attachment = get_object_or_404(
        get_clinical_attachment_queryset(
            include_inactive=True,
        ),
        pk=pk,
    )

    visit_pk = attachment.visit_id
    title = attachment.title
    stored_file_name = (
        attachment.attachment_file.name
        if attachment.attachment_file
        else ""
    )
    storage = (
        attachment.attachment_file.storage
        if attachment.attachment_file
        else None
    )
    attachment_pk = attachment.pk

    with transaction.atomic():
        log_activity(
            request,
            AuditLog.ActionType.UPDATE,
            "Clinical Attachments",
            (
                f"Permanently deleted clinical attachment "
                f"'{title}' from visit "
                f"{attachment.visit.visit_number}."
            ),
            object_id=attachment_pk,
            object_repr=title,
        )

        attachment.delete()

    if (
        storage is not None
        and stored_file_name
        and storage.exists(stored_file_name)
    ):
        storage.delete(stored_file_name)

    messages.success(
        request,
        "The clinical attachment and its stored file were "
        "permanently deleted.",
    )

    return redirect(
        "visit_clinical_attachment_list",
        visit_pk=visit_pk,
    )


@login_required
@clinical_staff_required
def clinical_attachment_compare_select(request, pk):
    """
    Let the clinician choose another active image or PDF belonging
    to the same patient for side-by-side comparison.
    """

    primary_attachment = get_object_or_404(
        get_clinical_attachment_queryset(),
        pk=pk,
    )

    if not (
        primary_attachment.is_image
        or primary_attachment.is_pdf
    ):
        messages.info(
            request,
            (
                "This file type cannot be used in the clinical "
                "comparison viewer."
            ),
        )

        return redirect(
            "clinical_attachment_detail",
            pk=primary_attachment.pk,
        )

    candidates = (
        get_clinical_attachment_queryset()
        .filter(patient=primary_attachment.patient)
        .exclude(pk=primary_attachment.pk)
        .filter(
            Q(file_extension__iexact=".jpg")
            | Q(file_extension__iexact=".jpeg")
            | Q(file_extension__iexact=".png")
            | Q(file_extension__iexact=".pdf")
        )
        .order_by(
            "-investigation_date",
            "-uploaded_at",
            "-pk",
        )
    )

    query = request.GET.get("q", "").strip()
    category = request.GET.get("category", "").strip()
    eye_side = request.GET.get("eye_side", "").strip()

    if query:
        candidates = candidates.filter(
            Q(title__icontains=query)
            | Q(description__icontains=query)
            | Q(original_filename__icontains=query)
            | Q(visit__visit_number__icontains=query)
            | Q(category_other__icontains=query)
        )

    valid_categories = {
        value
        for value, _label
        in ClinicalAttachment.Category.choices
    }

    if category in valid_categories:
        candidates = candidates.filter(category=category)
    else:
        category = ""

    valid_eye_sides = {
        value
        for value, _label
        in ClinicalAttachment.EyeSide.choices
    }

    if eye_side in valid_eye_sides:
        candidates = candidates.filter(eye_side=eye_side)
    else:
        eye_side = ""

    paginator = Paginator(candidates, 20)

    page_obj = paginator.get_page(
        request.GET.get("page")
    )

    return render(
        request,
        (
            "patients/clinical_attachments/"
            "attachment_compare_select.html"
        ),
        {
            "primary_attachment": primary_attachment,
            "patient": primary_attachment.patient,
            "visit": primary_attachment.visit,
            "candidate_attachments": page_obj,
            "page_obj": page_obj,
            "query": query,
            "selected_category": category,
            "selected_eye_side": eye_side,
            "category_choices": (
                ClinicalAttachment.Category.choices
            ),
            "eye_side_choices": (
                ClinicalAttachment.EyeSide.choices
            ),
        },
    )


@login_required
@clinical_staff_required
def clinical_attachment_compare(request, left_pk, right_pk):
    """
    Display two active clinical investigations side by side.

    Both records must belong to the same patient.
    """

    attachment_queryset = (
        get_clinical_attachment_queryset()
    )

    left_attachment = get_object_or_404(
        attachment_queryset,
        pk=left_pk,
    )

    right_attachment = get_object_or_404(
        attachment_queryset,
        pk=right_pk,
    )

    if left_attachment.pk == right_attachment.pk:
        messages.error(
            request,
            "Select two different clinical investigations to compare.",
        )

        return redirect(
            "clinical_attachment_compare_select",
            pk=left_attachment.pk,
        )

    if (
        left_attachment.patient_id
        != right_attachment.patient_id
    ):
        raise PermissionDenied(
            (
                "Clinical investigations belonging to different "
                "patients cannot be compared."
            )
        )

    for attachment in (
        left_attachment,
        right_attachment,
    ):
        if not (
            attachment.is_image
            or attachment.is_pdf
        ):
            messages.error(
                request,
                (
                    "Only image and PDF clinical investigations "
                    "can be compared."
                ),
            )

            return redirect(
                "clinical_attachment_detail",
                pk=attachment.pk,
            )

    log_activity(
        request,
        AuditLog.ActionType.PRINT,
        "Clinical Attachments",
        (
            f"Compared clinical attachments "
            f"'{left_attachment.title}' and "
            f"'{right_attachment.title}' for patient "
            f"{left_attachment.patient.file_number}."
        ),
        object_id=left_attachment.pk,
        object_repr=(
            f"{left_attachment.title} versus "
            f"{right_attachment.title}"
        ),
    )

    return render(
        request,
        (
            "patients/clinical_attachments/"
            "attachment_compare.html"
        ),
        {
            "patient": left_attachment.patient,
            "left_attachment": left_attachment,
            "right_attachment": right_attachment,
            "can_review_attachments": (
                user_can_review_clinical_attachments(
                    request.user
                )
            ),
        },
    )


@login_required
@clinical_staff_required
def clinical_attachment_annotation_workspace(request, pk):
    """
    Display the non-destructive annotation workspace for an active
    clinical image.
    """

    attachment = get_object_or_404(
        get_clinical_attachment_queryset(),
        pk=pk,
    )

    if not attachment.is_image:
        messages.info(
            request,
            (
                "Clinical annotation currently supports JPG, JPEG "
                "and PNG investigation images."
            ),
        )

        return redirect(
            "clinical_attachment_detail",
            pk=attachment.pk,
        )

    annotations = (
        attachment.image_annotations
        .filter(is_active=True)
        .select_related(
            "created_by",
            "updated_by",
        )
        .order_by(
            "-updated_at",
            "-pk",
        )
    )

    annotation_payloads = []

    for annotation in annotations:
        annotation_payloads.append(
            {
                "id": annotation.pk,
                "title": annotation.title,
                "clinicalNote": annotation.clinical_note,
                "status": annotation.status,
                "statusDisplay": annotation.get_status_display(),
                "version": annotation.version,
                "annotationData": annotation.annotation_data,
                "createdBy": timeline_user_name(
                    annotation.created_by
                ),
                "updatedBy": timeline_user_name(
                    annotation.updated_by
                ),
                "createdAt": timezone.localtime(
                    annotation.created_at
                ).strftime("%d %b %Y, %I:%M %p"),
                "updatedAt": timezone.localtime(
                    annotation.updated_at
                ).strftime("%d %b %Y, %I:%M %p"),
                "finalizedAt": (
                    timezone.localtime(
                        annotation.finalized_at
                    ).strftime(
                        "%d %b %Y, %I:%M %p"
                    )
                    if annotation.finalized_at
                    else ""
                ),
            }
        )

    log_activity(
        request,
        AuditLog.ActionType.PRINT,
        "Clinical Image Annotations",
        (
            f"Opened image annotation workspace for "
            f"'{attachment.title}' belonging to patient "
            f"{attachment.patient.file_number}."
        ),
        object_id=attachment.pk,
        object_repr=attachment.title,
    )

    return render(
        request,
        (
            "patients/clinical_attachments/"
            "attachment_annotation.html"
        ),
        {
            "attachment": attachment,
            "patient": attachment.patient,
            "visit": attachment.visit,
            "annotation_payloads": annotation_payloads,
            "can_edit_annotations": (
                user_can_annotate_clinical_attachments(
                    request.user
                )
            ),
        },
    )


@login_required
@clinical_staff_required
@require_POST
def clinical_attachment_annotation_save(request, pk):
    """
    Create or update one annotation layer using validated JSON.
    """

    attachment = get_object_or_404(
        get_clinical_attachment_queryset(),
        pk=pk,
    )

    ensure_annotation_permission(
        request.user
    )

    if not attachment.is_image:
        return JsonResponse(
            {
                "success": False,
                "message": (
                    "Only image investigations can be annotated."
                ),
            },
            status=400,
        )

    try:
        request_payload = json.loads(
            request.body.decode("utf-8")
        )
    except (
        UnicodeDecodeError,
        json.JSONDecodeError,
    ):
        return JsonResponse(
            {
                "success": False,
                "message": "Invalid annotation request.",
            },
            status=400,
        )

    annotation_id = request_payload.get(
        "annotationId"
    )

    title = str(
        request_payload.get(
            "title",
            "Clinical Annotation",
        )
    ).strip()

    clinical_note = str(
        request_payload.get(
            "clinicalNote",
            "",
        )
    ).strip()

    requested_status = str(
        request_payload.get(
            "status",
            ClinicalImageAnnotation
            .AnnotationStatus
            .DRAFT,
        )
    ).upper()

    if not title:
        return JsonResponse(
            {
                "success": False,
                "message": "Enter a title for the annotation.",
            },
            status=400,
        )

    if len(title) > 200:
        return JsonResponse(
            {
                "success": False,
                "message": (
                    "The annotation title must not exceed "
                    "200 characters."
                ),
            },
            status=400,
        )

    if len(clinical_note) > 5000:
        return JsonResponse(
            {
                "success": False,
                "message": (
                    "The clinical annotation note must not exceed "
                    "5,000 characters."
                ),
            },
            status=400,
        )

    valid_statuses = {
        value
        for value, _label
        in ClinicalImageAnnotation
        .AnnotationStatus
        .choices
    }

    if requested_status not in valid_statuses:
        return JsonResponse(
            {
                "success": False,
                "message": "Invalid annotation status.",
            },
            status=400,
        )

    try:
        annotation_data = validate_annotation_payload(
            request_payload.get(
                "annotationData",
                {},
            )
        )
    except ValueError as error:
        return JsonResponse(
            {
                "success": False,
                "message": str(error),
            },
            status=400,
        )

    with transaction.atomic():
        if annotation_id:
            annotation = get_object_or_404(
                ClinicalImageAnnotation.objects.select_related(
                    "attachment",
                ),
                pk=annotation_id,
                attachment=attachment,
                is_active=True,
            )

            annotation.version += 1
            action_type = (
                AuditLog.ActionType.UPDATE
            )
            action_word = "Updated"
        else:
            annotation = ClinicalImageAnnotation(
                attachment=attachment,
                created_by=request.user,
            )

            action_type = (
                AuditLog.ActionType.CREATE
            )
            action_word = "Created"

        annotation.title = title
        annotation.clinical_note = clinical_note
        annotation.annotation_data = annotation_data
        annotation.status = requested_status
        annotation.updated_by = request.user
        annotation.save()

        log_activity(
            request,
            action_type,
            "Clinical Image Annotations",
            (
                f"{action_word} annotation "
                f"'{annotation.title}' for attachment "
                f"'{attachment.title}'."
            ),
            object_id=annotation.pk,
            object_repr=annotation.title,
        )

    return JsonResponse(
        {
            "success": True,
            "message": (
                "Clinical annotation saved successfully."
            ),
            "annotation": {
                "id": annotation.pk,
                "title": annotation.title,
                "clinicalNote": annotation.clinical_note,
                "status": annotation.status,
                "statusDisplay": (
                    annotation.get_status_display()
                ),
                "version": annotation.version,
                "annotationData": annotation.annotation_data,
                "createdBy": timeline_user_name(
                    annotation.created_by
                ),
                "updatedBy": timeline_user_name(
                    annotation.updated_by
                ),
                "updatedAt": timezone.localtime(
                    annotation.updated_at
                ).strftime(
                    "%d %b %Y, %I:%M %p"
                ),
                "finalizedAt": (
                    timezone.localtime(
                        annotation.finalized_at
                    ).strftime(
                        "%d %b %Y, %I:%M %p"
                    )
                    if annotation.finalized_at
                    else ""
                ),
            },
        }
    )


@login_required
@clinical_staff_required
@require_POST
def clinical_image_annotation_deactivate(
    request,
    annotation_pk,
):
    annotation = get_object_or_404(
        ClinicalImageAnnotation.objects.select_related(
            "attachment",
            "attachment__patient",
        ),
        pk=annotation_pk,
        is_active=True,
    )

    ensure_annotation_permission(
        request.user
    )

    annotation.is_active = False
    annotation.updated_by = request.user

    annotation.save(
        update_fields=[
            "is_active",
            "updated_by",
            "updated_at",
        ]
    )

    log_activity(
        request,
        AuditLog.ActionType.UPDATE,
        "Clinical Image Annotations",
        (
            f"Deactivated annotation "
            f"'{annotation.title}' for attachment "
            f"'{annotation.attachment.title}'."
        ),
        object_id=annotation.pk,
        object_repr=annotation.title,
    )

    return JsonResponse(
        {
            "success": True,
            "message": (
                "Clinical annotation removed from active records."
            ),
        }
    )


@login_required
@clinical_staff_required
def diagnosis_treatment_create(request, visit_pk):
    visit = get_object_or_404(PatientVisit.objects.select_related("patient"), pk=visit_pk)

    if hasattr(visit, "diagnosis_treatment"):
        messages.info(request, "This visit already has a diagnosis/treatment record. You can edit it instead.")
        return redirect("diagnosis_treatment_update", pk=visit.diagnosis_treatment.pk)

    form = DiagnosisTreatmentForm(request.POST or None)

    if request.method == "POST":
        if form.is_valid():
            record = form.save(commit=False)
            record.visit = visit
            record.created_by = request.user
            record.save()

            visit.status = PatientVisit.VisitStatus.WITH_DOCTOR
            visit.save(update_fields=["status"])

            messages.success(request, "Diagnosis and treatment record saved successfully.")
            return redirect("visit_detail", pk=visit.pk)

        messages.error(request, "Please correct the errors below.")

    return render(request, "patients/diagnosis_treatment_form.html", {
        "form": form,
        "visit": visit,
        "title": "Create Diagnosis and Treatment Record",
        "button_text": "Save Diagnosis/Treatment",
    })


@login_required
@clinical_staff_required
def diagnosis_treatment_update(request, pk):
    record = get_object_or_404(
        DiagnosisTreatment.objects.select_related("visit", "visit__patient"),
        pk=pk,
    )

    form = DiagnosisTreatmentForm(request.POST or None, instance=record)

    if request.method == "POST":
        if form.is_valid():
            form.save()
            messages.success(request, "Diagnosis and treatment record updated successfully.")
            return redirect("visit_detail", pk=record.visit.pk)

        messages.error(request, "Please correct the errors below.")

    return render(request, "patients/diagnosis_treatment_form.html", {
        "form": form,
        "visit": record.visit,
        "record": record,
        "title": "Update Diagnosis and Treatment Record",
        "button_text": "Save Changes",
    })


@login_required
@clinical_staff_required
def prescription_create(request, visit_pk):
    visit = get_object_or_404(PatientVisit.objects.select_related("patient"), pk=visit_pk)
    form = PrescriptionForm(request.POST or None)

    if request.method == "POST":
        if form.is_valid():
            prescription = form.save(commit=False)
            prescription.visit = visit
            prescription.prescribed_by = request.user
            prescription.save()

            messages.success(request, "Prescription added successfully.")
            return redirect("visit_detail", pk=visit.pk)

        messages.error(request, "Please correct the errors below.")

    return render(request, "patients/prescription_form.html", {
        "form": form,
        "visit": visit,
        "title": "Add Prescription",
        "button_text": "Save Prescription",
    })


@login_required
@clinical_staff_required
def prescription_update(request, pk):
    prescription = get_object_or_404(
        Prescription.objects.select_related("visit", "visit__patient"),
        pk=pk,
    )

    form = PrescriptionForm(request.POST or None, instance=prescription)

    if request.method == "POST":
        if form.is_valid():
            form.save()
            messages.success(request, "Prescription updated successfully.")
            return redirect("visit_detail", pk=prescription.visit.pk)

        messages.error(request, "Please correct the errors below.")

    return render(request, "patients/prescription_form.html", {
        "form": form,
        "visit": prescription.visit,
        "prescription": prescription,
        "title": "Update Prescription",
        "button_text": "Save Changes",
    })



@login_required
def surgery_list(request):
    query = request.GET.get("q", "").strip()
    status = request.GET.get("status", "").strip()

    surgeries = SurgeryProcedure.objects.select_related(
        "patient",
        "visit",
        "surgeon",
        "assistant",
        "created_by",
    ).all()

    if query:
        surgeries = surgeries.filter(
            Q(procedure_number__icontains=query)
            | Q(patient__file_number__icontains=query)
            | Q(patient__full_name__icontains=query)
            | Q(patient__phone_number__icontains=query)
            | Q(procedure_name__icontains=query)
            | Q(procedure_type__icontains=query)
            | Q(pre_op_diagnosis__icontains=query)
            | Q(post_op_diagnosis__icontains=query)
        )

    if status:
        surgeries = surgeries.filter(status=status)

    return render(request, "patients/surgery_list.html", {
        "surgeries": surgeries,
        "query": query,
        "status": status,
        "status_choices": SurgeryProcedure.ProcedureStatus.choices,
    })


@login_required
@clinical_staff_required
def surgery_create(request):
    form = SurgeryProcedureForm(request.POST or None)

    if request.method == "POST":
        if form.is_valid():
            surgery = form.save(commit=False)
            surgery.created_by = request.user
            surgery.save()

            messages.success(
                request,
                f"Surgery/procedure record created successfully. Procedure Number: {surgery.procedure_number}"
            )
            return redirect("surgery_detail", pk=surgery.pk)

        messages.error(request, "Please correct the errors below.")

    return render(request, "patients/surgery_form.html", {
        "form": form,
        "title": "Create Surgery / Procedure Record",
        "button_text": "Save Surgery / Procedure",
    })


@login_required
@clinical_staff_required
def patient_surgery_create(request, patient_pk):
    patient = get_object_or_404(Patient, pk=patient_pk)
    form = SurgeryProcedureQuickForm(request.POST or None)

    if request.method == "POST":
        if form.is_valid():
            surgery = form.save(commit=False)
            surgery.patient = patient
            surgery.created_by = request.user
            surgery.save()

            messages.success(
                request,
                f"Surgery/procedure record created successfully. Procedure Number: {surgery.procedure_number}"
            )
            return redirect("surgery_detail", pk=surgery.pk)

        messages.error(request, "Please correct the errors below.")

    return render(request, "patients/patient_surgery_form.html", {
        "form": form,
        "patient": patient,
        "title": "Create Surgery / Procedure for Patient",
        "button_text": "Save Surgery / Procedure",
    })


@login_required
@clinical_staff_required
def visit_surgery_create(request, visit_pk):
    visit = get_object_or_404(PatientVisit.objects.select_related("patient"), pk=visit_pk)
    form = SurgeryProcedureQuickForm(request.POST or None)

    if request.method == "POST":
        if form.is_valid():
            surgery = form.save(commit=False)
            surgery.patient = visit.patient
            surgery.visit = visit
            surgery.created_by = request.user
            surgery.save()

            messages.success(
                request,
                f"Surgery/procedure record created successfully. Procedure Number: {surgery.procedure_number}"
            )
            return redirect("surgery_detail", pk=surgery.pk)

        messages.error(request, "Please correct the errors below.")

    return render(request, "patients/visit_surgery_form.html", {
        "form": form,
        "visit": visit,
        "patient": visit.patient,
        "title": "Create Surgery / Procedure for Visit",
        "button_text": "Save Surgery / Procedure",
    })


@login_required
def surgery_detail(request, pk):
    surgery = get_object_or_404(
        SurgeryProcedure.objects.select_related(
            "patient",
            "visit",
            "surgeon",
            "assistant",
            "created_by",
        ),
        pk=pk,
    )

    return render(request, "patients/surgery_detail.html", {
        "surgery": surgery,
    })


@login_required
@clinical_staff_required
def surgery_update(request, pk):
    surgery = get_object_or_404(SurgeryProcedure, pk=pk)
    form = SurgeryProcedureForm(request.POST or None, instance=surgery)

    if request.method == "POST":
        if form.is_valid():
            form.save()
            messages.success(request, "Surgery/procedure record updated successfully.")
            return redirect("surgery_detail", pk=surgery.pk)

        messages.error(request, "Please correct the errors below.")

    return render(request, "patients/surgery_form.html", {
        "form": form,
        "surgery": surgery,
        "title": "Update Surgery / Procedure Record",
        "button_text": "Save Changes",
    })


@login_required
def patient_surgery_history(request, patient_pk):
    patient = get_object_or_404(Patient, pk=patient_pk)
    surgeries = patient.surgeries.select_related("visit", "surgeon", "assistant", "created_by").all()

    return render(request, "patients/patient_surgery_history.html", {
        "patient": patient,
        "surgeries": surgeries,
    })



@login_required
def appointment_list(request):
    query = request.GET.get("q", "").strip()
    status = request.GET.get("status", "").strip()
    date = request.GET.get("date", "").strip()

    appointments = Appointment.objects.select_related(
        "patient",
        "assigned_to",
        "created_by",
    ).all()

    if query:
        appointments = appointments.filter(
            Q(appointment_number__icontains=query)
            | Q(patient__file_number__icontains=query)
            | Q(patient__full_name__icontains=query)
            | Q(patient__phone_number__icontains=query)
            | Q(reason__icontains=query)
        )

    if status:
        appointments = appointments.filter(status=status)

    if date:
        appointments = appointments.filter(appointment_date=date)

    return render(request, "patients/appointment_list.html", {
        "appointments": appointments,
        "query": query,
        "status": status,
        "date": date,
        "status_choices": Appointment.AppointmentStatus.choices,
    })


@login_required
@records_staff_required
def appointment_create(request):
    form = AppointmentForm(request.POST or None)

    if request.method == "POST":
        if form.is_valid():
            appointment = form.save(commit=False)
            appointment.created_by = request.user
            appointment.save()

            create_notification(
                title="New Appointment Created",
                message=f"Appointment {appointment.appointment_number} created for {appointment.patient.full_name}.",
                notification_type=Notification.NotificationType.APPOINTMENT,
                recipient=appointment.assigned_to,
                is_global=False,
                link=f"/patients/appointments/{appointment.pk}/",
            )

            messages.success(
                request,
                f"Appointment created successfully. Appointment Number: {appointment.appointment_number}"
            )
            return redirect("appointment_detail", pk=appointment.pk)

        messages.error(request, "Please correct the errors below.")

    return render(request, "patients/appointment_form.html", {
        "form": form,
        "title": "Create Appointment",
        "button_text": "Save Appointment",
    })


@login_required
@records_staff_required
def patient_appointment_create(request, patient_pk):
    patient = get_object_or_404(Patient, pk=patient_pk)
    form = AppointmentQuickForm(request.POST or None)

    if request.method == "POST":
        if form.is_valid():
            appointment = form.save(commit=False)
            appointment.patient = patient
            appointment.created_by = request.user
            appointment.save()

            create_notification(
                title="New Appointment Created",
                message=f"Appointment {appointment.appointment_number} created for {appointment.patient.full_name}.",
                notification_type=Notification.NotificationType.APPOINTMENT,
                recipient=appointment.assigned_to,
                is_global=False,
                link=f"/patients/appointments/{appointment.pk}/",
            )

            messages.success(
                request,
                f"Appointment created successfully. Appointment Number: {appointment.appointment_number}"
            )
            return redirect("appointment_detail", pk=appointment.pk)

        messages.error(request, "Please correct the errors below.")

    return render(request, "patients/patient_appointment_form.html", {
        "form": form,
        "patient": patient,
        "title": "Create Appointment for Patient",
        "button_text": "Save Appointment",
    })


@login_required
def appointment_detail(request, pk):
    appointment = get_object_or_404(
        Appointment.objects.select_related("patient", "assigned_to", "created_by"),
        pk=pk,
    )

    return render(request, "patients/appointment_detail.html", {
        "appointment": appointment,
    })


@login_required
@records_staff_required
def appointment_update(request, pk):
    appointment = get_object_or_404(Appointment, pk=pk)
    form = AppointmentForm(request.POST or None, instance=appointment)

    if request.method == "POST":
        if form.is_valid():
            form.save()
            messages.success(request, "Appointment updated successfully.")
            return redirect("appointment_detail", pk=appointment.pk)

        messages.error(request, "Please correct the errors below.")

    return render(request, "patients/appointment_form.html", {
        "form": form,
        "appointment": appointment,
        "title": "Update Appointment",
        "button_text": "Save Changes",
    })


@login_required
def patient_appointment_history(request, patient_pk):
    patient = get_object_or_404(Patient, pk=patient_pk)
    appointments = patient.appointments.select_related("assigned_to", "created_by").all()

    return render(request, "patients/patient_appointment_history.html", {
        "patient": patient,
        "appointments": appointments,
    })



@login_required
def appointment_calendar(request):
    today = timezone.localdate()
    selected_date = request.GET.get("date", "")

    if selected_date:
        calendar_date = selected_date
    else:
        calendar_date = today

    appointments = Appointment.objects.select_related(
        "patient",
        "assigned_to",
        "created_by",
    ).filter(
        appointment_date=calendar_date
    )

    return render(request, "patients/appointment_calendar.html", {
        "appointments": appointments,
        "calendar_date": calendar_date,
        "today": today,
        "status_choices": Appointment.AppointmentStatus.choices,
    })


@login_required
def clinic_queue(request):
    today = timezone.localdate()

    visits = PatientVisit.objects.select_related(
        "patient",
        "created_by",
    ).filter(
        visit_date__date=today
    )

    appointments = Appointment.objects.select_related(
        "patient",
        "assigned_to",
    ).filter(
        appointment_date=today
    )

    return render(request, "patients/clinic_queue.html", {
        "visits": visits,
        "appointments": appointments,
        "today": today,
    })


@login_required
@clinical_staff_required
def doctor_worklist(request):
    today = timezone.localdate()

    visits = PatientVisit.objects.select_related(
        "patient",
        "created_by",
    ).filter(
        visit_date__date=today
    ).exclude(
        status=PatientVisit.VisitStatus.COMPLETED
    )

    appointments = Appointment.objects.select_related(
        "patient",
        "assigned_to",
    ).filter(
        appointment_date=today
    )

    if not request.user.is_superuser:
        appointments = appointments.filter(
            Q(assigned_to=request.user) | Q(assigned_to__isnull=True)
        )

    return render(request, "patients/doctor_worklist.html", {
        "visits": visits,
        "appointments": appointments,
        "today": today,
    })


@login_required
@records_staff_required
def appointment_mark_arrived(request, pk):
    appointment = get_object_or_404(Appointment, pk=pk)
    appointment.status = Appointment.AppointmentStatus.ARRIVED
    appointment.save(update_fields=["status", "updated_at"])

    messages.success(request, "Appointment marked as arrived.")
    return redirect("clinic_queue")


@login_required
@records_staff_required
def appointment_mark_completed(request, pk):
    appointment = get_object_or_404(Appointment, pk=pk)
    appointment.status = Appointment.AppointmentStatus.COMPLETED
    appointment.save(update_fields=["status", "updated_at"])

    messages.success(request, "Appointment marked as completed.")
    return redirect("clinic_queue")


@login_required
@clinical_staff_required
def visit_mark_completed(request, pk):
    visit = get_object_or_404(PatientVisit, pk=pk)
    visit.status = PatientVisit.VisitStatus.COMPLETED
    visit.save(update_fields=["status", "updated_at"])

    messages.success(request, "Visit marked as completed.")
    return redirect("doctor_worklist")




@login_required
@finance_staff_required
def bill_list(request):
    query = request.GET.get("q", "").strip()
    status = request.GET.get("status", "").strip()

    bills = Bill.objects.select_related(
        "patient",
        "visit",
        "surgery",
        "appointment",
        "created_by",
    ).all()

    filtered_bills = bills

    if query:
        filtered_bills = filtered_bills.filter(
            Q(bill_number__icontains=query)
            | Q(patient__file_number__icontains=query)
            | Q(patient__full_name__icontains=query)
            | Q(patient__phone_number__icontains=query)
            | Q(bill_title__icontains=query)
        )

    if status:
        filtered_bills = filtered_bills.filter(status=status)

    total_bills = bills.count()
    unpaid_bills = bills.exclude(status=Bill.BillStatus.PAID).exclude(status=Bill.BillStatus.CANCELLED).count()

    return render(request, "patients/bill_list.html", {
        "bills": filtered_bills,
        "query": query,
        "status": status,
        "status_choices": Bill.BillStatus.choices,
        "total_bills": total_bills,
        "unpaid_bills": unpaid_bills,
    })


@login_required
@finance_staff_required
def bill_create(request):
    form = BillForm(request.POST or None)

    if request.method == "POST":
        if form.is_valid():
            bill = form.save(commit=False)
            bill.created_by = request.user
            bill.save()

            messages.success(request, f"Bill created successfully. Bill Number: {bill.bill_number}")
            return redirect("bill_detail", pk=bill.pk)

        messages.error(request, "Please correct the errors below.")

    return render(request, "patients/bill_form.html", {
        "form": form,
        "title": "Create Bill",
        "button_text": "Save Bill",
    })


@login_required
@finance_staff_required
def patient_bill_create(request, patient_pk):
    patient = get_object_or_404(Patient, pk=patient_pk)
    form = BillQuickForm(request.POST or None)

    if request.method == "POST":
        if form.is_valid():
            bill = form.save(commit=False)
            bill.patient = patient
            bill.created_by = request.user
            bill.save()

            messages.success(request, f"Bill created successfully. Bill Number: {bill.bill_number}")
            return redirect("bill_detail", pk=bill.pk)

        messages.error(request, "Please correct the errors below.")

    return render(request, "patients/patient_bill_form.html", {
        "form": form,
        "patient": patient,
        "title": "Create Bill for Patient",
        "button_text": "Save Bill",
    })


@login_required
@finance_staff_required
def bill_detail(request, pk):
    bill = get_object_or_404(
        Bill.objects.select_related(
            "patient",
            "visit",
            "surgery",
            "appointment",
            "created_by",
        ),
        pk=pk,
    )

    payments = bill.payments.select_related("received_by").all()

    context = {
        "bill": bill,
        "payments": payments,
        "can_pay": bill.balance > 0 and bill.status != Bill.BillStatus.CANCELLED,
    }

    return render(request, "patients/bill_detail.html", context)


@login_required
@finance_staff_required
def bill_update(request, pk):
    bill = get_object_or_404(Bill, pk=pk)
    form = BillForm(request.POST or None, instance=bill)

    if request.method == "POST":
        if form.is_valid():
            form.save()
            messages.success(request, "Bill updated successfully.")
            return redirect("bill_detail", pk=bill.pk)

        messages.error(request, "Please correct the errors below.")

    return render(request, "patients/bill_form.html", {
        "form": form,
        "bill": bill,
        "title": "Update Bill",
        "button_text": "Save Changes",
    })


@login_required
@finance_staff_required
def payment_create(request, bill_pk):
    bill = get_object_or_404(Bill.objects.select_related("patient"), pk=bill_pk)
    form = PaymentForm(request.POST or None, bill=bill)

    if request.method == "POST":
        if form.is_valid():
            payment = form.save(commit=False)
            payment.bill = bill
            payment.received_by = request.user
            payment.save()

            log_activity(
                request,
                AuditLog.ActionType.CREATE,
                "Payments",
                f"Recorded payment of ₦{payment.amount} for bill {bill.bill_number}.",
                object_id=payment.pk,
                object_repr=payment.receipt_number,
            )

            messages.success(request, f"Payment recorded successfully. Receipt Number: {payment.receipt_number}")
            return redirect("bill_detail", pk=bill.pk)

        messages.error(request, "Please correct the errors below.")

    return render(request, "patients/payment_form.html", {
        "form": form,
        "bill": bill,
        "title": "Record Payment",
        "button_text": "Save Payment",
    })

    

@login_required
@finance_staff_required
def patient_bill_history(request, patient_pk):
    patient = get_object_or_404(Patient, pk=patient_pk)
    bills = patient.bills.select_related("created_by").all()

    return render(request, "patients/patient_bill_history.html", {
        "patient": patient,
        "bills": bills,
    })



@login_required
def global_search(request):
    query = request.GET.get("q", "").strip()

    patients = Patient.objects.none()
    visits = PatientVisit.objects.none()
    appointments = Appointment.objects.none()
    surgeries = SurgeryProcedure.objects.none()
    bills = Bill.objects.none()
    prescriptions = Prescription.objects.none()
    diagnoses = DiagnosisTreatment.objects.none()
    consultations = Consultation.objects.none()

    if query:
        patients = Patient.objects.filter(
            Q(file_number__icontains=query)
            | Q(full_name__icontains=query)
            | Q(phone_number__icontains=query)
            | Q(address__icontains=query)
            | Q(diagnosis__icontains=query)
            | Q(treatment__icontains=query)
            | Q(eye_complaint__icontains=query)
        )[:20]

        visits = PatientVisit.objects.select_related("patient").filter(
            Q(visit_number__icontains=query)
            | Q(patient__file_number__icontains=query)
            | Q(patient__full_name__icontains=query)
            | Q(chief_complaint__icontains=query)
            | Q(brief_history__icontains=query)
        )[:20]

        appointments = Appointment.objects.select_related("patient").filter(
            Q(appointment_number__icontains=query)
            | Q(patient__file_number__icontains=query)
            | Q(patient__full_name__icontains=query)
            | Q(patient__phone_number__icontains=query)
            | Q(reason__icontains=query)
        )[:20]

        surgeries = SurgeryProcedure.objects.select_related("patient").filter(
            Q(procedure_number__icontains=query)
            | Q(patient__file_number__icontains=query)
            | Q(patient__full_name__icontains=query)
            | Q(procedure_name__icontains=query)
            | Q(procedure_type__icontains=query)
            | Q(pre_op_diagnosis__icontains=query)
            | Q(post_op_diagnosis__icontains=query)
        )[:20]

        bills = Bill.objects.select_related("patient").filter(
            Q(bill_number__icontains=query)
            | Q(patient__file_number__icontains=query)
            | Q(patient__full_name__icontains=query)
            | Q(patient__phone_number__icontains=query)
            | Q(bill_title__icontains=query)
        )[:20]

        prescriptions = Prescription.objects.select_related("visit", "visit__patient").filter(
            Q(drug_name__icontains=query)
            | Q(dosage__icontains=query)
            | Q(frequency__icontains=query)
            | Q(visit__visit_number__icontains=query)
            | Q(visit__patient__file_number__icontains=query)
            | Q(visit__patient__full_name__icontains=query)
        )[:20]

        diagnoses = DiagnosisTreatment.objects.select_related("visit", "visit__patient").filter(
            Q(primary_diagnosis__icontains=query)
            | Q(secondary_diagnosis__icontains=query)
            | Q(differential_diagnosis__icontains=query)
            | Q(treatment_plan__icontains=query)
            | Q(visit__visit_number__icontains=query)
            | Q(visit__patient__file_number__icontains=query)
            | Q(visit__patient__full_name__icontains=query)
        )[:20]

        consultations = Consultation.objects.select_related("visit", "visit__patient").filter(
            Q(presenting_complaint__icontains=query)
            | Q(provisional_diagnosis__icontains=query)
            | Q(final_diagnosis__icontains=query)
            | Q(treatment_plan__icontains=query)
            | Q(visit__visit_number__icontains=query)
            | Q(visit__patient__file_number__icontains=query)
            | Q(visit__patient__full_name__icontains=query)
        )[:20]

    return render(request, "patients/global_search.html", {
        "query": query,
        "patients": patients,
        "visits": visits,
        "appointments": appointments,
        "surgeries": surgeries,
        "bills": bills,
        "prescriptions": prescriptions,
        "diagnoses": diagnoses,
        "consultations": consultations,
    })



@login_required
@records_staff_required
def patient_excel_import(request):
    if request.method == "POST":
        excel_file = request.FILES.get("excel_file")

        if not excel_file:
            messages.error(request, "Please select an Excel file.")
            return redirect("patient_excel_import")

        if not excel_file.name.endswith(".xlsx"):
            messages.error(request, "Only .xlsx files are allowed.")
            return redirect("patient_excel_import")

        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active

        imported_count = 0
        skipped_count = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):
            full_name = row[0]
            gender = row[1]
            phone_number = row[2]
            address = row[3]
            occupation = row[4]
            next_of_kin_name = row[5]
            next_of_kin_phone = row[6]
            medical_history = row[7]
            allergy_history = row[8]
            eye_complaint = row[9]
            diagnosis = row[10]
            treatment = row[11]
            payment_status = row[12]
            notes = row[13]

            if not full_name:
                skipped_count += 1
                continue

            duplicate = Patient.objects.filter(
                full_name__iexact=str(full_name).strip(),
                phone_number__iexact=str(phone_number).strip() if phone_number else "",
            ).exists()

            if duplicate:
                skipped_count += 1
                continue

            Patient.objects.create(
                full_name=str(full_name).strip(),
                gender=str(gender).upper() if gender else Patient.Gender.MALE,
                phone_number=str(phone_number).strip() if phone_number else "",
                address=address or "",
                occupation=occupation or "",
                next_of_kin_name=next_of_kin_name or "",
                next_of_kin_phone=str(next_of_kin_phone).strip() if next_of_kin_phone else "",
                medical_history=medical_history or "",
                allergy_history=allergy_history or "",
                eye_complaint=eye_complaint or "",
                diagnosis=diagnosis or "",
                treatment=treatment or "",
                payment_status=payment_status if payment_status else Patient.PaymentStatus.NOT_APPLICABLE,
                notes=notes or "",
                registered_by=request.user,
            )

            imported_count += 1

        log_activity(
            request,
            AuditLog.ActionType.IMPORT,
            "Patients",
            f"Imported Excel patient records. Imported: {imported_count}, Skipped: {skipped_count}.",
        )

        messages.success(
            request,
            f"Excel import completed. Imported: {imported_count}, Skipped: {skipped_count}"
        )
        return redirect("patient_list")

    return render(request, "patients/patient_excel_import.html")



@login_required
def patient_excel_export(request):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Patients"

    headers = [
        "File Number",
        "Full Name",
        "Gender",
        "Age",
        "Date of Birth",
        "Phone Number",
        "Address",
        "Occupation",
        "Next of Kin",
        "Next of Kin Phone",
        "Medical History",
        "Allergy History",
        "Eye Complaint",
        "Diagnosis",
        "Treatment",
        "Payment Status",
        "Registration Date",
        "Registered By",
        "Notes",
    ]

    sheet.append(headers)

    patients = Patient.objects.select_related("registered_by").all()

    for patient in patients:
        sheet.append([
            patient.file_number,
            patient.full_name,
            patient.get_gender_display(),
            patient.display_age or "",
            patient.date_of_birth.strftime("%Y-%m-%d") if patient.date_of_birth else "",
            patient.phone_number,
            patient.address,
            patient.occupation,
            patient.next_of_kin_name,
            patient.next_of_kin_phone,
            patient.medical_history,
            patient.allergy_history,
            patient.eye_complaint,
            patient.diagnosis,
            patient.treatment,
            patient.get_payment_status_display(),
            patient.registration_date.strftime("%Y-%m-%d %H:%M"),
            str(patient.registered_by) if patient.registered_by else "",
            patient.notes,
        ])

    for col_num, column_title in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        sheet.column_dimensions[column_letter].width = 22

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    filename = f"capital_eye_patients_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    workbook.save(response)

    log_activity(
        request,
        AuditLog.ActionType.EXPORT,
        "Patients",
        "Exported patient records to Excel.",
    )

    return response



@login_required
@records_staff_required
def patient_excel_template(request):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Patient Import Template"

    headers = [
        "Full Name",
        "Gender",
        "Phone Number",
        "Address",
        "Occupation",
        "Next of Kin Name",
        "Next of Kin Phone",
        "Medical History",
        "Allergy History",
        "Eye Complaint",
        "Diagnosis",
        "Treatment",
        "Payment Status",
        "Notes",
    ]

    sheet.append(headers)

    sample = [
        "John Musa",
        "MALE",
        "08000000000",
        "Kano, Nigeria",
        "Trader",
        "Aisha Musa",
        "08111111111",
        "No known chronic illness",
        "No known allergy",
        "Blurred vision",
        "Cataract",
        "Eye drops and follow-up",
        "NOT_APPLICABLE",
        "Imported sample row. Delete before real import.",
    ]

    sheet.append(sample)

    for col_num, column_title in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        sheet.column_dimensions[column_letter].width = 25

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = 'attachment; filename="patient_import_template.xlsx"'

    workbook.save(response)
    return response



@login_required
def patient_card_pdf(request, pk):
    patient = get_object_or_404(Patient, pk=pk)

    log_activity(
        request,
        AuditLog.ActionType.PRINT,
        "PDF Reports",
        f"Generated PDF for patient {patient.file_number}.",
        object_id=patient.pk,
        object_repr=patient.file_number,
    )

    return render_to_pdf(
        "patients/pdf/patient_card_pdf.html",
        {"patient": patient},
        f"{patient.file_number}_patient_card.pdf",
    )


@login_required
def patient_report_pdf(request, pk):
    patient = get_object_or_404(Patient, pk=pk)

    visits = patient.visits.all()
    appointments = patient.appointments.all()
    surgeries = patient.surgeries.all()
    bills = patient.bills.all()

    log_activity(
        request,
        AuditLog.ActionType.PRINT,
        "PDF Reports",
        f"Generated PDF for patient {patient.file_number}.",
        object_id=patient.pk,
        object_repr=patient.file_number,
    )

    return render_to_pdf(
        "patients/pdf/patient_report_pdf.html",
        {
            "patient": patient,
            "visits": visits,
            "appointments": appointments,
            "surgeries": surgeries,
            "bills": bills,
        },
        f"{patient.file_number}_full_report.pdf",
    )


@login_required
def visit_report_pdf(request, pk):
    visit = get_object_or_404(
        PatientVisit.objects.select_related("patient"),
        pk=pk,
    )

    log_activity(
        request,
        AuditLog.ActionType.PRINT,
        "PDF Reports",
        f"Generated PDF for patient {visit.patient.file_number}.",
        object_id=visit.patient.pk,
        object_repr=visit.patient.file_number,
    )

    return render_to_pdf(
        "patients/pdf/visit_report_pdf.html",
        {"visit": visit},
        f"{visit.visit_number}_visit_report.pdf",
    )


@login_required
@finance_staff_required
def bill_receipt_pdf(request, pk):
    bill = get_object_or_404(
        Bill.objects.select_related("patient", "created_by"),
        pk=pk,
    )

    payments = bill.payments.select_related("received_by").all()

    log_activity(
        request,
        AuditLog.ActionType.PRINT,
        "PDF Reports",
        f"Generated PDF for patient {bill.patient.file_number}.",
        object_id=bill.patient.pk,
        object_repr=bill.patient.file_number,
    )

    return render_to_pdf(
        "patients/pdf/bill_receipt_pdf.html",
        {
            "bill": bill,
            "payments": payments,
        },
        f"{bill.bill_number}_receipt.pdf",
    )



@login_required
def surgery_theatre_dashboard(request):
    today = timezone.localdate()
    selected_date = request.GET.get("date", "")

    if selected_date:
        theatre_date = selected_date
    else:
        theatre_date = today

    surgeries = SurgeryProcedure.objects.select_related(
        "patient",
        "visit",
        "surgeon",
        "assistant",
        "created_by",
    ).filter(
        scheduled_date__date=theatre_date
    )

    planned_count = surgeries.filter(status=SurgeryProcedure.ProcedureStatus.PLANNED).count()
    in_progress_count = surgeries.filter(status=SurgeryProcedure.ProcedureStatus.IN_PROGRESS).count()
    completed_count = surgeries.filter(status=SurgeryProcedure.ProcedureStatus.COMPLETED).count()
    postponed_count = surgeries.filter(status=SurgeryProcedure.ProcedureStatus.POSTPONED).count()

    return render(request, "patients/surgery_theatre_dashboard.html", {
        "surgeries": surgeries,
        "theatre_date": theatre_date,
        "today": today,
        "planned_count": planned_count,
        "in_progress_count": in_progress_count,
        "completed_count": completed_count,
        "postponed_count": postponed_count,
    })


@login_required
@clinical_staff_required
def surgery_mark_in_progress(request, pk):
    surgery = get_object_or_404(SurgeryProcedure, pk=pk)
    surgery.status = SurgeryProcedure.ProcedureStatus.IN_PROGRESS
    surgery.save(update_fields=["status", "updated_at"])

    messages.success(request, "Surgery/procedure marked as in progress.")
    return redirect("surgery_theatre_dashboard")


@login_required
@clinical_staff_required
def surgery_mark_completed(request, pk):
    surgery = get_object_or_404(SurgeryProcedure, pk=pk)
    surgery.status = SurgeryProcedure.ProcedureStatus.COMPLETED

    if not surgery.procedure_date:
        surgery.procedure_date = timezone.now()

    surgery.save(update_fields=["status", "procedure_date", "updated_at"])

    messages.success(request, "Surgery/procedure marked as completed.")
    return redirect("surgery_theatre_dashboard")


@login_required
@clinical_staff_required
def surgery_mark_postponed(request, pk):
    surgery = get_object_or_404(SurgeryProcedure, pk=pk)
    surgery.status = SurgeryProcedure.ProcedureStatus.POSTPONED
    surgery.save(update_fields=["status", "updated_at"])

    messages.success(request, "Surgery/procedure marked as postponed.")
    return redirect("surgery_theatre_dashboard")


@login_required
@clinical_staff_required
def surgery_mark_cancelled(request, pk):
    surgery = get_object_or_404(SurgeryProcedure, pk=pk)
    surgery.status = SurgeryProcedure.ProcedureStatus.CANCELLED
    surgery.save(update_fields=["status", "updated_at"])

    messages.success(request, "Surgery/procedure marked as cancelled.")
    return redirect("surgery_theatre_dashboard")



@login_required
def reports_dashboard(request):
    today = timezone.localdate()

    total_patients = Patient.objects.count()
    total_visits = PatientVisit.objects.count()
    total_appointments = Appointment.objects.count()
    total_surgeries = SurgeryProcedure.objects.count()
    total_bills = Bill.objects.count()

    total_revenue = Payment.objects.aggregate(
        total=Sum("amount")
    )["total"] or 0

    today_revenue = Payment.objects.filter(
        payment_date__date=today
    ).aggregate(
        total=Sum("amount")
    )["total"] or 0

    outstanding_balance = sum(
        bill.balance
        for bill in Bill.objects.exclude(status=Bill.BillStatus.PAID).exclude(status=Bill.BillStatus.CANCELLED)
    )

    return render(request, "patients/reports_dashboard.html", {
        "today": today,
        "total_patients": total_patients,
        "total_visits": total_visits,
        "total_appointments": total_appointments,
        "total_surgeries": total_surgeries,
        "total_bills": total_bills,
        "total_revenue": total_revenue,
        "today_revenue": today_revenue,
        "outstanding_balance": outstanding_balance,
    })


@login_required
@finance_staff_required
def financial_report(request):
    today = timezone.localdate()
    start_date = request.GET.get("start_date") or today
    end_date = request.GET.get("end_date") or today

    payments = Payment.objects.select_related(
        "bill",
        "bill__patient",
        "received_by",
    ).filter(
        payment_date__date__gte=start_date,
        payment_date__date__lte=end_date,
    )

    bills = Bill.objects.select_related(
        "patient",
        "created_by",
    ).filter(
        created_at__date__gte=start_date,
        created_at__date__lte=end_date,
    )

    total_payments = payments.aggregate(total=Sum("amount"))["total"] or 0
    total_billed = bills.aggregate(total=Sum("total_amount"))["total"] or 0
    total_discount = bills.aggregate(total=Sum("discount"))["total"] or 0
    total_paid_on_bills = bills.aggregate(total=Sum("amount_paid"))["total"] or 0

    outstanding = sum(bill.balance for bill in bills)

    payment_method_summary = payments.values("payment_method").annotate(
        total=Sum("amount"),
        count=Count("id"),
    ).order_by("payment_method")

    return render(request, "patients/reports/financial_report.html", {
        "start_date": start_date,
        "end_date": end_date,
        "payments": payments,
        "bills": bills,
        "total_payments": total_payments,
        "total_billed": total_billed,
        "total_discount": total_discount,
        "total_paid_on_bills": total_paid_on_bills,
        "outstanding": outstanding,
        "payment_method_summary": payment_method_summary,
    })


@login_required
def clinical_report(request):
    today = timezone.localdate()
    start_date = request.GET.get("start_date") or today
    end_date = request.GET.get("end_date") or today

    visits = PatientVisit.objects.select_related(
        "patient",
        "created_by",
    ).filter(
        visit_date__date__gte=start_date,
        visit_date__date__lte=end_date,
    )

    consultations = Consultation.objects.select_related(
        "visit",
        "visit__patient",
        "doctor",
    ).filter(
        consultation_date__date__gte=start_date,
        consultation_date__date__lte=end_date,
    )

    prescriptions = Prescription.objects.select_related(
        "visit",
        "visit__patient",
        "prescribed_by",
    ).filter(
        prescribed_at__date__gte=start_date,
        prescribed_at__date__lte=end_date,
    )

    diagnosis_records = DiagnosisTreatment.objects.select_related(
        "visit",
        "visit__patient",
    ).filter(
        created_at__date__gte=start_date,
        created_at__date__lte=end_date,
    )

    visit_type_summary = visits.values("visit_type").annotate(
        count=Count("id")
    ).order_by("visit_type")

    return render(request, "patients/reports/clinical_report.html", {
        "start_date": start_date,
        "end_date": end_date,
        "visits": visits,
        "consultations": consultations,
        "prescriptions": prescriptions,
        "diagnosis_records": diagnosis_records,
        "visit_type_summary": visit_type_summary,
    })


@login_required
def appointment_report(request):
    today = timezone.localdate()
    start_date = request.GET.get("start_date") or today
    end_date = request.GET.get("end_date") or today
    status = request.GET.get("status", "").strip()

    appointments = Appointment.objects.select_related(
        "patient",
        "assigned_to",
        "created_by",
    ).filter(
        appointment_date__gte=start_date,
        appointment_date__lte=end_date,
    )

    if status:
        appointments = appointments.filter(status=status)

    status_summary = appointments.values("status").annotate(
        count=Count("id")
    ).order_by("status")

    return render(request, "patients/reports/appointment_report.html", {
        "start_date": start_date,
        "end_date": end_date,
        "status": status,
        "appointments": appointments,
        "status_summary": status_summary,
        "status_choices": Appointment.AppointmentStatus.choices,
    })


@login_required
def surgery_report(request):
    today = timezone.localdate()
    start_date = request.GET.get("start_date") or today
    end_date = request.GET.get("end_date") or today
    status = request.GET.get("status", "").strip()

    surgeries = SurgeryProcedure.objects.select_related(
        "patient",
        "visit",
        "surgeon",
        "assistant",
    ).filter(
        scheduled_date__date__gte=start_date,
        scheduled_date__date__lte=end_date,
    )

    if status:
        surgeries = surgeries.filter(status=status)

    status_summary = surgeries.values("status").annotate(
        count=Count("id")
    ).order_by("status")

    return render(request, "patients/reports/surgery_report.html", {
        "start_date": start_date,
        "end_date": end_date,
        "status": status,
        "surgeries": surgeries,
        "status_summary": status_summary,
        "status_choices": SurgeryProcedure.ProcedureStatus.choices,
    })

@login_required
@finance_staff_required
def financial_report_pdf(request):
    today = timezone.localdate()
    start_date = request.GET.get("start_date") or today
    end_date = request.GET.get("end_date") or today

    payments = Payment.objects.select_related(
        "bill",
        "bill__patient",
        "received_by",
    ).filter(
        payment_date__date__gte=start_date,
        payment_date__date__lte=end_date,
    )

    bills = Bill.objects.select_related(
        "patient",
        "created_by",
    ).filter(
        created_at__date__gte=start_date,
        created_at__date__lte=end_date,
    )

    total_payments = payments.aggregate(total=Sum("amount"))["total"] or 0
    total_billed = bills.aggregate(total=Sum("total_amount"))["total"] or 0
    total_discount = bills.aggregate(total=Sum("discount"))["total"] or 0
    outstanding = sum(bill.balance for bill in bills)

    payment_method_summary = payments.values("payment_method").annotate(
        total=Sum("amount"),
        count=Count("id"),
    ).order_by("payment_method")

    return render_to_pdf(
        "patients/reports/pdf/financial_report_pdf.html",
        {
            "start_date": start_date,
            "end_date": end_date,
            "payments": payments,
            "bills": bills,
            "total_payments": total_payments,
            "total_billed": total_billed,
            "total_discount": total_discount,
            "outstanding": outstanding,
            "payment_method_summary": payment_method_summary,
        },
        f"financial_report_{start_date}_to_{end_date}.pdf",
    )


@login_required
def clinical_report_pdf(request):
    today = timezone.localdate()
    start_date = request.GET.get("start_date") or today
    end_date = request.GET.get("end_date") or today

    visits = PatientVisit.objects.select_related(
        "patient",
        "created_by",
    ).filter(
        visit_date__date__gte=start_date,
        visit_date__date__lte=end_date,
    )

    consultations = Consultation.objects.select_related(
        "visit",
        "visit__patient",
        "doctor",
    ).filter(
        consultation_date__date__gte=start_date,
        consultation_date__date__lte=end_date,
    )

    prescriptions = Prescription.objects.select_related(
        "visit",
        "visit__patient",
        "prescribed_by",
    ).filter(
        prescribed_at__date__gte=start_date,
        prescribed_at__date__lte=end_date,
    )

    diagnosis_records = DiagnosisTreatment.objects.select_related(
        "visit",
        "visit__patient",
    ).filter(
        created_at__date__gte=start_date,
        created_at__date__lte=end_date,
    )

    visit_type_summary = visits.values("visit_type").annotate(
        count=Count("id")
    ).order_by("visit_type")

    return render_to_pdf(
        "patients/reports/pdf/clinical_report_pdf.html",
        {
            "start_date": start_date,
            "end_date": end_date,
            "visits": visits,
            "consultations": consultations,
            "prescriptions": prescriptions,
            "diagnosis_records": diagnosis_records,
            "visit_type_summary": visit_type_summary,
        },
        f"clinical_report_{start_date}_to_{end_date}.pdf",
    )


@login_required
def appointment_report_pdf(request):
    today = timezone.localdate()
    start_date = request.GET.get("start_date") or today
    end_date = request.GET.get("end_date") or today
    status = request.GET.get("status", "").strip()

    appointments = Appointment.objects.select_related(
        "patient",
        "assigned_to",
        "created_by",
    ).filter(
        appointment_date__gte=start_date,
        appointment_date__lte=end_date,
    )

    if status:
        appointments = appointments.filter(status=status)

    status_summary = appointments.values("status").annotate(
        count=Count("id")
    ).order_by("status")

    return render_to_pdf(
        "patients/reports/pdf/appointment_report_pdf.html",
        {
            "start_date": start_date,
            "end_date": end_date,
            "status": status,
            "appointments": appointments,
            "status_summary": status_summary,
        },
        f"appointment_report_{start_date}_to_{end_date}.pdf",
    )


@login_required
def surgery_report_pdf(request):
    today = timezone.localdate()
    start_date = request.GET.get("start_date") or today
    end_date = request.GET.get("end_date") or today
    status = request.GET.get("status", "").strip()

    surgeries = SurgeryProcedure.objects.select_related(
        "patient",
        "visit",
        "surgeon",
        "assistant",
    ).filter(
        scheduled_date__date__gte=start_date,
        scheduled_date__date__lte=end_date,
    )

    if status:
        surgeries = surgeries.filter(status=status)

    status_summary = surgeries.values("status").annotate(
        count=Count("id")
    ).order_by("status")

    return render_to_pdf(
        "patients/reports/pdf/surgery_report_pdf.html",
        {
            "start_date": start_date,
            "end_date": end_date,
            "status": status,
            "surgeries": surgeries,
            "status_summary": status_summary,
        },
        f"surgery_report_{start_date}_to_{end_date}.pdf",
    )


@login_required
@finance_staff_required
def financial_report_excel(request):
    today = timezone.localdate()
    start_date = request.GET.get("start_date") or today
    end_date = request.GET.get("end_date") or today

    payments = Payment.objects.select_related(
        "bill",
        "bill__patient",
        "received_by",
    ).filter(
        payment_date__date__gte=start_date,
        payment_date__date__lte=end_date,
    )

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Financial Report"

    sheet.append(["CAPITAL EYE HOSPITAL - FINANCIAL REPORT"])
    sheet.append([f"Period: {start_date} to {end_date}"])
    sheet.append([])

    headers = [
        "Payment Date",
        "Receipt Number",
        "Patient",
        "File Number",
        "Bill Number",
        "Bill Title",
        "Payment Method",
        "Reference",
        "Amount",
        "Received By",
    ]

    sheet.append(headers)

    for payment in payments:
        sheet.append([
            payment.payment_date.strftime("%Y-%m-%d %H:%M"),
            payment.receipt_number,
            payment.bill.patient.full_name,
            payment.bill.patient.file_number,
            payment.bill.bill_number,
            payment.bill.bill_title,
            payment.get_payment_method_display(),
            payment.reference_number,
            float(payment.amount),
            str(payment.received_by) if payment.received_by else "",
        ])

    autosize_excel_columns(sheet)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"financial_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    workbook.save(response)
    return response


@login_required
def clinical_report_excel(request):
    today = timezone.localdate()
    start_date = request.GET.get("start_date") or today
    end_date = request.GET.get("end_date") or today

    visits = PatientVisit.objects.select_related(
        "patient",
        "created_by",
    ).filter(
        visit_date__date__gte=start_date,
        visit_date__date__lte=end_date,
    )

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Clinical Report"

    sheet.append(["CAPITAL EYE HOSPITAL - CLINICAL REPORT"])
    sheet.append([f"Period: {start_date} to {end_date}"])
    sheet.append([])

    headers = [
        "Visit Date",
        "Visit Number",
        "Patient",
        "File Number",
        "Phone",
        "Visit Type",
        "Status",
        "Chief Complaint",
        "Brief History",
        "Created By",
    ]

    sheet.append(headers)

    for visit in visits:
        sheet.append([
            visit.visit_date.strftime("%Y-%m-%d %H:%M"),
            visit.visit_number,
            visit.patient.full_name,
            visit.patient.file_number,
            visit.patient.phone_number,
            visit.get_visit_type_display(),
            visit.get_status_display(),
            visit.chief_complaint,
            visit.brief_history,
            str(visit.created_by) if visit.created_by else "",
        ])

    autosize_excel_columns(sheet)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"clinical_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    workbook.save(response)
    return response


@login_required
def appointment_report_excel(request):
    today = timezone.localdate()
    start_date = request.GET.get("start_date") or today
    end_date = request.GET.get("end_date") or today
    status = request.GET.get("status", "").strip()

    appointments = Appointment.objects.select_related(
        "patient",
        "assigned_to",
        "created_by",
    ).filter(
        appointment_date__gte=start_date,
        appointment_date__lte=end_date,
    )

    if status:
        appointments = appointments.filter(status=status)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Appointment Report"

    sheet.append(["CAPITAL EYE HOSPITAL - APPOINTMENT REPORT"])
    sheet.append([f"Period: {start_date} to {end_date}"])
    sheet.append([])

    headers = [
        "Date",
        "Time",
        "Appointment Number",
        "Patient",
        "File Number",
        "Phone",
        "Type",
        "Status",
        "Assigned To",
        "Reason",
    ]

    sheet.append(headers)

    for appointment in appointments:
        sheet.append([
            appointment.appointment_date.strftime("%Y-%m-%d"),
            appointment.appointment_time.strftime("%H:%M"),
            appointment.appointment_number,
            appointment.patient.full_name,
            appointment.patient.file_number,
            appointment.patient.phone_number,
            appointment.get_appointment_type_display(),
            appointment.get_status_display(),
            str(appointment.assigned_to) if appointment.assigned_to else "",
            appointment.reason,
        ])

    autosize_excel_columns(sheet)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"appointment_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    workbook.save(response)
    return response


@login_required
def surgery_report_excel(request):
    today = timezone.localdate()
    start_date = request.GET.get("start_date") or today
    end_date = request.GET.get("end_date") or today
    status = request.GET.get("status", "").strip()

    surgeries = SurgeryProcedure.objects.select_related(
        "patient",
        "visit",
        "surgeon",
        "assistant",
    ).filter(
        scheduled_date__date__gte=start_date,
        scheduled_date__date__lte=end_date,
    )

    if status:
        surgeries = surgeries.filter(status=status)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Surgery Report"

    sheet.append(["CAPITAL EYE HOSPITAL - SURGERY / PROCEDURE REPORT"])
    sheet.append([f"Period: {start_date} to {end_date}"])
    sheet.append([])

    headers = [
        "Scheduled Date",
        "Procedure Number",
        "Patient",
        "File Number",
        "Phone",
        "Procedure Name",
        "Procedure Type",
        "Eye Side",
        "Status",
        "Surgeon",
        "Assistant",
        "Pre-op Diagnosis",
        "Post-op Diagnosis",
        "Outcome",
    ]

    sheet.append(headers)

    for surgery in surgeries:
        sheet.append([
            surgery.scheduled_date.strftime("%Y-%m-%d %H:%M") if surgery.scheduled_date else "",
            surgery.procedure_number,
            surgery.patient.full_name,
            surgery.patient.file_number,
            surgery.patient.phone_number,
            surgery.procedure_name,
            surgery.procedure_type,
            surgery.get_eye_side_display(),
            surgery.get_status_display(),
            str(surgery.surgeon) if surgery.surgeon else "",
            str(surgery.assistant) if surgery.assistant else "",
            surgery.pre_op_diagnosis,
            surgery.post_op_diagnosis,
            surgery.outcome,
        ])

    autosize_excel_columns(sheet)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"surgery_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    workbook.save(response)
    return response



@login_required
def patient_ophthalmology_timeline(request, patient_pk):
    """
    Display the patient's ophthalmology records as one unified,
    chronological clinical timeline.

    Existing records are queried and represented as timeline events.
    No clinical data is duplicated.
    """

    patient = get_object_or_404(
        Patient.objects.select_related(
            "family_group",
            "registered_by",
        ),
        pk=patient_pk,
    )

    timeline_events = []

    # =====================================================
    # VISITS
    # =====================================================

    visits = (
        patient.visits
        .select_related("created_by")
        .order_by("-visit_date", "-pk")
    )

    for visit in visits:
        event_datetime = normalize_timeline_datetime(
            visit.visit_date
        )

        timeline_events.append(
            {
                "event_type": "VISIT",
                "event_datetime": event_datetime,
                "date": event_datetime,
                "title": (
                    f"{visit.get_visit_type_display()} Visit"
                ),
                "subtitle": visit.visit_number,
                "description": (
                    getattr(visit, "chief_complaint", "")
                    or "Patient visit recorded."
                ),
                "staff_name": timeline_user_name(
                    getattr(visit, "created_by", None)
                ),
                "status": (
                    visit.get_status_display()
                    if hasattr(visit, "get_status_display")
                    else ""
                ),
                "icon": "🏥",
                "css_class": "visit",
                "url_name": "visit_detail",
                "url_kwargs": {
                    "pk": visit.pk,
                },
                "visit": visit,
            }
        )

    # =====================================================
    # CONSULTATIONS
    # =====================================================

    consultations = (
        Consultation.objects
        .filter(
            visit__patient=patient,
        )
        .select_related(
            "visit",
            "visit__patient",
            "doctor",
        )
    )

    for consultation in consultations:
        event_datetime = normalize_timeline_datetime(
            getattr(
                consultation,
                "consultation_date",
                None,
            )
            or getattr(
                consultation,
                "created_at",
                None,
            )
            or consultation.visit.visit_date
        )

        description_parts = []

        for field_name in [
            "history",
            "complaint",
            "examination_findings",
            "clinical_notes",
            "notes",
        ]:
            field_value = getattr(
                consultation,
                field_name,
                "",
            )

            if field_value:
                description_parts.append(
                    str(field_value)
                )
                break

        timeline_events.append(
            {
                "event_type": "CONSULTATION",
                "event_datetime": event_datetime,
                "date": event_datetime,
                "title": "Clinical Consultation",
                "subtitle": (
                    consultation.visit.visit_number
                ),
                "description": (
                    description_parts[0]
                    if description_parts
                    else "Consultation record completed."
                ),
                "staff_name": timeline_user_name(
                    getattr(
                        consultation,
                        "doctor",
                        None,
                    )
                ),
                "status": "Clinical Record",
                "icon": "🩺",
                "css_class": "consultation",
                "url_name": "visit_detail",
                "url_kwargs": {
                    "pk": consultation.visit.pk,
                },
                "visit": consultation.visit,
            }
        )

    # =====================================================
    # EYE EXAMINATIONS
    # =====================================================

    eye_examinations = (
        EyeExamination.objects
        .filter(
            visit__patient=patient,
        )
        .select_related(
            "visit",
            "examined_by",
            "subjective_refraction_approved_by",
        )
    )

    for examination in eye_examinations:
        event_datetime = normalize_timeline_datetime(
            getattr(
                examination,
                "examination_date",
                None,
            )
            or examination.visit.visit_date
        )

        examination_summary = []

        if examination.right_visual_acuity:
            examination_summary.append(
                "Right VA: "
                f"{examination.right_visual_acuity}"
            )

        if examination.left_visual_acuity:
            examination_summary.append(
                "Left VA: "
                f"{examination.left_visual_acuity}"
            )

        if examination.right_corrected_iop:
            examination_summary.append(
                "Right IOP: "
                f"{examination.right_corrected_iop}"
            )

        if examination.left_corrected_iop:
            examination_summary.append(
                "Left IOP: "
                f"{examination.left_corrected_iop}"
            )

        timeline_events.append(
            {
                "event_type": "EYE_EXAMINATION",
                "event_datetime": event_datetime,
                "date": event_datetime,
                "title": "Comprehensive Eye Examination",
                "subtitle": (
                    examination.visit.visit_number
                ),
                "description": (
                    " · ".join(examination_summary)
                    if examination_summary
                    else (
                        "Structured ophthalmology examination "
                        "recorded."
                    )
                ),
                "staff_name": timeline_user_name(
                    examination.examined_by
                ),
                "status": (
                    "Subjective Refraction Approved"
                    if examination.subjective_refraction_approved
                    else "Examination Recorded"
                ),
                "icon": "👁️",
                "css_class": "eye-examination",
                "url_name": "eye_examination_update",
                "url_kwargs": {
                    "pk": examination.pk,
                },
                "visit": examination.visit,
            }
        )

    # =====================================================
    # DIAGNOSES AND TREATMENTS
    # =====================================================

    diagnoses = (
        DiagnosisTreatment.objects
        .filter(
            visit__patient=patient,
        )
        .select_related(
            "visit",
            "created_by",
        )
    )

    for diagnosis in diagnoses:
        event_datetime = normalize_timeline_datetime(
            getattr(
                diagnosis,
                "diagnosis_date",
                None,
            )
            or getattr(
                diagnosis,
                "created_at",
                None,
            )
            or diagnosis.visit.visit_date
        )

        diagnosis_text = (
            getattr(diagnosis, "diagnosis", "")
            or getattr(
                diagnosis,
                "diagnosis_details",
                "",
            )
            or "Diagnosis and treatment recorded."
        )

        diagnosis_staff = (
            getattr(diagnosis, "diagnosed_by", None)
            or getattr(diagnosis, "created_by", None)
        )

        timeline_events.append(
            {
                "event_type": "DIAGNOSIS",
                "event_datetime": event_datetime,
                "date": event_datetime,
                "title": "Diagnosis and Treatment",
                "subtitle": diagnosis.visit.visit_number,
                "description": diagnosis_text,
                "staff_name": timeline_user_name(
                    diagnosis_staff
                ),
                "status": "Clinical Diagnosis",
                "icon": "🧠",
                "css_class": "diagnosis",
                "url_name": "visit_detail",
                "url_kwargs": {
                    "pk": diagnosis.visit.pk,
                },
                "visit": diagnosis.visit,
            }
        )

    # =====================================================
    # PRESCRIPTIONS
    # =====================================================

    prescriptions = (
        Prescription.objects
        .filter(
            visit__patient=patient,
        )
        .select_related(
            "visit",
            "prescribed_by",
        )
    )

    for prescription in prescriptions:
        event_datetime = normalize_timeline_datetime(
            getattr(
                prescription,
                "prescription_date",
                None,
            )
            or getattr(
                prescription,
                "created_at",
                None,
            )
            or prescription.visit.visit_date
        )

        medication_name = (
            getattr(prescription, "medication", "")
            or getattr(
                prescription,
                "drug_name",
                "",
            )
            or "Medication prescription"
        )

        medication_details = []

        for field_name in [
            "dosage",
            "frequency",
            "duration",
        ]:
            value = getattr(
                prescription,
                field_name,
                "",
            )

            if value:
                medication_details.append(
                    str(value)
                )

        description = medication_name

        if medication_details:
            description += (
                " — "
                + " · ".join(medication_details)
            )

        timeline_events.append(
            {
                "event_type": "PRESCRIPTION",
                "event_datetime": event_datetime,
                "date": event_datetime,
                "title": "Medication Prescription",
                "subtitle": (
                    prescription.visit.visit_number
                ),
                "description": description,
                "staff_name": timeline_user_name(
                    getattr(
                        prescription,
                        "prescribed_by",
                        None,
                    )
                ),
                "status": (
                    prescription.get_status_display()
                    if hasattr(
                        prescription,
                        "get_status_display",
                    )
                    else "Prescription"
                ),
                "icon": "💊",
                "css_class": "prescription",
                "url_name": "visit_detail",
                "url_kwargs": {
                    "pk": prescription.visit.pk,
                },
                "visit": prescription.visit,
            }
        )

    # =====================================================
    # SURGERIES AND PROCEDURES
    # =====================================================

    surgeries = (
        SurgeryProcedure.objects
        .filter(
            patient=patient,
        )
        .select_related(
            "visit",
            "surgeon",
            "assistant",
            "created_by",
        )
    )

    for surgery in surgeries:
        event_datetime = normalize_timeline_datetime(
            getattr(
                surgery,
                "scheduled_date",
                None,
            )
            or getattr(
                surgery,
                "procedure_date",
                None,
            )
            or getattr(
                surgery,
                "created_at",
                None,
            )
        )

        procedure_name = (
            getattr(
                surgery,
                "procedure_name",
                "",
            )
            or getattr(
                surgery,
                "surgery_type",
                "",
            )
            or "Surgery / Procedure"
        )

        timeline_events.append(
            {
                "event_type": "SURGERY",
                "event_datetime": event_datetime,
                "date": event_datetime,
                "title": procedure_name,
                "subtitle": (
                    surgery.visit.visit_number
                    if surgery.visit
                    else "Surgery Record"
                ),
                "description": (
                    getattr(
                        surgery,
                        "notes",
                        "",
                    )
                    or getattr(
                        surgery,
                        "operative_notes",
                        "",
                    )
                    or "Surgery or procedure recorded."
                ),
                "staff_name": timeline_user_name(
                    getattr(
                        surgery,
                        "surgeon",
                        None,
                    )
                ),
                "status": (
                    surgery.get_status_display()
                    if hasattr(
                        surgery,
                        "get_status_display",
                    )
                    else "Procedure"
                ),
                "icon": "🏥",
                "css_class": "surgery",
                "url_name": "visit_detail",
                "url_kwargs": {
                    "pk": (
                        surgery.visit.pk
                        if surgery.visit
                        else patient.visits.first().pk
                    ),
                },
                "visit": surgery.visit,
            }
        )

    # =====================================================
    # CLINICAL ATTACHMENTS
    # =====================================================

    attachments = (
        ClinicalAttachment.objects
        .filter(
            patient=patient,
            is_active=True,
        )
        .select_related(
            "visit",
            "uploaded_by",
            "reviewed_by",
        )
    )

    for attachment in attachments:
        event_datetime = normalize_timeline_datetime(
            attachment.investigation_date
            or attachment.uploaded_at
        )

        if attachment.category == (
            ClinicalAttachment.Category.OTHER
        ):
            category_name = attachment.category_other
        else:
            category_name = (
                attachment.get_category_display()
            )

        timeline_events.append(
            {
                "event_type": "ATTACHMENT",
                "event_datetime": event_datetime,
                "date": event_datetime,
                "title": attachment.title,
                "subtitle": (
                    f"{category_name} · "
                    f"{attachment.visit.visit_number}"
                ),
                "description": (
                    attachment.description
                    or (
                        f"{attachment.get_eye_side_display()} · "
                        f"{attachment.file_size_display}"
                    )
                ),
                "staff_name": timeline_user_name(
                    attachment.uploaded_by
                ),
                "status": (
                    attachment.get_review_status_display()
                ),
                "review_status": (
                    attachment.review_status
                ),
                "icon": (
                    "🖼️"
                    if attachment.is_image
                    else (
                        "📕"
                        if attachment.is_pdf
                        else "📘"
                    )
                ),
                "css_class": "attachment",
                "url_name": (
                    "clinical_attachment_viewer"
                    if (
                        attachment.is_image
                        or attachment.is_pdf
                    )
                    else "clinical_attachment_detail"
                ),
                "url_kwargs": {
                    "pk": attachment.pk,
                },
                "url_query": (
                    "?scope=patient"
                    if (
                        attachment.is_image
                        or attachment.is_pdf
                    )
                    else ""
                ),
                "visit": attachment.visit,
            }
        )

    # =====================================================
    # CONTACT LENS ASSESSMENTS
    # =====================================================

    contact_lens_assessments = (
        ContactLensAssessment.objects
        .filter(patient=patient)
        .select_related(
            "visit",
            "eye_examination",
            "assessed_by",
        )
        .order_by(
            "-assessment_date",
            "-pk",
        )
    )

    for assessment in contact_lens_assessments:
        event_datetime = normalize_timeline_datetime(
            assessment.assessment_date
        )

        purpose_display = (
            assessment.lens_purpose_other
            if assessment.lens_purpose == "OTHER"
            else assessment.get_lens_purpose_display()
        )

        assessment_description_parts = []

        if purpose_display:
            assessment_description_parts.append(
                f"Purpose: {purpose_display}"
            )

        assessment_description_parts.append(
            "Suitability: "
            f"{assessment.get_suitability_status_display()}"
        )

        if assessment.tear_film_status:
            assessment_description_parts.append(
                "Tear film: "
                f"{assessment.get_tear_film_status_display()}"
            )

        if assessment.suitability_reason:
            assessment_description_parts.append(
                assessment.suitability_reason
            )

        timeline_events.append(
            {
                "event_type": "CONTACT_LENS_ASSESSMENT",
                "event_datetime": event_datetime,
                "date": event_datetime,
                "title": "Contact Lens Suitability Assessment",
                "subtitle": (
                    assessment.visit.visit_number
                    if assessment.visit
                    else "Contact Lens Assessment"
                ),
                "description": " · ".join(
                    assessment_description_parts
                ),
                "staff_name": timeline_user_name(
                    assessment.assessed_by
                ),
                "status": (
                    assessment.get_suitability_status_display()
                ),
                "status_class": (
                    contact_lens_timeline_status_class(
                        assessment.suitability_status
                    )
                ),
                "icon": "👁️",
                "css_class": "contact-lens-assessment",
                "url_name": "contact_lens_assessment_detail",
                "url_kwargs": {
                    "pk": assessment.pk,
                },
                "visit": assessment.visit,
            }
        )

    # =====================================================
    # CONTACT LENS TRIALS
    # =====================================================

    contact_lens_trials = (
        ContactLensTrial.objects
        .filter(assessment__patient=patient)
        .select_related(
            "assessment",
            "assessment__visit",
            "fitted_by",
        )
        .order_by(
            "-trial_date",
            "-pk",
        )
    )

    for trial in contact_lens_trials:
        event_datetime = normalize_timeline_datetime(
            trial.trial_date
        )

        trial_design = (
            trial.lens_design_other
            if trial.lens_design == "OTHER"
            else trial.get_lens_design_display()
        )

        trial_parameters = []

        if trial_design:
            trial_parameters.append(str(trial_design))

        if trial.brand_name:
            trial_parameters.append(trial.brand_name)

        if trial.base_curve is not None:
            trial_parameters.append(f"BC {trial.base_curve}")

        if trial.diameter is not None:
            trial_parameters.append(f"DIA {trial.diameter}")

        if trial.sphere is not None:
            trial_parameters.append(f"SPH {trial.sphere}")

        if trial.cylinder is not None:
            trial_parameters.append(f"CYL {trial.cylinder}")

        if trial.axis is not None:
            trial_parameters.append(f"Axis {trial.axis}")

        if trial.final_visual_acuity:
            trial_parameters.append(
                f"Final VA {trial.final_visual_acuity}"
            )

        timeline_events.append(
            {
                "event_type": "CONTACT_LENS_TRIAL",
                "event_datetime": event_datetime,
                "date": event_datetime,
                "title": (
                    f"{trial.get_eye_side_display()} "
                    f"Trial Lens #{trial.trial_number}"
                ),
                "subtitle": (
                    trial.assessment.visit.visit_number
                    if trial.assessment.visit
                    else "Contact Lens Trial"
                ),
                "description": ", ".join(trial_parameters),
                "staff_name": timeline_user_name(
                    trial.fitted_by
                ),
                "status": (
                    "Accepted for Prescription"
                    if trial.accepted_for_prescription
                    else trial.get_fit_assessment_display()
                ),
                "status_class": (
                    "completed"
                    if trial.accepted_for_prescription
                    else (
                        "danger"
                        if trial.fit_assessment
                        == ContactLensTrial.FitAssessment.UNACCEPTABLE
                        else "default"
                    )
                ),
                "icon": "◉",
                "css_class": "contact-lens-trial",
                "url_name": "contact_lens_assessment_detail",
                "url_kwargs": {
                    "pk": trial.assessment_id,
                },
                "visit": trial.assessment.visit,
            }
        )

    # =====================================================
    # CONTACT LENS PRESCRIPTIONS
    # =====================================================

    contact_lens_prescriptions = (
        ContactLensPrescription.objects
        .filter(
            patient=patient,
            is_active=True,
        )
        .select_related(
            "visit",
            "assessment",
            "prescribed_by",
            "approved_by",
            "dispensed_by",
        )
        .order_by(
            "-prescription_date",
            "-pk",
        )
    )

    for prescription in contact_lens_prescriptions:
        event_datetime = normalize_timeline_datetime(
            prescription.created_at
        )

        eye_summaries = []

        right_summary = contact_lens_eye_parameter_summary(
            prescription,
            prefix="right",
            eye_label="OD",
        )

        left_summary = contact_lens_eye_parameter_summary(
            prescription,
            prefix="left",
            eye_label="OS",
        )

        if right_summary:
            eye_summaries.append(right_summary)

        if left_summary:
            eye_summaries.append(left_summary)

        timeline_events.append(
            {
                "event_type": "CONTACT_LENS_PRESCRIPTION",
                "event_datetime": event_datetime,
                "date": event_datetime,
                "title": (
                    f"{prescription.prescription_number} "
                    f"— Version {prescription.version}"
                ),
                "subtitle": (
                    prescription.visit.visit_number
                    if prescription.visit
                    else "Contact Lens Prescription"
                ),
                "description": " · ".join(eye_summaries),
                "staff_name": timeline_user_name(
                    prescription.prescribed_by
                ),
                "status": prescription.get_status_display(),
                "status_class": (
                    contact_lens_timeline_status_class(
                        prescription.status
                    )
                ),
                "icon": "◉",
                "css_class": "contact-lens-prescription",
                "url_name": "contact_lens_prescription_detail",
                "url_kwargs": {
                    "pk": prescription.pk,
                },
                "visit": prescription.visit,
            }
        )

        if prescription.approved_at:
            approval_datetime = normalize_timeline_datetime(
                prescription.approved_at
            )

            timeline_events.append(
                {
                    "event_type": "CONTACT_LENS_APPROVAL",
                    "event_datetime": approval_datetime,
                    "date": approval_datetime,
                    "title": (
                        "Contact Lens Prescription Approved"
                    ),
                    "subtitle": (
                        f"{prescription.prescription_number} "
                        f"— Version {prescription.version}"
                    ),
                    "description": (
                        f"{prescription.prescription_number} "
                        f"— Version {prescription.version}"
                    ),
                    "staff_name": timeline_user_name(
                        prescription.approved_by
                    ),
                    "status": "Approved",
                    "status_class": "completed",
                    "icon": "✓",
                    "css_class": "contact-lens-approval",
                    "url_name": "contact_lens_prescription_detail",
                    "url_kwargs": {
                        "pk": prescription.pk,
                    },
                    "visit": prescription.visit,
                }
            )

        if prescription.dispensed_at:
            dispensed_datetime = normalize_timeline_datetime(
                prescription.dispensed_at
            )

            timeline_events.append(
                {
                    "event_type": "CONTACT_LENS_DISPENSING",
                    "event_datetime": dispensed_datetime,
                    "date": dispensed_datetime,
                    "title": "Contact Lenses Dispensed",
                    "subtitle": (
                        f"{prescription.prescription_number} "
                        f"— Version {prescription.version}"
                    ),
                    "description": (
                        f"{prescription.prescription_number} "
                        f"— Version {prescription.version}"
                    ),
                    "staff_name": timeline_user_name(
                        prescription.dispensed_by
                    ),
                    "status": "Dispensed",
                    "status_class": "completed",
                    "icon": "📦",
                    "css_class": "contact-lens-dispensing",
                    "url_name": "contact_lens_prescription_detail",
                    "url_kwargs": {
                        "pk": prescription.pk,
                    },
                    "visit": prescription.visit,
                }
            )

    # =====================================================
    # CONTACT LENS FOLLOW-UPS
    # =====================================================

    contact_lens_follow_ups = (
        ContactLensFollowUp.objects
        .filter(patient=patient)
        .select_related(
            "prescription",
            "prescription__visit",
            "reviewed_by",
        )
        .order_by(
            "-follow_up_date",
            "-pk",
        )
    )

    for follow_up in contact_lens_follow_ups:
        event_datetime = normalize_timeline_datetime(
            follow_up.follow_up_date
        )

        follow_up_description = []

        if follow_up.wearing_time_per_day is not None:
            follow_up_description.append(
                "Daily wear "
                f"{follow_up.wearing_time_per_day} hours"
            )

        if follow_up.comfort_score_right is not None:
            follow_up_description.append(
                "R comfort "
                f"{follow_up.comfort_score_right}/10"
            )

        if follow_up.comfort_score_left is not None:
            follow_up_description.append(
                "L comfort "
                f"{follow_up.comfort_score_left}/10"
            )

        if follow_up.lens_condition:
            follow_up_description.append(
                follow_up.get_lens_condition_display()
            )

        if follow_up.complications:
            follow_up_description.append(
                f"Complication: {follow_up.complications}"
            )

        timeline_events.append(
            {
                "event_type": "CONTACT_LENS_FOLLOW_UP",
                "event_datetime": event_datetime,
                "date": event_datetime,
                "title": (
                    f"{follow_up.get_status_display()} "
                    f"Contact Lens Review"
                ),
                "subtitle": (
                    follow_up.prescription.visit.visit_number
                    if follow_up.prescription
                    and follow_up.prescription.visit
                    else "Contact Lens Follow-up"
                ),
                "description": " · ".join(
                    follow_up_description
                ),
                "staff_name": timeline_user_name(
                    follow_up.reviewed_by
                ),
                "status": follow_up.get_status_display(),
                "status_class": (
                    contact_lens_timeline_status_class(
                        follow_up.status
                    )
                ),
                "icon": "📅",
                "css_class": "contact-lens-follow-up",
                "url_name": "contact_lens_follow_up_detail",
                "url_kwargs": {
                    "pk": follow_up.pk,
                },
                "visit": (
                    follow_up.prescription.visit
                    if follow_up.prescription
                    else None
                ),
            }
        )

    # =====================================================
    # CLEAN, FILTER AND SORT EVENTS
    # =====================================================

    timeline_events = [
        event
        for event in timeline_events
        if event.get("event_datetime") is not None
    ]

    timeline_events.sort(
        key=lambda item: item["event_datetime"],
        reverse=True,
    )

    selected_event_type = (
        request.GET.get("event_type", "")
        .strip()
        .upper()
    )

    valid_event_types = {
        "VISIT",
        "CONSULTATION",
        "EYE_EXAMINATION",
        "DIAGNOSIS",
        "PRESCRIPTION",
        "SURGERY",
        "ATTACHMENT",
        "CONTACT_LENS_ASSESSMENT",
        "CONTACT_LENS_TRIAL",
        "CONTACT_LENS_PRESCRIPTION",
        "CONTACT_LENS_APPROVAL",
        "CONTACT_LENS_DISPENSING",
        "CONTACT_LENS_FOLLOW_UP",
    }

    if selected_event_type in valid_event_types:
        filtered_events = [
            event
            for event in timeline_events
            if event["event_type"] == selected_event_type
        ]
    else:
        selected_event_type = ""
        filtered_events = timeline_events

    query = request.GET.get(
        "q",
        "",
    ).strip().lower()

    if query:
        filtered_events = [
            event
            for event in filtered_events
            if (
                query in str(
                    event.get("title", "")
                ).lower()
                or query in str(
                    event.get("subtitle", "")
                ).lower()
                or query in str(
                    event.get("description", "")
                ).lower()
                or query in str(
                    event.get("staff_name", "")
                ).lower()
                or query in str(
                    event.get("status", "")
                ).lower()
            )
        ]

    paginator = Paginator(
        filtered_events,
        30,
    )

    page_obj = paginator.get_page(
        request.GET.get("page")
    )

    event_counts = {
        "all": len(timeline_events),
        "visits": sum(
            event["event_type"] == "VISIT"
            for event in timeline_events
        ),
        "examinations": sum(
            event["event_type"] == "EYE_EXAMINATION"
            for event in timeline_events
        ),
        "diagnoses": sum(
            event["event_type"] == "DIAGNOSIS"
            for event in timeline_events
        ),
        "prescriptions": sum(
            event["event_type"] == "PRESCRIPTION"
            for event in timeline_events
        ),
        "surgeries": sum(
            event["event_type"] == "SURGERY"
            for event in timeline_events
        ),
        "attachments": sum(
            event["event_type"] == "ATTACHMENT"
            for event in timeline_events
        ),
        "contact_lens_assessments": sum(
            event["event_type"] == "CONTACT_LENS_ASSESSMENT"
            for event in timeline_events
        ),
        "contact_lens_trials": sum(
            event["event_type"] == "CONTACT_LENS_TRIAL"
            for event in timeline_events
        ),
        "contact_lens_prescriptions": sum(
            event["event_type"] == "CONTACT_LENS_PRESCRIPTION"
            for event in timeline_events
        ),
        "contact_lens_follow_ups": sum(
            event["event_type"] == "CONTACT_LENS_FOLLOW_UP"
            for event in timeline_events
        ),
    }

    # =====================================================
    # CONTACT LENS SUMMARY COUNTS (unfiltered totals, for the
    # dedicated Contact Lens summary cards on this page)
    # =====================================================

    contact_lens_timeline_counts = {
        "assessments": contact_lens_assessments.count(),
        "trials": contact_lens_trials.count(),
        "prescriptions": contact_lens_prescriptions.count(),
        "follow_ups": contact_lens_follow_ups.count(),
        "approved_prescriptions": (
            contact_lens_prescriptions.filter(
                status__in=[
                    ContactLensPrescription
                    .PrescriptionStatus
                    .APPROVED,

                    ContactLensPrescription
                    .PrescriptionStatus
                    .DISPENSED,
                ]
            ).count()
        ),
    }

    return render(
        request,
        "patients/ophthalmology_timeline.html",
        {
            "patient": patient,
            "timeline_events": page_obj,
            "page_obj": page_obj,
            "event_counts": event_counts,
            "contact_lens_timeline_counts": (
                contact_lens_timeline_counts
            ),
            "selected_event_type": selected_event_type,
            "query": request.GET.get("q", "").strip(),
            "event_type_choices": [
                ("", "All Clinical Events"),
                ("VISIT", "Visits"),
                ("CONSULTATION", "Consultations"),
                (
                    "EYE_EXAMINATION",
                    "Eye Examinations",
                ),
                ("DIAGNOSIS", "Diagnoses"),
                (
                    "PRESCRIPTION",
                    "Prescriptions",
                ),
                (
                    "SURGERY",
                    "Surgeries / Procedures",
                ),
                (
                    "ATTACHMENT",
                    "Investigation Files",
                ),
                (
                    "CONTACT_LENS_ASSESSMENT",
                    "Contact Lens Assessments",
                ),
                (
                    "CONTACT_LENS_TRIAL",
                    "Contact Lens Trials",
                ),
                (
                    "CONTACT_LENS_PRESCRIPTION",
                    "Contact Lens Prescriptions",
                ),
                (
                    "CONTACT_LENS_APPROVAL",
                    "Contact Lens Approvals",
                ),
                (
                    "CONTACT_LENS_DISPENSING",
                    "Contact Lens Dispensing",
                ),
                (
                    "CONTACT_LENS_FOLLOW_UP",
                    "Contact Lens Follow-ups",
                ),
            ],
        },
    )


@login_required
@clinical_staff_required
def ophthalmology_dashboard(request):
    """
    Specialist ophthalmology dashboard.

    Displays operational KPIs, review queues, investigation statistics,
    recent annotation activity and monthly clinical trends.
    """

    date_range = ophthalmology_dashboard_date_range(
        request
    )

    selected_period = date_range[
        "selected_period"
    ]

    start_date = date_range["start_date"]
    end_date = date_range["end_date"]

    # =====================================================
    # BASE QUERYSETS
    # =====================================================

    patients_queryset = Patient.objects.all()

    visits_queryset = PatientVisit.objects.select_related(
        "patient",
        "created_by",
    )

    examinations_queryset = (
        EyeExamination.objects
        .select_related(
            "visit",
            "visit__patient",
            "examined_by",
        )
    )

    attachments_queryset = (
        ClinicalAttachment.objects
        .filter(is_active=True)
        .select_related(
            "patient",
            "visit",
            "uploaded_by",
            "reviewed_by",
        )
    )

    annotations_queryset = (
        ClinicalImageAnnotation.objects
        .filter(is_active=True)
        .select_related(
            "attachment",
            "attachment__patient",
            "created_by",
            "updated_by",
        )
    )

    surgeries_queryset = (
        SurgeryProcedure.objects
        .select_related(
            "patient",
            "surgeon",
            "created_by",
        )
    )

    appointments_queryset = (
        Appointment.objects
        .select_related(
            "patient",
            "assigned_to",
            "created_by",
        )
    )

    contact_lens_assessments_queryset = (
        ContactLensAssessment.objects
        .select_related(
            "patient",
            "visit",
            "assessed_by",
        )
    )

    contact_lens_trials_queryset = (
        ContactLensTrial.objects
        .select_related(
            "assessment",
            "assessment__patient",
            "fitted_by",
        )
    )

    contact_lens_prescriptions_queryset = (
        ContactLensPrescription.objects
        .filter(is_active=True)
        .select_related(
            "patient",
            "visit",
            "assessment",
            "prescribed_by",
            "approved_by",
            "dispensed_by",
        )
    )

    contact_lens_follow_ups_queryset = (
        ContactLensFollowUp.objects
        .select_related(
            "patient",
            "prescription",
            "prescription__visit",
            "reviewed_by",
        )
    )

    # =====================================================
    # DATE FILTERS
    # =====================================================

    if start_date and end_date:
        if model_has_field(
            Patient,
            "created_at",
        ):
            patients_queryset = (
                patients_queryset.filter(
                    created_at__date__range=(
                        start_date,
                        end_date,
                    )
                )
            )

        visits_queryset = visits_queryset.filter(
            visit_date__date__range=(
                start_date,
                end_date,
            )
        )

        examinations_queryset = (
            examinations_queryset.filter(
                visit__visit_date__date__range=(
                    start_date,
                    end_date,
                )
            )
        )

        attachments_queryset = (
            attachments_queryset.filter(
                uploaded_at__date__range=(
                    start_date,
                    end_date,
                )
            )
        )

        annotations_queryset = (
            annotations_queryset.filter(
                updated_at__date__range=(
                    start_date,
                    end_date,
                )
            )
        )

        surgery_date_field = (
            first_existing_model_field(
                SurgeryProcedure,
                [
                    "scheduled_date",
                    "procedure_date",
                    "surgery_date",
                    "created_at",
                ],
            )
        )

        if surgery_date_field:
            surgery_filter = {
                (
                    f"{surgery_date_field}__date__range"
                ): (
                    start_date,
                    end_date,
                )
            }

            surgeries_queryset = (
                surgeries_queryset.filter(
                    **surgery_filter
                )
            )

        appointment_date_field = (
            first_existing_model_field(
                Appointment,
                [
                    "appointment_date",
                    "scheduled_date",
                    "created_at",
                ],
            )
        )

        if appointment_date_field:
            appointment_filter = {
                (
                    f"{appointment_date_field}"
                    + (
                        "__date__range"
                        if appointment_date_field
                        != "appointment_date"
                        else "__range"
                    )
                ): (
                    start_date,
                    end_date,
                )
            }

            try:
                appointments_queryset = (
                    appointments_queryset.filter(
                        **appointment_filter
                    )
                )
            except Exception:
                # Retain the unfiltered appointment queryset when the
                # actual model stores this value in another compatible
                # date representation.
                pass

        contact_lens_assessments_queryset = (
            contact_lens_assessments_queryset.filter(
                assessment_date__date__range=(
                    start_date,
                    end_date,
                )
            )
        )

        contact_lens_trials_queryset = (
            contact_lens_trials_queryset.filter(
                trial_date__date__range=(
                    start_date,
                    end_date,
                )
            )
        )

        contact_lens_prescriptions_queryset = (
            contact_lens_prescriptions_queryset.filter(
                prescription_date__range=(
                    start_date,
                    end_date,
                )
            )
        )

        contact_lens_follow_ups_queryset = (
            contact_lens_follow_ups_queryset.filter(
                follow_up_date__date__range=(
                    start_date,
                    end_date,
                )
            )
        )

    # =====================================================
    # KPI COUNTS
    # =====================================================

    patient_count = patients_queryset.count()
    visit_count = visits_queryset.count()
    examination_count = (
        examinations_queryset.count()
    )
    surgery_count = surgeries_queryset.count()
    appointment_count = (
        appointments_queryset.count()
    )
    attachment_count = (
        attachments_queryset.count()
    )
    annotation_count = (
        annotations_queryset.count()
    )

    pending_attachment_count = (
        attachments_queryset.filter(
            review_status=(
                ClinicalAttachment
                .ReviewStatus
                .PENDING
            )
        ).count()
    )

    reviewed_attachment_count = (
        attachments_queryset.filter(
            review_status=(
                ClinicalAttachment
                .ReviewStatus
                .REVIEWED
            )
        ).count()
    )

    needs_attention_count = (
        attachments_queryset.filter(
            review_status=(
                ClinicalAttachment
                .ReviewStatus
                .NEEDS_ATTENTION
            )
        ).count()
    )

    finalized_annotation_count = (
        annotations_queryset.filter(
            status=(
                ClinicalImageAnnotation
                .AnnotationStatus
                .FINAL
            )
        ).count()
    )

    approved_refraction_count = (
        examinations_queryset.filter(
            subjective_refraction_approved=True,
        ).count()
    )

    # =====================================================
    # CONTACT LENS KPI COUNTS (reporting period)
    # =====================================================

    contact_lens_assessment_count = (
        contact_lens_assessments_queryset.count()
    )

    contact_lens_trial_count = (
        contact_lens_trials_queryset.count()
    )

    contact_lens_prescription_count = (
        contact_lens_prescriptions_queryset.count()
    )

    contact_lens_pending_approval_count = (
        contact_lens_prescriptions_queryset.filter(
            status=(
                ContactLensPrescription
                .PrescriptionStatus
                .PENDING_APPROVAL
            )
        ).count()
    )

    contact_lens_approved_count = (
        contact_lens_prescriptions_queryset.filter(
            status=(
                ContactLensPrescription
                .PrescriptionStatus
                .APPROVED
            )
        ).count()
    )

    contact_lens_dispensed_count = (
        contact_lens_prescriptions_queryset.filter(
            status=(
                ContactLensPrescription
                .PrescriptionStatus
                .DISPENSED
            )
        ).count()
    )

    contact_lens_follow_up_count = (
        contact_lens_follow_ups_queryset.count()
    )

    # =====================================================
    # CONTACT LENS OPERATIONAL QUEUE COUNTS
    #
    # These always reflect the current operational state and are
    # intentionally NOT limited to the selected reporting period.
    # =====================================================

    dashboard_today = timezone.localdate()

    contact_lens_overdue_follow_up_count = (
        ContactLensFollowUp.objects.filter(
            status=(
                ContactLensFollowUp
                .FollowUpStatus
                .SCHEDULED
            ),
            follow_up_date__date__lt=dashboard_today,
        ).count()
    )

    contact_lens_due_today_count = (
        ContactLensFollowUp.objects.filter(
            status=(
                ContactLensFollowUp
                .FollowUpStatus
                .SCHEDULED
            ),
            follow_up_date__date=dashboard_today,
        ).count()
    )

    contact_lens_due_soon_count = (
        ContactLensFollowUp.objects.filter(
            status=(
                ContactLensFollowUp
                .FollowUpStatus
                .SCHEDULED
            ),
            follow_up_date__date__gt=dashboard_today,
            follow_up_date__date__lte=(
                dashboard_today + timedelta(days=7)
            ),
        ).count()
    )

    contact_lens_complication_count = (
        ContactLensFollowUp.objects.filter(
            Q(complications__isnull=False)
            & ~Q(complications="")
        ).count()
    )

    # =====================================================
    # REVIEW QUEUES
    # =====================================================

    pending_review_queue = (
        attachments_queryset.filter(
            review_status=(
                ClinicalAttachment
                .ReviewStatus
                .PENDING
            )
        )
        .order_by(
            "uploaded_at",
            "pk",
        )[:10]
    )

    attention_queue = (
        attachments_queryset.filter(
            review_status=(
                ClinicalAttachment
                .ReviewStatus
                .NEEDS_ATTENTION
            )
        )
        .order_by(
            "-reviewed_at",
            "-updated_at",
        )[:10]
    )

    recent_finalized_annotations = (
        annotations_queryset.filter(
            status=(
                ClinicalImageAnnotation
                .AnnotationStatus
                .FINAL
            )
        )
        .order_by(
            "-finalized_at",
            "-updated_at",
        )[:8]
    )

    recent_examinations = (
        examinations_queryset.order_by(
            "-visit__visit_date",
            "-pk",
        )[:8]
    )

    # =====================================================
    # INVESTIGATION CATEGORY STATISTICS
    # =====================================================

    category_statistics = list(
        attachments_queryset.values(
            "category"
        )
        .annotate(
            total=Count("id")
        )
        .order_by("-total")
    )

    category_label_map = dict(
        ClinicalAttachment.Category.choices
    )

    for category_item in category_statistics:
        category_value = category_item[
            "category"
        ]

        category_item["label"] = (
            category_label_map.get(
                category_value,
                category_value,
            )
        )

    category_statistics = (
        category_statistics[:10]
    )

    # =====================================================
    # EYE-SIDE STATISTICS
    # =====================================================

    eye_side_statistics = list(
        attachments_queryset.values(
            "eye_side"
        )
        .annotate(
            total=Count("id")
        )
        .order_by("-total")
    )

    eye_side_label_map = dict(
        ClinicalAttachment.EyeSide.choices
    )

    for eye_item in eye_side_statistics:
        eye_value = eye_item["eye_side"]

        eye_item["label"] = (
            eye_side_label_map.get(
                eye_value,
                eye_value,
            )
        )

    # =====================================================
    # MONTHLY TREND — LAST 12 MONTHS
    # =====================================================

    trend_start_date = (
        timezone.localdate()
        - timedelta(days=365)
    )

    monthly_visit_data = list(
        PatientVisit.objects.filter(
            visit_date__date__gte=(
                trend_start_date
            )
        )
        .annotate(
            month=TruncMonth(
                "visit_date"
            )
        )
        .values("month")
        .annotate(
            total=Count("id")
        )
        .order_by("month")
    )

    monthly_attachment_data = list(
        ClinicalAttachment.objects.filter(
            is_active=True,
            uploaded_at__date__gte=(
                trend_start_date
            ),
        )
        .annotate(
            month=TruncMonth(
                "uploaded_at"
            )
        )
        .values("month")
        .annotate(
            total=Count("id")
        )
        .order_by("month")
    )

    monthly_examination_data = list(
        EyeExamination.objects.filter(
            visit__visit_date__date__gte=(
                trend_start_date
            )
        )
        .annotate(
            month=TruncMonth(
                "visit__visit_date"
            )
        )
        .values("month")
        .annotate(
            total=Count("id")
        )
        .order_by("month")
    )

    # Convert datetimes into JSON-safe labels.
    monthly_visit_labels = [
        item["month"].strftime("%b %Y")
        for item in monthly_visit_data
        if item["month"]
    ]

    monthly_visit_values = [
        item["total"]
        for item in monthly_visit_data
        if item["month"]
    ]

    monthly_attachment_labels = [
        item["month"].strftime("%b %Y")
        for item in monthly_attachment_data
        if item["month"]
    ]

    monthly_attachment_values = [
        item["total"]
        for item in monthly_attachment_data
        if item["month"]
    ]

    monthly_examination_labels = [
        item["month"].strftime("%b %Y")
        for item in monthly_examination_data
        if item["month"]
    ]

    monthly_examination_values = [
        item["total"]
        for item in monthly_examination_data
        if item["month"]
    ]

    # =====================================================
    # DIAGNOSIS SUMMARY
    # =====================================================

    diagnosis_name_field = (
        first_existing_model_field(
            DiagnosisTreatment,
            [
                "diagnosis",
                "diagnosis_details",
                "primary_diagnosis",
                "diagnosis_name",
            ],
        )
    )

    top_diagnoses = []

    if diagnosis_name_field:
        diagnosis_queryset = (
            DiagnosisTreatment.objects
            .exclude(
                **{
                    f"{diagnosis_name_field}": ""
                }
            )
        )

        if start_date and end_date:
            diagnosis_queryset = (
                diagnosis_queryset.filter(
                    visit__visit_date__date__range=(
                        start_date,
                        end_date,
                    )
                )
            )

        top_diagnoses = list(
            diagnosis_queryset.values(
                diagnosis_name_field
            )
            .annotate(
                total=Count("id")
            )
            .order_by("-total")[:10]
        )

        for diagnosis_item in top_diagnoses:
            diagnosis_item["label"] = (
                diagnosis_item.get(
                    diagnosis_name_field
                )
                or "Unspecified Diagnosis"
            )

    # =====================================================
    # CONTACT LENS PENDING APPROVAL QUEUE
    # =====================================================

    contact_lens_pending_approval_queue = (
        ContactLensPrescription.objects
        .filter(
            is_active=True,
            status=(
                ContactLensPrescription
                .PrescriptionStatus
                .PENDING_APPROVAL
            ),
        )
        .select_related(
            "patient",
            "visit",
            "prescribed_by",
        )
        .order_by(
            "created_at",
            "pk",
        )[:8]
    )

    # =====================================================
    # RECENT CONTACT LENS ACTIVITY
    # =====================================================

    recent_contact_lens_assessments = (
        ContactLensAssessment.objects
        .select_related(
            "patient",
            "visit",
            "assessed_by",
        )
        .order_by(
            "-assessment_date",
            "-pk",
        )[:6]
    )

    recent_contact_lens_prescriptions = (
        ContactLensPrescription.objects
        .filter(is_active=True)
        .select_related(
            "patient",
            "visit",
            "prescribed_by",
            "approved_by",
        )
        .order_by(
            "-created_at",
            "-pk",
        )[:6]
    )

    recent_contact_lens_follow_ups = (
        ContactLensFollowUp.objects
        .select_related(
            "patient",
            "prescription",
            "reviewed_by",
        )
        .order_by(
            "-follow_up_date",
            "-pk",
        )[:6]
    )

    # =====================================================
    # CONTACT LENS DESIGN STATISTICS
    # =====================================================

    right_lens_design_statistics = list(
        contact_lens_prescriptions_queryset
        .exclude(right_lens_design="")
        .values("right_lens_design")
        .annotate(total=Count("id"))
        .order_by("-total")
    )

    left_lens_design_statistics = list(
        contact_lens_prescriptions_queryset
        .exclude(left_lens_design="")
        .values("left_lens_design")
        .annotate(total=Count("id"))
        .order_by("-total")
    )

    lens_design_totals = {}

    lens_design_label_map = dict(
        ContactLensPrescription.LensDesign.choices
    )

    for item in (
        right_lens_design_statistics
        + left_lens_design_statistics
    ):
        design_value = (
            item.get("right_lens_design")
            or item.get("left_lens_design")
            or ""
        )

        if not design_value:
            continue

        lens_design_totals[design_value] = (
            lens_design_totals.get(design_value, 0)
            + item["total"]
        )

    contact_lens_design_statistics = [
        {
            "value": design_value,
            "label": lens_design_label_map.get(
                design_value,
                design_value,
            ),
            "total": total,
        }
        for design_value, total in sorted(
            lens_design_totals.items(),
            key=lambda item: item[1],
            reverse=True,
        )
    ][:8]

    # =====================================================
    # DASHBOARD CONTEXT
    # =====================================================

    context = {
        "selected_period": selected_period,
        "start_date": start_date,
        "end_date": end_date,

        "patient_count": patient_count,
        "visit_count": visit_count,
        "examination_count": examination_count,
        "surgery_count": surgery_count,
        "appointment_count": appointment_count,
        "attachment_count": attachment_count,
        "annotation_count": annotation_count,

        "pending_attachment_count": (
            pending_attachment_count
        ),
        "reviewed_attachment_count": (
            reviewed_attachment_count
        ),
        "needs_attention_count": (
            needs_attention_count
        ),
        "finalized_annotation_count": (
            finalized_annotation_count
        ),
        "approved_refraction_count": (
            approved_refraction_count
        ),

        "pending_review_queue": (
            pending_review_queue
        ),
        "attention_queue": attention_queue,
        "recent_finalized_annotations": (
            recent_finalized_annotations
        ),
        "recent_examinations": (
            recent_examinations
        ),

        "category_statistics": (
            category_statistics
        ),
        "eye_side_statistics": (
            eye_side_statistics
        ),
        "top_diagnoses": top_diagnoses,

        "monthly_visit_labels": (
            monthly_visit_labels
        ),
        "monthly_visit_values": (
            monthly_visit_values
        ),
        "monthly_attachment_labels": (
            monthly_attachment_labels
        ),
        "monthly_attachment_values": (
            monthly_attachment_values
        ),
        "monthly_examination_labels": (
            monthly_examination_labels
        ),
        "monthly_examination_values": (
            monthly_examination_values
        ),

        "can_review_attachments": (
            user_can_review_clinical_attachments(
                request.user
            )
        ),

        "contact_lens_assessment_count": (
            contact_lens_assessment_count
        ),
        "contact_lens_trial_count": (
            contact_lens_trial_count
        ),
        "contact_lens_prescription_count": (
            contact_lens_prescription_count
        ),
        "contact_lens_pending_approval_count": (
            contact_lens_pending_approval_count
        ),
        "contact_lens_approved_count": (
            contact_lens_approved_count
        ),
        "contact_lens_dispensed_count": (
            contact_lens_dispensed_count
        ),
        "contact_lens_follow_up_count": (
            contact_lens_follow_up_count
        ),
        "contact_lens_overdue_follow_up_count": (
            contact_lens_overdue_follow_up_count
        ),
        "contact_lens_due_today_count": (
            contact_lens_due_today_count
        ),
        "contact_lens_due_soon_count": (
            contact_lens_due_soon_count
        ),
        "contact_lens_complication_count": (
            contact_lens_complication_count
        ),
        "contact_lens_pending_approval_queue": (
            contact_lens_pending_approval_queue
        ),
        "recent_contact_lens_assessments": (
            recent_contact_lens_assessments
        ),
        "recent_contact_lens_prescriptions": (
            recent_contact_lens_prescriptions
        ),
        "recent_contact_lens_follow_ups": (
            recent_contact_lens_follow_ups
        ),
        "contact_lens_design_statistics": (
            contact_lens_design_statistics
        ),
        "can_manage_contact_lenses": (
            user_can_manage_contact_lenses(
                request.user
            )
        ),
        "can_approve_contact_lenses": (
            user_can_approve_contact_lens_prescriptions(
                request.user
            )
        ),
    }

    return render(
        request,
        "patients/ophthalmology_dashboard.html",
        context,
    )


@login_required
@clinical_staff_required
def contact_lens_assessment_detail(request, pk):
    assessment = get_object_or_404(
        get_contact_lens_assessment_queryset(),
        pk=pk,
    )

    trials = (
        assessment.trial_lenses
        .select_related("fitted_by")
        .order_by(
            "eye_side",
            "trial_number",
            "pk",
        )
    )

    prescriptions = (
        assessment.prescriptions
        .filter(is_active=True)
        .select_related(
            "prescribed_by",
            "approved_by",
            "dispensed_by",
        )
        .order_by(
            "-version",
            "-prescription_date",
            "-pk",
        )
    )

    accepted_right_trial = (
        trials.filter(
            eye_side=ContactLensTrial.EyeSide.RIGHT,
            accepted_for_prescription=True,
        )
        .order_by("-trial_number")
        .first()
    )

    accepted_left_trial = (
        trials.filter(
            eye_side=ContactLensTrial.EyeSide.LEFT,
            accepted_for_prescription=True,
        )
        .order_by("-trial_number")
        .first()
    )

    return render(
        request,
        "patients/contact_lens/assessment_detail.html",
        {
            "assessment": assessment,
            "patient": assessment.patient,
            "visit": assessment.visit,
            "trials": trials,
            "prescriptions": prescriptions,
            "accepted_right_trial": (
                accepted_right_trial
            ),
            "accepted_left_trial": (
                accepted_left_trial
            ),
            "can_manage_contact_lenses": (
                user_can_manage_contact_lenses(
                    request.user
                )
            ),
            "can_approve_contact_lenses": (
                user_can_approve_contact_lens_prescriptions(
                    request.user
                )
            ),
        },
    )


@login_required
@clinical_staff_required
def contact_lens_assessment_create(
    request,
    visit_pk,
):
    ensure_contact_lens_management_permission(
        request.user
    )

    visit = get_object_or_404(
        PatientVisit.objects.select_related(
            "patient",
            "created_by",
        ),
        pk=visit_pk,
    )

    existing_assessment = (
        ContactLensAssessment.objects
        .filter(visit=visit)
        .first()
    )

    if existing_assessment:
        messages.info(
            request,
            (
                "A Contact Lens assessment already exists for "
                "this visit. The existing assessment has been opened."
            ),
        )

        return redirect(
            "contact_lens_assessment_detail",
            pk=existing_assessment.pk,
        )

    if request.method == "POST":
        form = ContactLensAssessmentForm(
            request.POST,
            visit=visit,
            request_user=request.user,
        )

        if form.is_valid():
            with transaction.atomic():
                assessment = form.save()

                log_activity(
                    request,
                    AuditLog.ActionType.CREATE,
                    "Contact Lens",
                    (
                        "Created Contact Lens assessment for "
                        f"{assessment.patient.file_number}, "
                        f"visit {assessment.visit.visit_number}."
                    ),
                    object_id=assessment.pk,
                    object_repr=str(assessment),
                )

            messages.success(
                request,
                "Contact Lens assessment created successfully.",
            )

            return redirect(
                "contact_lens_assessment_detail",
                pk=assessment.pk,
            )

    else:
        form = ContactLensAssessmentForm(
            visit=visit,
            request_user=request.user,
        )

    return render(
        request,
        "patients/contact_lens/assessment_form.html",
        {
            "form": form,
            "visit": visit,
            "patient": visit.patient,
            "page_title": (
                "New Contact Lens Assessment"
            ),
            "submit_label": (
                "Save Contact Lens Assessment"
            ),
            "is_update": False,
        },
    )


@login_required
@clinical_staff_required
def contact_lens_assessment_update(
    request,
    pk,
):
    ensure_contact_lens_management_permission(
        request.user
    )

    assessment = get_object_or_404(
        get_contact_lens_assessment_queryset(),
        pk=pk,
    )

    if request.method == "POST":
        form = ContactLensAssessmentForm(
            request.POST,
            instance=assessment,
            visit=assessment.visit,
            request_user=request.user,
        )

        if form.is_valid():
            with transaction.atomic():
                updated_assessment = form.save()

                log_activity(
                    request,
                    AuditLog.ActionType.UPDATE,
                    "Contact Lens",
                    (
                        "Updated Contact Lens assessment for "
                        f"{updated_assessment.patient.file_number}, "
                        f"visit "
                        f"{updated_assessment.visit.visit_number}."
                    ),
                    object_id=updated_assessment.pk,
                    object_repr=str(
                        updated_assessment
                    ),
                )

            messages.success(
                request,
                "Contact Lens assessment updated successfully.",
            )

            return redirect(
                "contact_lens_assessment_detail",
                pk=updated_assessment.pk,
            )

    else:
        form = ContactLensAssessmentForm(
            instance=assessment,
            visit=assessment.visit,
            request_user=request.user,
        )

    return render(
        request,
        "patients/contact_lens/assessment_form.html",
        {
            "form": form,
            "assessment": assessment,
            "visit": assessment.visit,
            "patient": assessment.patient,
            "page_title": (
                "Edit Contact Lens Assessment"
            ),
            "submit_label": (
                "Update Contact Lens Assessment"
            ),
            "is_update": True,
        },
    )


@login_required
@clinical_staff_required
def contact_lens_trial_create(
    request,
    assessment_pk,
):
    ensure_contact_lens_management_permission(
        request.user
    )

    assessment = get_object_or_404(
        get_contact_lens_assessment_queryset(),
        pk=assessment_pk,
    )

    if (
        assessment.suitability_status
        in {
            ContactLensAssessment
            .SuitabilityStatus
            .UNSUITABLE,

            ContactLensAssessment
            .SuitabilityStatus
            .TEMPORARILY_UNSUITABLE,
        }
    ):
        messages.warning(
            request,
            (
                "This patient is currently marked as unsuitable "
                "for Contact Lens fitting. Update the suitability "
                "assessment before recording a trial lens."
            ),
        )

        return redirect(
            "contact_lens_assessment_detail",
            pk=assessment.pk,
        )

    if request.method == "POST":
        form = ContactLensTrialForm(
            request.POST,
            assessment=assessment,
            request_user=request.user,
        )

        if form.is_valid():
            with transaction.atomic():
                trial = form.save()

                if trial.accepted_for_prescription:
                    (
                        ContactLensTrial.objects
                        .filter(
                            assessment=assessment,
                            eye_side=trial.eye_side,
                            accepted_for_prescription=True,
                        )
                        .exclude(pk=trial.pk)
                        .update(
                            accepted_for_prescription=False
                        )
                    )

                log_activity(
                    request,
                    AuditLog.ActionType.CREATE,
                    "Contact Lens",
                    (
                        f"Created Contact Lens trial "
                        f"{trial.trial_number} for "
                        f"{trial.get_eye_side_display()}, patient "
                        f"{assessment.patient.file_number}."
                    ),
                    object_id=trial.pk,
                    object_repr=str(trial),
                )

            messages.success(
                request,
                "Contact Lens trial recorded successfully.",
            )

            return redirect(
                "contact_lens_assessment_detail",
                pk=assessment.pk,
            )

    else:
        form = ContactLensTrialForm(
            assessment=assessment,
            request_user=request.user,
        )

    return render(
        request,
        "patients/contact_lens/trial_form.html",
        {
            "form": form,
            "assessment": assessment,
            "patient": assessment.patient,
            "visit": assessment.visit,
            "page_title": "New Trial Lens",
            "submit_label": "Save Trial Lens",
            "is_update": False,
        },
    )


@login_required
@clinical_staff_required
def contact_lens_trial_update(
    request,
    pk,
):
    ensure_contact_lens_management_permission(
        request.user
    )

    trial = get_object_or_404(
        get_contact_lens_trial_queryset(),
        pk=pk,
    )

    assessment = trial.assessment

    if request.method == "POST":
        form = ContactLensTrialForm(
            request.POST,
            instance=trial,
            assessment=assessment,
            request_user=request.user,
        )

        if form.is_valid():
            with transaction.atomic():
                updated_trial = form.save()

                if updated_trial.accepted_for_prescription:
                    (
                        ContactLensTrial.objects
                        .filter(
                            assessment=assessment,
                            eye_side=updated_trial.eye_side,
                            accepted_for_prescription=True,
                        )
                        .exclude(pk=updated_trial.pk)
                        .update(
                            accepted_for_prescription=False
                        )
                    )

                log_activity(
                    request,
                    AuditLog.ActionType.UPDATE,
                    "Contact Lens",
                    (
                        f"Updated Contact Lens trial "
                        f"{updated_trial.trial_number} for "
                        f"{updated_trial.get_eye_side_display()}, "
                        f"patient "
                        f"{assessment.patient.file_number}."
                    ),
                    object_id=updated_trial.pk,
                    object_repr=str(
                        updated_trial
                    ),
                )

            messages.success(
                request,
                "Contact Lens trial updated successfully.",
            )

            return redirect(
                "contact_lens_assessment_detail",
                pk=assessment.pk,
            )

    else:
        form = ContactLensTrialForm(
            instance=trial,
            assessment=assessment,
            request_user=request.user,
        )

    return render(
        request,
        "patients/contact_lens/trial_form.html",
        {
            "form": form,
            "trial": trial,
            "assessment": assessment,
            "patient": assessment.patient,
            "visit": assessment.visit,
            "page_title": "Edit Trial Lens",
            "submit_label": "Update Trial Lens",
            "is_update": True,
        },
    )


@login_required
@clinical_staff_required
def contact_lens_prescription_detail(
    request,
    pk,
):
    prescription = get_object_or_404(
        get_contact_lens_prescription_queryset(
            include_inactive=True
        ),
        pk=pk,
    )

    follow_ups = (
        prescription.follow_ups
        .select_related(
            "reviewed_by",
        )
        .order_by(
            "-follow_up_date",
            "-pk",
        )
    )

    return render(
        request,
        "patients/contact_lens/prescription_detail.html",
        {
            "prescription": prescription,
            "assessment": prescription.assessment,
            "patient": prescription.patient,
            "visit": prescription.visit,
            "follow_ups": follow_ups,
            "can_manage_contact_lenses": (
                user_can_manage_contact_lenses(
                    request.user
                )
            ),
            "can_approve_contact_lenses": (
                user_can_approve_contact_lens_prescriptions(
                    request.user
                )
            ),
            "can_dispense_contact_lenses": (
                user_can_dispense_contact_lens_prescriptions(
                    request.user
                )
            ),
        },
    )


@login_required
@clinical_staff_required
def contact_lens_prescription_create(
    request,
    assessment_pk,
):
    ensure_contact_lens_management_permission(
        request.user
    )

    assessment = get_object_or_404(
        get_contact_lens_assessment_queryset(),
        pk=assessment_pk,
    )

    accepted_trial_exists = (
        assessment.trial_lenses
        .filter(
            accepted_for_prescription=True
        )
        .exists()
    )

    if not accepted_trial_exists:
        messages.warning(
            request,
            (
                "Accept at least one right-eye or left-eye trial "
                "before generating a Contact Lens prescription."
            ),
        )

        return redirect(
            "contact_lens_assessment_detail",
            pk=assessment.pk,
        )

    if request.method == "POST":
        form = ContactLensPrescriptionForm(
            request.POST,
            assessment=assessment,
            visit=assessment.visit,
            request_user=request.user,
        )

        if form.is_valid():
            with transaction.atomic():
                latest_version = (
                    ContactLensPrescription.objects
                    .filter(
                        assessment=assessment
                    )
                    .aggregate(
                        maximum_version=Max(
                            "version"
                        )
                    )
                    .get("maximum_version")
                    or 0
                )

                prescription = form.save(
                    commit=False
                )

                prescription.version = (
                    latest_version + 1
                )

                prescription.status = (
                    ContactLensPrescription
                    .PrescriptionStatus
                    .DRAFT
                )

                prescription.save()
                form.save_m2m()

                log_activity(
                    request,
                    AuditLog.ActionType.CREATE,
                    "Contact Lens",
                    (
                        "Created Contact Lens prescription "
                        f"{prescription.prescription_number}, "
                        f"version {prescription.version}, for "
                        f"patient "
                        f"{prescription.patient.file_number}."
                    ),
                    object_id=prescription.pk,
                    object_repr=str(
                        prescription
                    ),
                )

            messages.success(
                request,
                (
                    "Contact Lens prescription created as a draft. "
                    "Review it and submit it for approval."
                ),
            )

            return redirect(
                "contact_lens_prescription_detail",
                pk=prescription.pk,
            )

    else:
        form = ContactLensPrescriptionForm(
            assessment=assessment,
            visit=assessment.visit,
            request_user=request.user,
        )

    return render(
        request,
        "patients/contact_lens/prescription_form.html",
        {
            "form": form,
            "assessment": assessment,
            "patient": assessment.patient,
            "visit": assessment.visit,
            "page_title": (
                "Create Contact Lens Prescription"
            ),
            "submit_label": (
                "Save Draft Prescription"
            ),
            "is_update": False,
        },
    )


@login_required
@clinical_staff_required
def contact_lens_prescription_update(
    request,
    pk,
):
    ensure_contact_lens_management_permission(
        request.user
    )

    prescription = get_object_or_404(
        get_contact_lens_prescription_queryset(),
        pk=pk,
    )

    editable_statuses = {
        ContactLensPrescription
        .PrescriptionStatus
        .DRAFT,

        ContactLensPrescription
        .PrescriptionStatus
        .PENDING_APPROVAL,
    }

    if prescription.status not in editable_statuses:
        messages.warning(
            request,
            (
                "Approved, dispensed, expired or cancelled "
                "prescriptions cannot be edited. Create a new "
                "prescription version instead."
            ),
        )

        return redirect(
            "contact_lens_prescription_detail",
            pk=prescription.pk,
        )

    if request.method == "POST":
        form = ContactLensPrescriptionForm(
            request.POST,
            instance=prescription,
            assessment=prescription.assessment,
            visit=prescription.visit,
            request_user=request.user,
        )

        if form.is_valid():
            with transaction.atomic():
                updated_prescription = form.save(
                    commit=False
                )

                # Editing a pending prescription returns it to draft
                # so it must be submitted and reviewed again.
                updated_prescription.status = (
                    ContactLensPrescription
                    .PrescriptionStatus
                    .DRAFT
                )

                updated_prescription.approved_by = None
                updated_prescription.approved_at = None

                updated_prescription.save()
                form.save_m2m()

                log_activity(
                    request,
                    AuditLog.ActionType.UPDATE,
                    "Contact Lens",
                    (
                        "Updated Contact Lens prescription "
                        f"{updated_prescription.prescription_number}, "
                        f"version "
                        f"{updated_prescription.version}."
                    ),
                    object_id=updated_prescription.pk,
                    object_repr=str(
                        updated_prescription
                    ),
                )

            messages.success(
                request,
                (
                    "Contact Lens prescription updated. "
                    "It remains a draft until resubmitted."
                ),
            )

            return redirect(
                "contact_lens_prescription_detail",
                pk=updated_prescription.pk,
            )

    else:
        form = ContactLensPrescriptionForm(
            instance=prescription,
            assessment=prescription.assessment,
            visit=prescription.visit,
            request_user=request.user,
        )

    return render(
        request,
        "patients/contact_lens/prescription_form.html",
        {
            "form": form,
            "prescription": prescription,
            "assessment": prescription.assessment,
            "patient": prescription.patient,
            "visit": prescription.visit,
            "page_title": (
                "Edit Contact Lens Prescription"
            ),
            "submit_label": (
                "Update Draft Prescription"
            ),
            "is_update": True,
        },
    )


@login_required
@clinical_staff_required
@require_POST
def contact_lens_prescription_submit(
    request,
    pk,
):
    ensure_contact_lens_management_permission(
        request.user
    )

    prescription = get_object_or_404(
        get_contact_lens_prescription_queryset(),
        pk=pk,
    )

    if (
        prescription.status
        != ContactLensPrescription
        .PrescriptionStatus
        .DRAFT
    ):
        messages.warning(
            request,
            (
                "Only a draft Contact Lens prescription "
                "can be submitted for approval."
            ),
        )

        return redirect(
            "contact_lens_prescription_detail",
            pk=prescription.pk,
        )

    right_has_parameters = any(
        value not in {
            None,
            "",
        }
        for value in [
            prescription.right_lens_design,
            prescription.right_sphere,
            prescription.right_base_curve,
            prescription.right_diameter,
        ]
    )

    left_has_parameters = any(
        value not in {
            None,
            "",
        }
        for value in [
            prescription.left_lens_design,
            prescription.left_sphere,
            prescription.left_base_curve,
            prescription.left_diameter,
        ]
    )

    if not right_has_parameters and not left_has_parameters:
        messages.error(
            request,
            (
                "The prescription does not contain sufficient "
                "right-eye or left-eye lens parameters."
            ),
        )

        return redirect(
            "contact_lens_prescription_update",
            pk=prescription.pk,
        )

    prescription.status = (
        ContactLensPrescription
        .PrescriptionStatus
        .PENDING_APPROVAL
    )

    prescription.save(
        update_fields=[
            "status",
            "updated_at",
        ]
    )

    log_activity(
        request,
        AuditLog.ActionType.UPDATE,
        "Contact Lens",
        (
            "Submitted Contact Lens prescription "
            f"{prescription.prescription_number} "
            "for clinical approval."
        ),
        object_id=prescription.pk,
        object_repr=str(prescription),
    )

    messages.success(
        request,
        "Contact Lens prescription submitted for approval.",
    )

    return redirect(
        "contact_lens_prescription_detail",
        pk=prescription.pk,
    )


@login_required
@clinical_staff_required
def contact_lens_prescription_approve(
    request,
    pk,
):
    ensure_contact_lens_approval_permission(
        request.user
    )

    prescription = get_object_or_404(
        get_contact_lens_prescription_queryset(),
        pk=pk,
    )

    if (
        prescription.status
        != ContactLensPrescription
        .PrescriptionStatus
        .PENDING_APPROVAL
    ):
        messages.warning(
            request,
            (
                "Only a prescription awaiting approval "
                "can be approved."
            ),
        )

        return redirect(
            "contact_lens_prescription_detail",
            pk=prescription.pk,
        )

    if request.method == "POST":
        form = ContactLensPrescriptionApprovalForm(
            request.POST,
            prescription=prescription,
            request_user=request.user,
        )

        if form.is_valid():
            with transaction.atomic():
                prescription.status = (
                    ContactLensPrescription
                    .PrescriptionStatus
                    .APPROVED
                )

                prescription.approved_by = (
                    request.user
                )

                prescription.approved_at = (
                    timezone.now()
                )

                approval_note = (
                    form.cleaned_data.get(
                        "approval_note"
                    )
                    or ""
                ).strip()

                if approval_note:
                    existing_notes = (
                        prescription.clinical_notes
                        or ""
                    ).strip()

                    approval_entry = (
                        "\n\nApproval Note:\n"
                        f"{approval_note}"
                    )

                    prescription.clinical_notes = (
                        existing_notes
                        + approval_entry
                    ).strip()

                prescription.save(
                    update_fields=[
                        "status",
                        "approved_by",
                        "approved_at",
                        "clinical_notes",
                        "updated_at",
                    ]
                )

                log_activity(
                    request,
                    AuditLog.ActionType.UPDATE,
                    "Contact Lens",
                    (
                        "Approved Contact Lens prescription "
                        f"{prescription.prescription_number} "
                        f"for patient "
                        f"{prescription.patient.file_number}."
                    ),
                    object_id=prescription.pk,
                    object_repr=str(
                        prescription
                    ),
                )

            messages.success(
                request,
                (
                    "Contact Lens prescription approved "
                    "successfully."
                ),
            )

            return redirect(
                "contact_lens_prescription_detail",
                pk=prescription.pk,
            )

    else:
        form = (
            ContactLensPrescriptionApprovalForm(
                prescription=prescription,
                request_user=request.user,
            )
        )

    return render(
        request,
        "patients/contact_lens/prescription_approve.html",
        {
            "form": form,
            "prescription": prescription,
            "patient": prescription.patient,
            "assessment": prescription.assessment,
        },
    )


@login_required
@clinical_staff_required
def contact_lens_prescription_dispense(
    request,
    pk,
):
    ensure_contact_lens_dispensing_permission(
        request.user
    )

    prescription = get_object_or_404(
        get_contact_lens_prescription_queryset(),
        pk=pk,
    )

    if (
        prescription.status
        != ContactLensPrescription
        .PrescriptionStatus
        .APPROVED
    ):
        messages.warning(
            request,
            (
                "Only an approved Contact Lens prescription "
                "can be dispensed."
            ),
        )

        return redirect(
            "contact_lens_prescription_detail",
            pk=prescription.pk,
        )

    if request.method == "POST":
        form = (
            ContactLensPrescriptionDispensingForm(
                request.POST,
                prescription=prescription,
                request_user=request.user,
            )
        )

        if form.is_valid():
            with transaction.atomic():
                prescription.status = (
                    ContactLensPrescription
                    .PrescriptionStatus
                    .DISPENSED
                )

                prescription.dispensed_by = (
                    request.user
                )

                prescription.dispensed_at = (
                    timezone.now()
                )

                dispensing_note = (
                    form.cleaned_data.get(
                        "dispensing_note"
                    )
                    or ""
                ).strip()

                if dispensing_note:
                    existing_instructions = (
                        prescription
                        .dispensing_instructions
                        or ""
                    ).strip()

                    dispensing_entry = (
                        "\n\nDispensing Note:\n"
                        f"{dispensing_note}"
                    )

                    prescription.dispensing_instructions = (
                        existing_instructions
                        + dispensing_entry
                    ).strip()

                prescription.save(
                    update_fields=[
                        "status",
                        "dispensed_by",
                        "dispensed_at",
                        "dispensing_instructions",
                        "updated_at",
                    ]
                )

                log_activity(
                    request,
                    AuditLog.ActionType.UPDATE,
                    "Contact Lens",
                    (
                        "Dispensed Contact Lens prescription "
                        f"{prescription.prescription_number} "
                        f"for patient "
                        f"{prescription.patient.file_number}."
                    ),
                    object_id=prescription.pk,
                    object_repr=str(
                        prescription
                    ),
                )

            messages.success(
                request,
                (
                    "Contact Lens prescription marked as "
                    "dispensed."
                ),
            )

            return redirect(
                "contact_lens_prescription_detail",
                pk=prescription.pk,
            )

    else:
        form = (
            ContactLensPrescriptionDispensingForm(
                prescription=prescription,
                request_user=request.user,
            )
        )

    return render(
        request,
        "patients/contact_lens/prescription_dispense.html",
        {
            "form": form,
            "prescription": prescription,
            "patient": prescription.patient,
            "assessment": prescription.assessment,
        },
    )


@login_required
@clinical_staff_required
@require_POST
def contact_lens_prescription_new_version(
    request,
    pk,
):
    ensure_contact_lens_management_permission(
        request.user
    )

    source_prescription = get_object_or_404(
        get_contact_lens_prescription_queryset(),
        pk=pk,
    )

    with transaction.atomic():
        latest_version = (
            ContactLensPrescription.objects
            .filter(
                assessment=(
                    source_prescription.assessment
                )
            )
            .aggregate(
                maximum_version=Max("version")
            )
            .get("maximum_version")
            or 0
        )

        copied_field_names = [
            "right_lens_design",
            "right_lens_design_other",
            "right_manufacturer",
            "right_brand_name",
            "right_material",
            "right_base_curve",
            "right_diameter",
            "right_sphere",
            "right_cylinder",
            "right_axis",
            "right_add_power",
            "right_colour",

            "left_lens_design",
            "left_lens_design_other",
            "left_manufacturer",
            "left_brand_name",
            "left_material",
            "left_base_curve",
            "left_diameter",
            "left_sphere",
            "left_cylinder",
            "left_axis",
            "left_add_power",
            "left_colour",

            "replacement_schedule",
            "replacement_schedule_other",
            "wearing_schedule",
            "maximum_daily_wear_hours",

            "cleaning_solution",
            "cleaning_instructions",

            "insertion_removal_training_completed",
            "hygiene_training_completed",
            "emergency_warning_signs_explained",

            "clinical_notes",
            "dispensing_instructions",
        ]

        copied_values = {
            field_name: getattr(
                source_prescription,
                field_name,
            )
            for field_name in copied_field_names
        }

        new_prescription = (
            ContactLensPrescription.objects.create(
                assessment=(
                    source_prescription.assessment
                ),
                patient=source_prescription.patient,
                visit=source_prescription.visit,
                version=latest_version + 1,
                prescription_date=(
                    timezone.localdate()
                ),
                valid_until=(
                    source_prescription.valid_until
                ),
                status=(
                    ContactLensPrescription
                    .PrescriptionStatus
                    .DRAFT
                ),
                prescribed_by=request.user,
                **copied_values,
            )
        )

        log_activity(
            request,
            AuditLog.ActionType.CREATE,
            "Contact Lens",
            (
                "Created Contact Lens prescription "
                f"{new_prescription.prescription_number}, "
                f"version {new_prescription.version}, from "
                f"{source_prescription.prescription_number}."
            ),
            object_id=new_prescription.pk,
            object_repr=str(
                new_prescription
            ),
        )

    messages.success(
        request,
        (
            "A new draft prescription version was created. "
            "The previous prescription remains unchanged."
        ),
    )

    return redirect(
        "contact_lens_prescription_update",
        pk=new_prescription.pk,
    )


@login_required
@clinical_staff_required
def contact_lens_follow_up_create(
    request,
    prescription_pk,
):
    ensure_contact_lens_management_permission(
        request.user
    )

    prescription = get_object_or_404(
        get_contact_lens_prescription_queryset(),
        pk=prescription_pk,
    )

    allowed_statuses = {
        ContactLensPrescription
        .PrescriptionStatus
        .APPROVED,

        ContactLensPrescription
        .PrescriptionStatus
        .DISPENSED,
    }

    if prescription.status not in allowed_statuses:
        messages.warning(
            request,
            (
                "Follow-up reviews may be created only for "
                "approved or dispensed prescriptions."
            ),
        )

        return redirect(
            "contact_lens_prescription_detail",
            pk=prescription.pk,
        )

    if request.method == "POST":
        form = ContactLensFollowUpForm(
            request.POST,
            prescription=prescription,
            request_user=request.user,
        )

        if form.is_valid():
            with transaction.atomic():
                follow_up = form.save()

                log_activity(
                    request,
                    AuditLog.ActionType.CREATE,
                    "Contact Lens",
                    (
                        "Created Contact Lens follow-up for "
                        f"{prescription.prescription_number}, "
                        f"patient "
                        f"{prescription.patient.file_number}."
                    ),
                    object_id=follow_up.pk,
                    object_repr=str(follow_up),
                )

            messages.success(
                request,
                "Contact Lens follow-up saved successfully.",
            )

            return redirect(
                "contact_lens_prescription_detail",
                pk=prescription.pk,
            )

    else:
        form = ContactLensFollowUpForm(
            prescription=prescription,
            request_user=request.user,
        )

    return render(
        request,
        "patients/contact_lens/follow_up_form.html",
        {
            "form": form,
            "prescription": prescription,
            "patient": prescription.patient,
            "assessment": prescription.assessment,
            "page_title": (
                "New Contact Lens Follow-up"
            ),
            "submit_label": (
                "Save Follow-up"
            ),
            "is_update": False,
        },
    )


@login_required
@clinical_staff_required
def contact_lens_follow_up_update(
    request,
    pk,
):
    ensure_contact_lens_management_permission(
        request.user
    )

    follow_up = get_object_or_404(
        get_contact_lens_follow_up_queryset(),
        pk=pk,
    )

    prescription = follow_up.prescription

    if request.method == "POST":
        form = ContactLensFollowUpForm(
            request.POST,
            instance=follow_up,
            prescription=prescription,
            request_user=request.user,
        )

        if form.is_valid():
            with transaction.atomic():
                updated_follow_up = form.save()

                log_activity(
                    request,
                    AuditLog.ActionType.UPDATE,
                    "Contact Lens",
                    (
                        "Updated Contact Lens follow-up for "
                        f"{prescription.prescription_number}, "
                        f"patient "
                        f"{prescription.patient.file_number}."
                    ),
                    object_id=updated_follow_up.pk,
                    object_repr=str(
                        updated_follow_up
                    ),
                )

            messages.success(
                request,
                "Contact Lens follow-up updated successfully.",
            )

            return redirect(
                "contact_lens_prescription_detail",
                pk=prescription.pk,
            )

    else:
        form = ContactLensFollowUpForm(
            instance=follow_up,
            prescription=prescription,
            request_user=request.user,
        )

    return render(
        request,
        "patients/contact_lens/follow_up_form.html",
        {
            "form": form,
            "follow_up": follow_up,
            "prescription": prescription,
            "patient": prescription.patient,
            "assessment": prescription.assessment,
            "page_title": (
                "Edit Contact Lens Follow-up"
            ),
            "submit_label": (
                "Update Follow-up"
            ),
            "is_update": True,
        },
    )


@login_required
@clinical_staff_required
def contact_lens_follow_up_detail(
    request,
    pk,
):
    follow_up = get_object_or_404(
        get_contact_lens_follow_up_queryset(),
        pk=pk,
    )

    return render(
        request,
        "patients/contact_lens/follow_up_detail.html",
        {
            "follow_up": follow_up,
            "prescription": follow_up.prescription,
            "assessment": (
                follow_up.prescription.assessment
            ),
            "patient": follow_up.patient,
            "can_manage_contact_lenses": (
                user_can_manage_contact_lenses(
                    request.user
                )
            ),
        },
    )


@login_required
@clinical_staff_required
def contact_lens_follow_up_queue(request):
    """
    Operational Contact Lens follow-up and clinical safety queue.
    """

    today = timezone.localdate()

    selected_queue = (
        request.GET.get(
            "queue",
            "active",
        )
        .strip()
        .lower()
    )

    valid_queues = {
        "active",
        "overdue",
        "today",
        "due_soon",
        "upcoming",
        "completed",
        "missed",
        "complications",
        "all",
    }

    if selected_queue not in valid_queues:
        selected_queue = "active"

    selected_status = (
        request.GET.get(
            "status",
            "",
        )
        .strip()
        .upper()
    )

    valid_statuses = {
        value
        for value, _label
        in ContactLensFollowUp
        .FollowUpStatus
        .choices
    }

    if selected_status not in valid_statuses:
        selected_status = ""

    query = (
        request.GET.get(
            "q",
            "",
        )
        .strip()
    )

    base_queryset = (
        ContactLensFollowUp.objects
        .select_related(
            "prescription",
            "prescription__assessment",
            "prescription__visit",
            "patient",
            "reviewed_by",
        )
        .order_by(
            "follow_up_date",
            "pk",
        )
    )

    if selected_status:
        base_queryset = base_queryset.filter(
            status=selected_status
        )

    if query:
        base_queryset = base_queryset.filter(
            Q(
                patient__first_name__icontains=query
            )
            | Q(
                patient__last_name__icontains=query
            )
            | Q(
                patient__file_number__icontains=query
            )
            | Q(
                prescription__prescription_number__icontains=query
            )
            | Q(
                prescription__visit__visit_number__icontains=query
            )
            | Q(
                complications__icontains=query
            )
            | Q(
                management_plan__icontains=query
            )
        )

    active_status = (
        ContactLensFollowUp
        .FollowUpStatus
        .SCHEDULED
    )

    if selected_queue == "active":
        base_queryset = base_queryset.filter(
            status=active_status,
            follow_up_date__date__lte=(
                today + timedelta(days=7)
            ),
        )

    elif selected_queue == "overdue":
        base_queryset = base_queryset.filter(
            status=active_status,
            follow_up_date__date__lt=today,
        )

    elif selected_queue == "today":
        base_queryset = base_queryset.filter(
            status=active_status,
            follow_up_date__date=today,
        )

    elif selected_queue == "due_soon":
        base_queryset = base_queryset.filter(
            status=active_status,
            follow_up_date__date__gt=today,
            follow_up_date__date__lte=(
                today + timedelta(days=7)
            ),
        )

    elif selected_queue == "upcoming":
        base_queryset = base_queryset.filter(
            status=active_status,
            follow_up_date__date__gt=(
                today + timedelta(days=7)
            ),
        )

    elif selected_queue == "completed":
        base_queryset = base_queryset.filter(
            status=(
                ContactLensFollowUp
                .FollowUpStatus
                .COMPLETED
            )
        ).order_by(
            "-follow_up_date",
            "-pk",
        )

    elif selected_queue == "missed":
        base_queryset = base_queryset.filter(
            status=(
                ContactLensFollowUp
                .FollowUpStatus
                .MISSED
            )
        ).order_by(
            "-follow_up_date",
            "-pk",
        )

    elif selected_queue == "complications":
        base_queryset = base_queryset.filter(
            Q(
                complications__isnull=False
            )
            & ~Q(
                complications=""
            )
        ).order_by(
            "-follow_up_date",
            "-pk",
        )

    elif selected_queue == "all":
        base_queryset = base_queryset.order_by(
            "-follow_up_date",
            "-pk",
        )

    paginator = Paginator(
        base_queryset,
        25,
    )

    page_obj = paginator.get_page(
        request.GET.get("page")
    )

    follow_up_rows = []

    for follow_up in page_obj.object_list:
        due_category = (
            contact_lens_follow_up_due_category(
                follow_up,
                today=today,
            )
        )

        safety_flags = (
            contact_lens_follow_up_safety_flags(
                follow_up
            )
        )

        follow_up_rows.append(
            {
                "record": follow_up,
                "due_category": due_category,
                "safety_flags": safety_flags,
                "has_safety_flags": bool(
                    safety_flags
                ),
            }
        )

    queue_counts = {
        "overdue": (
            ContactLensFollowUp.objects.filter(
                status=active_status,
                follow_up_date__date__lt=today,
            ).count()
        ),

        "today": (
            ContactLensFollowUp.objects.filter(
                status=active_status,
                follow_up_date__date=today,
            ).count()
        ),

        "due_soon": (
            ContactLensFollowUp.objects.filter(
                status=active_status,
                follow_up_date__date__gt=today,
                follow_up_date__date__lte=(
                    today + timedelta(days=7)
                ),
            ).count()
        ),

        "upcoming": (
            ContactLensFollowUp.objects.filter(
                status=active_status,
                follow_up_date__date__gt=(
                    today + timedelta(days=7)
                ),
            ).count()
        ),

        "missed": (
            ContactLensFollowUp.objects.filter(
                status=(
                    ContactLensFollowUp
                    .FollowUpStatus
                    .MISSED
                )
            ).count()
        ),

        "complications": (
            ContactLensFollowUp.objects.filter(
                Q(
                    complications__isnull=False
                )
                & ~Q(
                    complications=""
                )
            ).count()
        ),
    }

    expiring_prescriptions = (
        ContactLensPrescription.objects
        .filter(
            is_active=True,
            status__in=[
                ContactLensPrescription
                .PrescriptionStatus
                .APPROVED,

                ContactLensPrescription
                .PrescriptionStatus
                .DISPENSED,
            ],
            valid_until__isnull=False,
            valid_until__gte=today,
            valid_until__lte=(
                today + timedelta(days=30)
            ),
        )
        .select_related(
            "patient",
            "visit",
            "approved_by",
        )
        .order_by(
            "valid_until",
            "pk",
        )[:10]
    )

    expired_prescription_count = (
        ContactLensPrescription.objects
        .filter(
            is_active=True,
            status__in=[
                ContactLensPrescription
                .PrescriptionStatus
                .APPROVED,

                ContactLensPrescription
                .PrescriptionStatus
                .DISPENSED,
            ],
            valid_until__lt=today,
        )
        .count()
    )

    return render(
        request,
        (
            "patients/contact_lens/"
            "follow_up_queue.html"
        ),
        {
            "follow_up_rows": follow_up_rows,
            "page_obj": page_obj,
            "selected_queue": selected_queue,
            "selected_status": selected_status,
            "status_choices": (
                ContactLensFollowUp
                .FollowUpStatus
                .choices
            ),
            "query": query,
            "today": today,
            "queue_counts": queue_counts,
            "expiring_prescriptions": (
                expiring_prescriptions
            ),
            "expired_prescription_count": (
                expired_prescription_count
            ),
            "can_manage_contact_lenses": (
                user_can_manage_contact_lenses(
                    request.user
                )
            ),
        },
    )


@login_required
@clinical_staff_required
def patient_contact_lens_history(
    request,
    patient_pk,
):
    patient = get_object_or_404(
        Patient,
        pk=patient_pk,
    )

    assessments = (
        ContactLensAssessment.objects
        .filter(patient=patient)
        .select_related(
            "visit",
            "eye_examination",
            "assessed_by",
        )
        .prefetch_related(
            "trial_lenses",
            "prescriptions",
        )
        .order_by(
            "-assessment_date",
            "-pk",
        )
    )

    prescriptions = (
        ContactLensPrescription.objects
        .filter(
            patient=patient,
            is_active=True,
        )
        .select_related(
            "assessment",
            "visit",
            "prescribed_by",
            "approved_by",
            "dispensed_by",
        )
        .order_by(
            "-prescription_date",
            "-version",
            "-pk",
        )
    )

    follow_ups = (
        ContactLensFollowUp.objects
        .filter(patient=patient)
        .select_related(
            "prescription",
            "reviewed_by",
        )
        .order_by(
            "-follow_up_date",
            "-pk",
        )
    )

    selected_status = (
        request.GET.get("status", "")
        .strip()
        .upper()
    )

    valid_statuses = {
        value
        for value, _label
        in ContactLensPrescription
        .PrescriptionStatus
        .choices
    }

    if selected_status in valid_statuses:
        prescriptions = prescriptions.filter(
            status=selected_status
        )
    else:
        selected_status = ""

    query = (
        request.GET.get("q", "")
        .strip()
    )

    if query:
        prescriptions = prescriptions.filter(
            Q(
                prescription_number__icontains=query
            )
            | Q(
                right_brand_name__icontains=query
            )
            | Q(
                left_brand_name__icontains=query
            )
            | Q(
                visit__visit_number__icontains=query
            )
        )

    paginator = Paginator(
        prescriptions,
        20,
    )

    page_obj = paginator.get_page(
        request.GET.get("page")
    )

    prescription_counts = {
        "all": (
            ContactLensPrescription.objects
            .filter(
                patient=patient,
                is_active=True,
            )
            .count()
        ),
        "draft": (
            ContactLensPrescription.objects
            .filter(
                patient=patient,
                is_active=True,
                status=(
                    ContactLensPrescription
                    .PrescriptionStatus
                    .DRAFT
                ),
            )
            .count()
        ),
        "pending": (
            ContactLensPrescription.objects
            .filter(
                patient=patient,
                is_active=True,
                status=(
                    ContactLensPrescription
                    .PrescriptionStatus
                    .PENDING_APPROVAL
                ),
            )
            .count()
        ),
        "approved": (
            ContactLensPrescription.objects
            .filter(
                patient=patient,
                is_active=True,
                status=(
                    ContactLensPrescription
                    .PrescriptionStatus
                    .APPROVED
                ),
            )
            .count()
        ),
        "dispensed": (
            ContactLensPrescription.objects
            .filter(
                patient=patient,
                is_active=True,
                status=(
                    ContactLensPrescription
                    .PrescriptionStatus
                    .DISPENSED
                ),
            )
            .count()
        ),
    }

    return render(
        request,
        "patients/contact_lens/patient_history.html",
        {
            "patient": patient,
            "assessments": assessments,
            "prescriptions": page_obj,
            "page_obj": page_obj,
            "follow_ups": follow_ups[:10],
            "prescription_counts": (
                prescription_counts
            ),
            "selected_status": selected_status,
            "query": query,
            "status_choices": (
                ContactLensPrescription
                .PrescriptionStatus
                .choices
            ),
            "can_manage_contact_lenses": (
                user_can_manage_contact_lenses(
                    request.user
                )
            ),
        },
    )


# ============================================================
# CONTACT LENS PRESCRIPTION PRINT / PDF EXPORT
# ============================================================


@login_required
@clinical_staff_required
def contact_lens_prescription_print(
    request,
    pk,
):
    """
    Display a professional print-ready Contact Lens prescription.

    Draft and pending prescriptions may be previewed but are clearly
    watermarked and cannot be mistaken for approved final prescriptions.
    """

    ensure_contact_lens_print_permission(
        request.user
    )

    prescription = get_object_or_404(
        get_contact_lens_prescription_queryset(
            include_inactive=True
        ),
        pk=pk,
    )

    if not prescription.is_active:
        messages.warning(
            request,
            (
                "This prescription is inactive. It is being shown "
                "for historical reference only."
            ),
        )

    log_activity(
        request,
        AuditLog.ActionType.PRINT,
        "Contact Lens",
        (
            "Opened printable Contact Lens prescription "
            f"{prescription.prescription_number}, "
            f"version {prescription.version}, for patient "
            f"{prescription.patient.file_number}."
        ),
        object_id=prescription.pk,
        object_repr=str(prescription),
    )

    context = contact_lens_prescription_print_context(
        prescription,
        request=request,
    )

    return render(
        request,
        (
            "patients/contact_lens/"
            "prescription_print.html"
        ),
        context,
    )


@login_required
@clinical_staff_required
def contact_lens_prescription_pdf(
    request,
    pk,
):
    """
    Export an approved or dispensed Contact Lens prescription to PDF.

    WeasyPrint is imported inside the view so a missing package does
    not break the entire application during startup.
    """

    ensure_contact_lens_print_permission(
        request.user
    )

    prescription = get_object_or_404(
        get_contact_lens_prescription_queryset(
            include_inactive=True
        ),
        pk=pk,
    )

    allowed_statuses = {
        ContactLensPrescription
        .PrescriptionStatus
        .APPROVED,

        ContactLensPrescription
        .PrescriptionStatus
        .DISPENSED,
    }

    if prescription.status not in allowed_statuses:
        messages.warning(
            request,
            (
                "Only an approved or dispensed Contact Lens "
                "prescription can be exported as a final PDF. "
                "Use Print Preview to review the current draft."
            ),
        )

        return redirect(
            "contact_lens_prescription_print",
            pk=prescription.pk,
        )

    try:
        from weasyprint import HTML
    except ImportError:
        messages.error(
            request,
            (
                "PDF generation is unavailable because WeasyPrint "
                "is not installed in the current environment. "
                "The browser Print command can still save the "
                "prescription as PDF."
            ),
        )

        return redirect(
            "contact_lens_prescription_print",
            pk=prescription.pk,
        )

    context = contact_lens_prescription_print_context(
        prescription,
        request=request,
    )

    html_string = render_to_string(
        (
            "patients/contact_lens/"
            "prescription_print.html"
        ),
        context,
        request=request,
    )

    try:
        pdf_bytes = HTML(
            string=html_string,
            base_url=request.build_absolute_uri("/"),
        ).write_pdf()

    except Exception as error:
        messages.error(
            request,
            (
                "The Contact Lens PDF could not be generated. "
                "Use the browser Print command while the PDF "
                "configuration is being checked."
            ),
        )

        log_activity(
            request,
            AuditLog.ActionType.UPDATE,
            "Contact Lens",
            (
                "Contact Lens PDF generation failed for "
                f"{prescription.prescription_number}: "
                f"{error.__class__.__name__}."
            ),
            object_id=prescription.pk,
            object_repr=str(prescription),
        )

        return redirect(
            "contact_lens_prescription_print",
            pk=prescription.pk,
        )

    filename = (
        f"Contact_Lens_Prescription_"
        f"{prescription.prescription_number}_"
        f"V{prescription.version}.pdf"
    )

    response = HttpResponse(
        pdf_bytes,
        content_type="application/pdf",
    )

    response["Content-Disposition"] = (
        f'inline; filename="{filename}"'
    )

    log_activity(
        request,
        AuditLog.ActionType.PRINT,
        "Contact Lens",
        (
            "Generated Contact Lens prescription PDF "
            f"{prescription.prescription_number}, "
            f"version {prescription.version}, for patient "
            f"{prescription.patient.file_number}."
        ),
        object_id=prescription.pk,
        object_repr=str(prescription),
    )

    return response
